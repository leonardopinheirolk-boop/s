import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import hashlib
import re
from datetime import datetime, timedelta
import os
import json
import secrets

# Carregar variáveis de ambiente
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv não instalado

# Importações opcionais para funcionalidades de email/WhatsApp
try:
    import yagmail
    EMAIL_AVAILABLE = True
except ImportError:
    EMAIL_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

# Importações para sistema de monitoramento
try:
    from firebase_config import firebase_manager
    from ip_utils import get_client_info
    from admin_page import tela_admin, dashboard_admin, relatorio_completo, estatisticas_usuario
    MONITORING_AVAILABLE = True
    
    # Inicializar Firebase se disponível
    try:
        firebase_manager.initialize()
    except Exception as e:
        print(f"Firebase não inicializado: {e}")
        MONITORING_AVAILABLE = False
except ImportError:
    MONITORING_AVAILABLE = False

# -----------------------------
# Sistema de Autenticação
# -----------------------------
def carregar_usuarios():
    """Carrega a planilha de usuários"""
    try:
        # Tenta carregar a planilha de login
        df_usuarios = pd.read_excel("login_senha.xlsx")
        return df_usuarios
    except FileNotFoundError:
        st.error("Arquivo 'login_senha.xlsx' não encontrado!")
        return None
    except Exception as e:
        st.error(f"Erro ao carregar usuários: {str(e)}")
        return None

def validar_cpf(cpf):
    """Valida formato do CPF"""
    cpf = re.sub(r'[^0-9]', '', str(cpf))
    if len(cpf) != 11:
        return False
    return True

def _has_recent_access(usuario_nome):
    """Verifica se o usuário já está logado (evita registros duplicados na mesma sessão)"""
    try:
        # Verificar se o usuário já está na sessão atual
        # Se já está logado, não registrar novo acesso
        if st.session_state.get('logado') and st.session_state.get('usuario', {}).get('nome') == usuario_nome:
            return True
        
        return False
        
    except Exception as e:
        print(f"Erro ao verificar acesso recente: {e}")
        return False

def autenticar_usuario(identificador, senha):
    """Autentica usuário com CPF ou INEP e senha"""
    df_usuarios = carregar_usuarios()
    if df_usuarios is None:
        return None
    
    # Normalizar identificador (remover pontos, traços, espaços) - aceita apenas números
    id_limpo = re.sub(r'[^0-9]', '', str(identificador))
    if not id_limpo:
        return None
    
    # Buscar usuário na planilha
    for _, usuario in df_usuarios.iterrows():
        # Verificar CPF - remover formatação e tratar como string para preservar zeros à esquerda
        cpf_valor = usuario.get('CPF', '')
        if pd.isna(cpf_valor) or cpf_valor == '':
            cpf_usuario = ''
        else:
            # Converter para string primeiro (importante para CPFs com zeros à esquerda)
            cpf_str = str(cpf_valor)
            # Remover formatação (pontos, traços, espaços)
            cpf_usuario = re.sub(r'[^0-9]', '', cpf_str)
        
        # Verificar INEP - tratar NaN e float
        inep_valor = usuario.get('INEP', '')
        if pd.isna(inep_valor) or inep_valor == '':
            inep_usuario = ''
        else:
            # Converter float para int primeiro para remover o .0, depois para string
            inep_str = str(int(float(inep_valor)))
            inep_usuario = re.sub(r'[^0-9]', '', inep_str)
        
        # Comparar com CPF ou INEP (comparação exata de strings)
        if (cpf_usuario and cpf_usuario == id_limpo) or (inep_usuario and inep_usuario == id_limpo):
            # Verificar senha (comparação direta)
            if str(usuario.get('SENHA', '')) == str(senha):
                # Registrar acesso apenas no momento do login
                if MONITORING_AVAILABLE:
                    try:
                        client_info = get_client_info()
                        # Sempre registrar o acesso (removida verificação de acesso recente)
                        firebase_manager.log_access(
                            usuario=usuario.get('NOME', 'Usuário'),
                            ip=client_info['ip'],
                            user_agent=client_info['user_agent']
                        )
                    except Exception as e:
                        print(f"Erro ao registrar acesso: {e}")
                
                return {
                    'nome': usuario.get('NOME', 'Usuário'),
                    'cpf': cpf_usuario if cpf_usuario else None,
                    'inep': inep_usuario if inep_usuario else None,
                    'senha_atual': str(usuario.get('SENHA', '')),
                    'linha': _
                }
    return None

def alterar_senha(identificador, senha_atual, nova_senha):
    """Altera a senha do usuário na planilha"""
    try:
        df_usuarios = carregar_usuarios()
        if df_usuarios is None:
            return False, "Erro ao carregar planilha"
        
        # Normalizar identificador (remover pontos, traços, espaços) - aceita apenas números
        id_limpo = re.sub(r'[^0-9]', '', str(identificador))
        if not id_limpo:
            return False, "Identificador inválido"
        
        # Encontrar usuário
        for idx, usuario in df_usuarios.iterrows():
            # Verificar CPF - remover formatação e tratar como string para preservar zeros à esquerda
            cpf_valor = usuario.get('CPF', '')
            if pd.isna(cpf_valor) or cpf_valor == '':
                cpf_usuario = ''
            else:
                # Converter para string primeiro (importante para CPFs com zeros à esquerda)
                cpf_str = str(cpf_valor)
                # Remover formatação (pontos, traços, espaços)
                cpf_usuario = re.sub(r'[^0-9]', '', cpf_str)
            # Verificar INEP - tratar NaN e float
            inep_valor = usuario.get('INEP', '')
            if pd.isna(inep_valor) or inep_valor == '':
                inep_usuario = ''
            else:
                # Converter float para int primeiro para remover o .0, depois para string
                inep_str = str(int(float(inep_valor)))
                inep_usuario = re.sub(r'[^0-9]', '', inep_str)
            
            # Comparar com CPF ou INEP (comparação exata de strings)
            if (cpf_usuario and cpf_usuario == id_limpo) or (inep_usuario and inep_usuario == id_limpo):
                if str(usuario.get('SENHA', '')) == str(senha_atual):
                    # Atualizar senha
                    df_usuarios.at[idx, 'SENHA'] = nova_senha
                    
                    # Salvar planilha
                    df_usuarios.to_excel("login_senha.xlsx", index=False)
                    return True, "Senha alterada com sucesso!"
                else:
                    return False, "Senha atual incorreta"
        
        return False, "Usuário não encontrado"
    except Exception as e:
        return False, f"Erro ao alterar senha: {str(e)}"


# -----------------------------
# Login por e-mail + código (envio do código para o e-mail)
# -----------------------------
CODIGOS_LOGIN_FILE = "codigos_login.json"
CODIGO_EXPIRA_MINUTOS = 10

# Sessão: expira após 30 minutos e exige novo login
SESSAO_EXPIRA_MINUTOS = 30


def _carregar_codigos():
    try:
        if os.path.exists(CODIGOS_LOGIN_FILE):
            with open(CODIGOS_LOGIN_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _salvar_codigos(data):
    try:
        with open(CODIGOS_LOGIN_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _limpar_expirados(data):
    agora = datetime.now()
    for email in list(data.keys()):
        expira_str = data[email].get('expira', '')
        try:
            expira = datetime.strptime(expira_str, '%Y-%m-%dT%H:%M:%S')
            if agora > expira:
                del data[email]
        except Exception:
            del data[email]
    return data

def buscar_usuario_por_email(email):
    """Busca usuário na planilha pela coluna EMAIL (ou E-mail / email)."""
    df_usuarios = carregar_usuarios()
    if df_usuarios is None:
        return None
    email_limpo = str(email).strip().lower() if email else ''
    if not email_limpo:
        return None
    for _, usuario in df_usuarios.iterrows():
        for col in ('EMAIL', 'E-mail', 'email'):
            val = usuario.get(col, '')
            if pd.notna(val) and str(val).strip() != '':
                if str(val).strip().lower() == email_limpo:
                    cpf_valor = usuario.get('CPF', '')
                    cpf_usuario = re.sub(r'[^0-9]', '', str(cpf_valor)) if pd.notna(cpf_valor) and cpf_valor != '' else ''
                    inep_valor = usuario.get('INEP', '')
                    inep_usuario = ''
                    if pd.notna(inep_valor) and inep_valor != '':
                        inep_usuario = re.sub(r'[^0-9]', '', str(int(float(inep_valor))))
                    return {
                        'nome': usuario.get('NOME', 'Usuário'),
                        'cpf': cpf_usuario or None,
                        'inep': inep_usuario or None,
                        'senha_atual': '',
                        'linha': _
                    }
    return None

def _gmail_esta_configurado():
    gmail_user = (os.getenv('GMAIL_USER') or '').strip()
    gmail_pass = (os.getenv('GMAIL_PASSWORD') or '').strip()
    return bool(gmail_user and gmail_pass and gmail_user != 'seu_email@gmail.com' and gmail_pass != 'sua_senha_app')

def gerar_e_enviar_codigo(email):
    """Gera código de 6 dígitos, salva e envia por e-mail. Retorna (ok, msg, email_enviado)."""
    usuario = buscar_usuario_por_email(email)
    if usuario is None:
        return False, "E-mail não cadastrado na planilha. Cadastre o e-mail na coluna EMAIL do arquivo login_senha.xlsx.", False
    codigo = ''.join(secrets.choice('0123456789') for _ in range(6))
    expira = datetime.now() + timedelta(minutes=CODIGO_EXPIRA_MINUTOS)
    expira_str = expira.strftime('%Y-%m-%dT%H:%M:%S')
    data = _limpar_expirados(_carregar_codigos())
    data[str(email).strip().lower()] = {'codigo': codigo, 'expira': expira_str}
    _salvar_codigos(data)

    if not _gmail_esta_configurado():
        return True, (
            "Código gerado, mas **o e-mail não foi enviado** porque o envio não está configurado. "
            "No **Streamlit Cloud**: Settings → Secrets → adicione **GMAIL_USER** e **GMAIL_PASSWORD** (senha de app do Gmail). "
            "No **computador**: crie o arquivo **.env** com as mesmas variáveis. Depois solicite o código novamente."
        ), False

    assunto = "Seu código de acesso - Painel SGE"
    corpo = f"""Olá, {usuario['nome']}!

Seu código de acesso ao Painel SGE é:

    {codigo}

Válido por {CODIGO_EXPIRA_MINUTOS} minutos. Não compartilhe.

Se não solicitou, ignore este e-mail.
"""
    ok, msg = enviar_email(destinatario=email, assunto=assunto, corpo=corpo, remetente_log="Sistema")
    if ok:
        return True, f"Código enviado para {email}. Verifique a caixa de entrada (e spam).", True
    return False, msg, False

def validar_codigo_email(email, codigo):
    codigo_digitado = str(codigo).strip().replace(' ', '')
    if not codigo_digitado:
        return None
    data = _limpar_expirados(_carregar_codigos())
    email_key = str(email).strip().lower()
    if email_key not in data or data[email_key].get('codigo') != codigo_digitado:
        return None
    del data[email_key]
    _salvar_codigos(data)
    return buscar_usuario_por_email(email)


def tela_instrucoes():
    """Exibe tela de instruções de uso do sistema"""
    
    # Header moderno
    st.markdown("""
    <div style="background: linear-gradient(135deg, #4a90e2 0%, #357abd 100%); 
                color: white; padding: 2rem; border-radius: 15px; 
                text-align: center; margin-bottom: 2rem; 
                box-shadow: 0 8px 25px rgba(74, 144, 226, 0.15);">
        <h1 style="margin: 0; font-size: 2.5rem; font-weight: 700; 
                   text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
            Guia Completo de Uso
        </h1>
        <p style="font-size: 1.2rem; margin: 0.5rem 0 0 0; opacity: 0.9;">
            Como analisar os dados da sua escola de forma eficiente
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Botão para voltar
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("Voltar ao Login", use_container_width=True, type="primary", key="btn_voltar_login"):
            st.session_state.mostrar_instrucoes = False
            st.rerun()
    
    st.markdown("---")
    
    # PASSO 1
    st.markdown("### 📥 PASSO 1: Baixar Dados do SGE")
    
    st.markdown("""
    **1.1 - Acesse o SGE:**
    - Entre no sistema SGE da sua escola
    - Faça login com suas credenciais
    
    **1.2 - Navegue até Relatórios:**
    - No menu lateral, clique em "Relatórios"
    - Selecione "Escola"
    - Escolha "Ata/Mapa de rendimento"
    
    **1.3 - Configure o Relatório:**
    - **Escola:** Selecione sua escola
    - **Modelo:** Escolha "Conselho"
    - **Tipo:** Selecione "Anual"
    - **Ano Letivo:** Escolha o ano atual
    - **1º Semestre:** Selecione o semestre desejado
    - **Status:** Deixe em branco para incluir todos
    
    **1.4 - Baixe a Planilha:**
    - Clique no botão "Exportar"
    - Escolha "Excel"
    - Salve o arquivo no seu computador
    """)
    
    # PASSO 2
    st.markdown("### 📤 PASSO 2: Carregar Dados no Painel")
    
    st.markdown("""
    **2.1 - Faça Login:**
    - Use seu CPF ou INEP da escola
    - Digite sua senha
    - Clique em "Entrar"
    
    **2.2 - Carregue a Planilha:**
    - Na tela principal, clique em "Escolher arquivo"
    - Selecione a planilha baixada do SGE
    - Aguarde o carregamento dos dados
    """)
    
    # PASSO 3
    st.markdown("### ⚙️ PASSO 3: Configurar Filtros Obrigatórios")
    
    st.warning("⚠️ **IMPORTANTE: Filtros Obrigatórios** - Estes filtros são essenciais para análise correta!")
    
    st.markdown("""
    **3.1 - Filtro de Escola:**
    - No menu lateral, selecione sua escola
    - Este filtro é obrigatório
    
    **3.2 - Filtro de Status:**
    - Sempre selecione "Cursando"
    - Este filtro é obrigatório para análise correta
    - Desmarque outros status se aparecerem
    """)
    
    # PASSO 4
    st.markdown("### 🔍 PASSO 4: Filtros Opcionais")
    
    st.markdown("""
    **4.1 - Por Disciplina:**
    - Selecione disciplinas específicas
    - Útil para análise de matérias problemáticas
    
    **4.2 - Por Turma:**
    - Escolha turmas específicas
    - Ideal para análise de classes individuais
    
    **4.3 - Por Aluno:**
    - Selecione alunos específicos
    - Para acompanhamento individual
    """)
    
    # PASSO 5
    st.markdown("### 📊 PASSO 5: Analisar os Dados")
    
    st.markdown("""
    **5.1 - Visão Geral:**
    - Veja o resumo geral da escola
    - Métricas de aprovação e reprovação
    - Indicadores de frequência
    
    **5.2 - Análise por Bimestre:**
    - Compare 1º e 2º bimestres
    - Identifique tendências de melhoria ou piora
    
    **5.3 - Alertas e Riscos:**
    - Alunos em situação de risco
    - Necessidade de intervenção pedagógica
    
    **5.4 - Gráficos e Visualizações:**
    - Gráficos de notas por disciplina
    - Análise de frequência
    - Comparações entre turmas
    """)
    
    # Dicas importantes
    st.markdown("### 💡 Dicas Importantes")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **Dados Obrigatórios:**
        - Escola e Status "Cursando" são sempre necessários
        - Sem esses filtros, a análise pode ficar incorreta
        """)
    
    with col2:
        st.markdown("""
        **Interpretação dos Resultados:**
        - Notas abaixo de 6 indicam necessidade de atenção
        - Frequência abaixo de 75% é preocupante
        - Alunos em "Corda Bamba" precisam de acompanhamento
        """)
    
    # Problemas comuns
    st.markdown("### ❓ Problemas Comuns")
    
    st.markdown("""
    **Erro ao carregar planilha:**
    - Verifique se o arquivo é do SGE
    - Confirme se tem as colunas necessárias
    - Tente salvar novamente no SGE
    
    **Dados não aparecem:**
    - Verifique os filtros obrigatórios
    - Confirme se selecionou "Cursando"
    - Verifique se a escola está correta
    
    **Gráficos não carregam:**
    - Aguarde o processamento dos dados
    - Verifique se há dados suficientes
    - Tente recarregar a página
    """)
    
    # Assinatura
    st.markdown("---")
    st.markdown("<div style='text-align: center; padding: 2rem;'><strong style='color: #4a90e2; font-size: 1.1rem;'>© 2025 – desenvolvido por Alexandre Tolentino</strong></div>", unsafe_allow_html=True)

def tela_login():
    """Login apenas por e-mail + código enviado ao e-mail. Sem CPF/INEP."""
    st.markdown("""
    <style>
    .stButton > button[kind="primary"] { background-color: #1f77b4; color: white; border: none; border-radius: 0.5rem; padding: 0.75rem 1.5rem; font-size: 1.1rem; font-weight: 600; min-height: 3rem; }
    .stButton > button[kind="primary"]:hover { background-color: #0d5a8a; color: white; }
    </style>
    """, unsafe_allow_html=True)
    
    col_inst, col_main, col_empty = st.columns([1, 2, 1])
    with col_inst:
        if st.button("Instruções", use_container_width=True, help="Como usar o sistema", type="primary", key="btn_instrucoes"):
            st.session_state.mostrar_instrucoes = True
            st.rerun()
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### Acesso ao Painel SGE")
        if st.session_state.get("sessao_expirada"):
            st.warning(f"Sua sessão expirou após {SESSAO_EXPIRA_MINUTOS} minutos. Faça login novamente.")
            st.session_state.sessao_expirada = False

        email_aguardando = st.session_state.get("email_aguardando_codigo") or ""
        if not email_aguardando:
            st.info("Informe seu e-mail. Um código de acesso será enviado para esse e-mail.")
            with st.form("login_form_email"):
                email_digitado = st.text_input("E-mail:", placeholder="seu_email@exemplo.com", key="email_login")
                enviar_btn = st.form_submit_button("Enviar código por e-mail")
            if enviar_btn:
                email = (email_digitado or "").strip()
                if not email or "@" not in email:
                    st.error("Digite um e-mail válido.")
                else:
                    ok, msg, email_enviado = gerar_e_enviar_codigo(email)
                    if ok:
                        if email_enviado:
                            st.session_state.email_aguardando_codigo = email
                            st.success(msg)
                            st.rerun()
                        else:
                            st.warning(msg)
                    else:
                        st.error(msg)
        else:
            st.success(f"Código enviado para **{email_aguardando}**. Digite o código recebido abaixo.")
            if st.button("Usar outro e-mail", key="btn_outro_email"):
                st.session_state.email_aguardando_codigo = ""
                st.rerun()
            with st.form("login_form_codigo"):
                codigo_digitado = st.text_input("Código recebido:", placeholder="Ex: 123456", max_chars=6, key="codigo_login")
                entrar_btn = st.form_submit_button("Entrar")
            if entrar_btn:
                usuario = validar_codigo_email(email_aguardando, codigo_digitado)
                if usuario:
                    if MONITORING_AVAILABLE:
                        try:
                            client_info = get_client_info()
                            firebase_manager.log_access(usuario=usuario.get('nome', 'Usuário'), ip=client_info['ip'], user_agent=client_info['user_agent'])
                        except Exception as e:
                            print(f"Erro ao registrar acesso: {e}")
                    st.session_state.logado = True
                    st.session_state.usuario = usuario
                    st.session_state.login_time = datetime.now()
                    st.session_state.email_aguardando_codigo = ""
                    st.success("Login realizado com sucesso!")
                    st.rerun()
                else:
                    st.error("Código inválido ou expirado. Solicite um novo código.")

        st.markdown("---")
        st.markdown("<div style='text-align: center;'><strong>© 2025 – desenvolvido por Alexandre Tolentino</strong></div>", unsafe_allow_html=True)

def tela_sobre():
    """Exibe modal com informações sobre o sistema"""
    st.title("História do Painel SGE")
    st.markdown("---")
    
    # CSS para aumentar fonte e justificar texto
    st.markdown("""
    <style>
    .sobre-texto {
        font-size: 16px;
        line-height: 1.8;
        text-align: justify;
        margin-bottom: 20px;
    }
    .sobre-titulo {
        font-size: 18px;
        font-weight: bold;
        margin-bottom: 10px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-titulo">Desenvolvedor: Alexandre Tolentino</div>
    <div class="sobre-titulo">Cargo: Técnico de Currículo da Superintendência Regional de Educação de Gurupi - TO</div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    Este painel foi desenvolvido para atender uma necessidade crítica identificada no processo de análise de dados educacionais da Superintendência Regional de Educação de Gurupi. O Sistema de Gestão Escolar (SGE) fornecia uma grande quantidade de informações sobre o desempenho dos alunos, frequência e indicadores educacionais, porém o processo de análise e interpretação desses dados era extremamente trabalhoso e demorado.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    O trabalho manual de coleta, organização e análise dos dados das escolas levava horas para ser concluído, comprometendo significativamente a eficiência do processo de acompanhamento pedagógico. Quando os técnicos chegavam às escolas para apresentar os resultados e discutir estratégias de intervenção para melhoria do processo de ensino-aprendizagem, muitas vezes não dispunham de tempo suficiente para uma análise aprofundada e contextualizada dos dados.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    Diante dessa realidade, Alexandre Tolentino, reconhecendo a necessidade de uma ferramenta mais eficiente e ágil, desenvolveu este painel interativo. A solução permite que, com a simples inserção da planilha de dados do SGE, em questão de segundos sejam apresentados todos os indicadores da unidade escolar de forma organizada, visual e facilmente interpretável.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    No dia seguinte à sua criação, o painel foi apresentado à equipe técnica da Superintendência Regional de Educação de Gurupi, sendo imediatamente aprovado por todos os membros da equipe. A ferramenta foi incorporada ao processo de trabalho de forma instantânea, permitindo que os técnicos levassem às escolas análises mais precisas e discussões mais produtivas sobre estratégias de melhoria dos índices educacionais.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    Posteriormente, o próprio Sistema de Gestão Escolar desenvolveu um mapa de acompanhamento que oferecia funcionalidades similares. No entanto, a Superintendência Regional de Educação de Gurupi optou por continuar utilizando este painel personalizado, pois ele apresentava características específicas e funcionalidades particulares que o mapa oficial do SGE não possuía, atendendo melhor às necessidades específicas da região.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    O sucesso da ferramenta despertou o interesse das próprias escolas, que passaram a solicitar acesso ao painel para realizar seus próprios levantamentos e análises internas. Diante dessa demanda, a ferramenta foi disponibilizada para que cada unidade escolar pudesse fazer o monitoramento de seus próprios indicadores e implementar intervenções direcionadas e contextualizadas. Esta iniciativa tem se mostrado extremamente eficaz, contribuindo significativamente para a melhoria da qualidade educacional na região.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    Desde sua implementação inicial até os dias atuais, o painel tem passado por constantes atualizações e melhorias, sempre baseadas nas necessidades reais identificadas pelos usuários. Alexandre Tolentino continua desenvolvendo e aprimorando o sistema conforme novas demandas surgem, garantindo que a ferramenta permaneça atual, útil e eficaz para o processo de gestão educacional.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="sobre-texto">
    Este trabalho representa um exemplo concreto de como a tecnologia, quando aplicada de forma inteligente e direcionada às necessidades reais do sistema educacional, pode transformar dados brutos em informações estratégicas que contribuem efetivamente para a melhoria da qualidade do ensino e do aprendizado.
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    st.info("**Missão:** Transformar dados em ações educacionais eficazes")
    st.markdown("---")
    
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
    with col_btn2:
        if st.button("Fechar", use_container_width=True, type="primary", key="btn_fechar_sobre"):
            st.session_state.mostrar_sobre = False
            st.rerun()

def tela_alterar_senha():
    """Exibe tela para alterar senha"""
    st.title("🔑 Alterar Senha")
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("### Alterar sua senha")
        
        with st.form("alterar_senha_form"):
            senha_atual = st.text_input("Senha atual:", type="password")
            nova_senha = st.text_input("Nova senha:", type="password")
            confirmar_senha = st.text_input("Confirmar nova senha:", type="password")
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                alterar_btn = st.form_submit_button("Alterar Senha", use_container_width=True)
            with col_btn2:
                if st.form_submit_button("Cancelar", use_container_width=True):
                    st.session_state.mostrar_alterar_senha = False
                    st.rerun()
        
        if alterar_btn:
            if not senha_atual or not nova_senha or not confirmar_senha:
                st.error("Por favor, preencha todos os campos!")
            elif nova_senha != confirmar_senha:
                st.error("As senhas não coincidem!")
            elif len(nova_senha) < 4:
                st.error("A nova senha deve ter pelo menos 4 caracteres!")
            else:
                # Usar CPF ou INEP dependendo do que estiver disponível
                identificador = st.session_state.usuario.get('cpf') or st.session_state.usuario.get('inep')
                sucesso, mensagem = alterar_senha(
                    identificador, 
                    senha_atual, 
                    nova_senha
                )
                if sucesso:
                    st.success(mensagem)
                    st.session_state.mostrar_alterar_senha = False
                    st.rerun()
                else:
                    st.error(mensagem)

# -----------------------------
# Sistema de Relatórios e Envio
# -----------------------------
def gerar_relatorio_excel(df, tipo_relatorio="completo", filtros=None):
    """Gera relatório em Excel com os dados filtrados"""
    try:
        # Criar novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório SGE"
        
        # Adicionar cabeçalho
        ws['A1'] = "RELATÓRIO SGE - SISTEMA DE GESTÃO ESCOLAR"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=12)
        ws['A3'] = f"Usuário: {st.session_state.usuario['nome']}"
        ws['A3'].font = Font(size=12)
        
        # Adicionar dados
        if not df.empty:
            # Converter DataFrame para Excel
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formatar cabeçalho
            for cell in ws[4]:  # Linha do cabeçalho
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    except Exception as e:
        st.error(f"Erro ao gerar relatório: {str(e)}")
        return None

def carregar_config_email():
    """Carrega configurações de email do arquivo"""
    try:
        with open('email_config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except:
        return {"email_remetente": "", "senha_email": "", "configurado": False}

def salvar_config_email(email, senha):
    """Salva configurações de email no arquivo"""
    try:
        config = {
            "email_remetente": email,
            "senha_email": senha,
            "configurado": True
        }
        with open('email_config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except:
        return False

def enviar_email(destinatario, assunto, corpo, anexo=None, remetente_log=None):
    """Envia email real com Gmail SMTP. remetente_log: nome no log (se None, usa usuário logado ou 'Sistema')."""
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        import os
        
        gmail_user = os.getenv('GMAIL_USER', 'seu_email@gmail.com')
        gmail_password = os.getenv('GMAIL_PASSWORD', 'sua_senha_app')
        
        if gmail_user == 'seu_email@gmail.com' or not gmail_password:
            return enviar_email_simulado(destinatario, assunto, corpo, anexo, remetente_log)
        
        msg = MIMEMultipart()
        msg['From'] = gmail_user
        msg['To'] = destinatario
        msg['Subject'] = assunto
        msg.attach(MIMEText(corpo, 'plain', 'utf-8'))
        
        if anexo:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(anexo)
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= relatorio_sge_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
            )
            msg.attach(part)
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, destinatario, msg.as_string())
        server.quit()
        
        remetente = remetente_log if remetente_log is not None else (st.session_state.get('usuario') or {}).get('nome', 'Sistema')
        log_info = {
            "destinatario": destinatario,
            "assunto": assunto,
            "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "remetente": remetente,
            "status": "Enviado (Real)"
        }
        try:
            with open("email_log.json", "a", encoding="utf-8") as f:
                f.write(f"{json.dumps(log_info, ensure_ascii=False)}\n")
        except Exception:
            pass
        return True, f"Email enviado para {destinatario} com sucesso!"
    except Exception as e:
        return False, f"Erro ao enviar email: {str(e)}"

def enviar_email_simulado(destinatario, assunto, corpo, anexo=None, remetente_log=None):
    """Simula envio de email (fallback). remetente_log: nome no log (se None, usa usuário ou 'Sistema')."""
    try:
        import time
        time.sleep(1)
        remetente = remetente_log if remetente_log is not None else (st.session_state.get('usuario') or {}).get('nome', 'Sistema')
        log_info = {
            "destinatario": destinatario,
            "assunto": assunto,
            "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "remetente": remetente,
            "status": "Enviado (Simulado)"
        }
        try:
            with open("email_log.json", "a", encoding="utf-8") as f:
                f.write(f"{json.dumps(log_info, ensure_ascii=False)}\n")
        except Exception:
            pass
        return True, f"Email simulado enviado para {destinatario} com sucesso!"
    except Exception as e:
        return False, f"Erro ao simular envio: {str(e)}"



# -----------------------------
# Configuração inicial
# -----------------------------
st.set_page_config(page_title="Painel SGE – Notas e Alertas", layout="wide")

st.markdown("""
<style>
    .stApp {
        background: linear-gradient(180deg, #f8fafc 0%, #eef4ff 100%);
    }
    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1400px;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
        border-right: 1px solid rgba(255,255,255,0.08);
    }
    [data-testid="stSidebar"] * {
        color: #e5eefc;
    }
    div[data-testid="stMetric"] {
        background: rgba(255,255,255,0.92);
        border: 1px solid rgba(59,130,246,0.12);
        padding: 14px 16px;
        border-radius: 18px;
        box-shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
    }
    .stButton > button {
        border-radius: 14px;
        border: none;
        min-height: 2.8rem;
        font-weight: 700;
        box-shadow: 0 8px 20px rgba(15, 23, 42, 0.12);
    }
    .stSelectbox > div > div,
    .stMultiSelect > div > div,
    .stTextInput > div > div,
    .stDateInput > div > div,
    .stFileUploader > div {
        border-radius: 14px !important;
    }
    .stDataFrame, .stPlotlyChart {
        background: rgba(255,255,255,0.95);
        border-radius: 18px;
        box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
        padding: 0.2rem;
    }
</style>
""", unsafe_allow_html=True)

MEDIA_APROVACAO = 6.0
MEDIA_FINAL_ALVO = 6.0   # média final desejada após 4 bimestres
SOMA_FINAL_ALVO = MEDIA_FINAL_ALVO * 4  # 24 pontos no ano

# -----------------------------
# Utilidades
# -----------------------------
def detectar_tipo_planilha(df):
    """
    Detecta automaticamente o tipo de planilha baseado nas colunas disponíveis
    Retorna: 'notas_frequencia', 'conteudo_aplicado' ou 'censo_escolar'
    """
    colunas = [col.lower().strip() for col in df.columns]

    # Verificar se é planilha de censo escolar
    censo_indicators = [
        'código', 'superv', 'convên', 'entidade', 'inep', 'situação', 'classific',
        'nome', 'endereço', 'bairro', 'distrito', 'cep', 'cnpj', 'telefone', 'email',
        'nível de', 'categoria', 'tipo de estrutura', 'etapas', 'ano letivo', 'calendário',
        'curso', 'avaliação', 'conceito', 'servidor', 'turno', 'horário', 'tempo',
        'média', 'salário', 'língua', 'professor', 'área de cargo', 'data na', 'cpf'
    ]

    # Verificar se é planilha de conteúdo aplicado
    conteudo_indicators = [
        'componente curricu', 'atividade/conteúdo', 'situação', 'data', 'horário'
    ]

    # Verificar se é planilha de notas/frequência
    notas_indicators = [
        'aluno', 'nota', 'frequencia', 'turma', 'escola', 'disciplina', 'periodo'
    ]

    censo_score = sum(1 for indicator in censo_indicators
                      if any(indicator in col for col in colunas))
    conteudo_score = sum(1 for indicator in conteudo_indicators
                         if any(indicator in col for col in colunas))
    notas_score = sum(1 for indicator in notas_indicators
                      if any(indicator in col for col in colunas))

    # Se tem mais indicadores de censo escolar, é esse tipo
    if censo_score >= 8:
        return 'censo_escolar'
    elif conteudo_score >= 3:
        return 'conteudo_aplicado'
    elif notas_score >= 3:
        return 'notas_frequencia'
    else:
        # Se não conseguir detectar claramente, assume notas/frequência como padrão
        return 'notas_frequencia'

@st.cache_data(show_spinner=False)
def carregar_dados(arquivo, sheet=None):
    if arquivo is None:
        # Tenta ler o padrão local "dados.xlsx"
        df = pd.read_excel("dados.xlsx", sheet_name=sheet) if sheet else pd.read_excel("dados.xlsx")
    else:
        df = pd.read_excel(arquivo, sheet_name=sheet) if sheet else pd.read_excel(arquivo)

    # Normalizar nomes de colunas
    df.columns = [c.strip() for c in df.columns]
    
    # Detectar tipo de planilha
    tipo_planilha = detectar_tipo_planilha(df)
    
    if tipo_planilha == 'conteudo_aplicado':
        # Processar planilha de conteúdo aplicado
        return processar_conteudo_aplicado(df)
    elif tipo_planilha == 'censo_escolar':
        # Processar planilha do censo escolar
        return processar_censo_escolar(df)
    else:
        # Processar planilha de notas/frequência (padrão atual)
        return processar_notas_frequencia(df)

def processar_conteudo_aplicado(df):
    """Processa planilha de conteúdo aplicado"""
    # Mapear colunas para nomes padronizados
    mapeamento_colunas = {}
    
    for col in df.columns:
        col_lower = col.lower().strip()
        if 'componente curricu' in col_lower:
            mapeamento_colunas[col] = 'Disciplina'
        elif 'atividade/conteúdo' in col_lower or 'atividade' in col_lower:
            mapeamento_colunas[col] = 'Atividade'
        elif 'situação' in col_lower:
            mapeamento_colunas[col] = 'Status'
        elif 'data' in col_lower:
            mapeamento_colunas[col] = 'Data'
        elif 'horário' in col_lower:
            mapeamento_colunas[col] = 'Horario'
    
    df = df.rename(columns=mapeamento_colunas)
    
    # Converter Data para datetime se possível
    if 'Data' in df.columns:
        # Tentar diferentes formatos de data
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        # Se não funcionar, tentar formato automático
        if df['Data'].isna().all():
            df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
    
    # Padronizar texto dos campos principais
    for col in ['Disciplina', 'Atividade', 'Status']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    # Adicionar tipo de planilha para identificação
    df.attrs['tipo_planilha'] = 'conteudo_aplicado'
    
    return df

def processar_notas_frequencia(df):
    """Processa planilha de notas/frequência (processamento atual)"""
    # Garantir colunas esperadas (flexível aos nomes encontrados)
    # Esperados: Escola, Turma, Turno, Aluno, Periodo, Disciplina, Nota, Falta, Frequência, Frequência Anual
    # Algumas planilhas têm "Período" com acento; vamos padronizar para "Periodo"
    if "Período" in df.columns and "Periodo" not in df.columns:
        df = df.rename(columns={"Período": "Periodo"})
    if "Frequência" in df.columns and "Frequencia" not in df.columns:
        df = df.rename(columns={"Frequência": "Frequencia"})
    if "Frequência Anual" in df.columns and "Frequencia Anual" not in df.columns:
        df = df.rename(columns={"Frequência Anual": "Frequencia Anual"})

    # Converter Nota (vírgula -> ponto, texto -> float)
    if "Nota" in df.columns:
        df["Nota"] = (
            df["Nota"]
            .astype(str)
            .str.replace(",", ".", regex=False)
            .str.replace(" ", "", regex=False)
        )
        df["Nota"] = pd.to_numeric(df["Nota"], errors="coerce")

    # Falta -> numérico
    if "Falta" in df.columns:
        df["Falta"] = pd.to_numeric(df["Falta"], errors="coerce").fillna(0).astype(int)

    # Frequências -> numérico
    if "Frequencia" in df.columns:
        df["Frequencia"] = pd.to_numeric(df["Frequencia"], errors="coerce")
    if "Frequencia Anual" in df.columns:
        df["Frequencia Anual"] = pd.to_numeric(df["Frequencia Anual"], errors="coerce")

    # Padronizar texto dos campos principais (evita diferenças por espaços)
    for col in ["Escola", "Turma", "Turno", "Status", "Periodo", "Disciplina"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    # Detectar coluna de aluno/estudante
    coluna_aluno = None
    for col in ["Aluno", "Nome_Estudante", "Estudante"]:
        if col in df.columns:
            coluna_aluno = col
            break
    
    if coluna_aluno:
        df[coluna_aluno] = df[coluna_aluno].astype(str).str.strip()
    
    # Adicionar tipo de planilha para identificação
    df.attrs['tipo_planilha'] = 'notas_frequencia'
    
    return df

def processar_censo_escolar(df):
    """
    Processa dados do Censo Escolar - Lista de Estudantes
    """
    # Normalizar nomes das colunas
    df.columns = df.columns.str.strip()
    
    # Mapear colunas específicas da planilha ListaDeEstudantes_TurmaEscolarização
    colunas_mapeadas = {}
    for col in df.columns:
        col_lower = col.lower()
        if col == 'Nome':
            colunas_mapeadas[col] = 'Nome_Estudante'
        elif col == 'Escola':
            colunas_mapeadas[col] = 'Escola'
        elif col == 'CPF':
            colunas_mapeadas[col] = 'CPF'
        elif col == 'INEP':
            colunas_mapeadas[col] = 'Codigo_Estudante'
        elif col == 'Situação da Matrícula':
            colunas_mapeadas[col] = 'Situacao'
        elif col == 'Turno':
            colunas_mapeadas[col] = 'Turno'
        elif col == 'Data Nascimento':
            colunas_mapeadas[col] = 'Data_Nascimento'
        elif col == 'Nível de Ensino':
            colunas_mapeadas[col] = 'Nivel_Educacao'
        elif col == 'Ano/Série':
            colunas_mapeadas[col] = 'Ano_Serie'
        elif col == 'Descrição Turma':
            colunas_mapeadas[col] = 'Turma'
        elif col == 'Entidade Conveniada':
            colunas_mapeadas[col] = 'Entidade'
        elif col == 'Superintendência Regional':
            colunas_mapeadas[col] = 'Supervisao'
        elif col == 'Convênio':
            colunas_mapeadas[col] = 'Convenio'
        elif col == 'INEP da Escola':
            colunas_mapeadas[col] = 'INEP_Escola'
        elif col == 'Classificação da Escola':
            colunas_mapeadas[col] = 'Classificacao'
        elif col == 'Endereço':
            colunas_mapeadas[col] = 'Endereco'
        elif col == 'Bairro':
            colunas_mapeadas[col] = 'Bairro'
        elif col == 'Distrito':
            colunas_mapeadas[col] = 'Distrito'
        elif col == 'Cep':
            colunas_mapeadas[col] = 'CEP'
        elif col == 'Telefone Principal':
            colunas_mapeadas[col] = 'Telefone'
        elif col == 'E-mail':
            colunas_mapeadas[col] = 'Email'
        elif col == 'CNPJ':
            colunas_mapeadas[col] = 'CNPJ'
        elif col == 'Carga Horária':
            colunas_mapeadas[col] = 'Carga_Horaria'
        elif col == 'Entrada':
            colunas_mapeadas[col] = 'Data_Entrada'
        elif col == 'Data de saída':
            colunas_mapeadas[col] = 'Data_Saida'
        elif col == 'Cor/Raça':
            colunas_mapeadas[col] = 'Cor_Raca'
    
    # Renomear colunas
    df = df.rename(columns=colunas_mapeadas)
    
    # Converter tipos de dados
    if 'Data_Nascimento' in df.columns:
        df['Data_Nascimento'] = pd.to_datetime(df['Data_Nascimento'], dayfirst=True, errors='coerce')
    
    if 'Data_Entrada' in df.columns:
        df['Data_Entrada'] = pd.to_datetime(df['Data_Entrada'], dayfirst=True, errors='coerce')
    
    if 'Data_Saida' in df.columns:
        df['Data_Saida'] = pd.to_datetime(df['Data_Saida'], dayfirst=True, errors='coerce')
    
    # Padronizar texto dos campos principais
    for col in ['Nome_Estudante', 'Escola', 'Situacao', 'Turno', 'Nivel_Educacao', 'Ano_Serie', 'Turma']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    # Limpar formatação do CPF (remover pontos e traços)
    if 'CPF' in df.columns:
        df['CPF'] = df['CPF'].astype(str).apply(lambda x: re.sub(r'[^0-9]', '', x) if pd.notna(x) and x != 'nan' else x)
    
    # Marcar tipo de planilha
    df.attrs['tipo_planilha'] = 'censo_escolar'
    
    return df

def criar_interface_censo_escolar(df):
    """Cria interface específica para análise do Censo Escolar"""
    
    # Header específico para censo escolar
    st.markdown("""
    <div style="background: linear-gradient(90deg, #1e40af 0%, #3b82f6 100%); 
                padding: 2rem; border-radius: 10px; margin-bottom: 2rem; text-align: center;">
        <h1 style="color: white; margin: 0; font-size: 2.5rem; font-weight: bold;">
            📊 Painel Censo Escolar
        </h1>
        <p style="color: #e0e7ff; margin: 0.5rem 0 0 0; font-size: 1.2rem;">
            Identificação de Duplicatas - Estudantes em Múltiplas Escolas/Turmas
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Resumo Geral Simples
    st.markdown("### 📊 Resumo Geral")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total de Registros", f"{len(df):,}")
    
    with col2:
        escolas_unicas = df['Escola'].nunique() if 'Escola' in df.columns else 0
        st.metric("Escolas", escolas_unicas)
    
    with col3:
        estudantes_unicos = df['Nome_Estudante'].nunique() if 'Nome_Estudante' in df.columns else 0
        st.metric("Estudantes Únicos", estudantes_unicos)
    
    with col4:
        turmas_unicas = df['Turma'].nunique() if 'Turma' in df.columns else 0
        st.metric("Turmas", turmas_unicas)
    
    # Filtros Simples
    st.sidebar.markdown("### 🔍 Filtros")
    
    # Filtro por Escola
    if 'Escola' in df.columns:
        escolas_disponiveis = ['Todas as Escolas'] + sorted(df['Escola'].dropna().unique().tolist())
        escola_sel = st.sidebar.selectbox("Escola", escolas_disponiveis)
        
        if escola_sel != 'Todas as Escolas':
            df_filt = df[df['Escola'] == escola_sel].copy()
        else:
            df_filt = df.copy()
    else:
        df_filt = df.copy()
        escola_sel = 'Todas as Escolas'
    
    # Filtro por Situação (apenas Matriculado)
    if 'Situacao' in df.columns:
        situacoes_disponiveis = ['Todas as Situações'] + sorted(df_filt['Situacao'].dropna().unique().tolist())
        situacao_sel = st.sidebar.selectbox("Situação", situacoes_disponiveis)
        
        if situacao_sel != 'Todas as Situações':
            df_filt = df_filt[df_filt['Situacao'] == situacao_sel].copy()
    else:
        situacao_sel = 'Todas as Situações'
    
    # Resumo dos Dados Filtrados
    st.markdown("### 📋 Dados Após Filtros")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Registros", f"{len(df_filt):,}")
    
    with col2:
        estudantes_filtrados = df_filt['Nome_Estudante'].nunique() if 'Nome_Estudante' in df_filt.columns else 0
        st.metric("Estudantes", estudantes_filtrados)
    
    with col3:
        escolas_filtradas = df_filt['Escola'].nunique() if 'Escola' in df_filt.columns else 0
        st.metric("Escolas", escolas_filtradas)
    
    # Análise de Duplicatas - Foco Principal
    st.markdown("### 🔍 Duplicatas Encontradas")
    
    if 'Nome_Estudante' in df_filt.columns and 'Escola' in df_filt.columns:
        # 1. Duplicatas por Escola (estudante em múltiplas escolas)
        duplicatas_escola = df_filt.groupby('Nome_Estudante').agg({
            'Escola': 'nunique',
            'Turma': 'nunique' if 'Turma' in df_filt.columns else 'count'
        }).reset_index()
        
        estudantes_multiplas_escolas = duplicatas_escola[duplicatas_escola['Escola'] > 1]
        
        # 2. Duplicatas por Turma (estudante em múltiplas turmas na mesma escola)
        duplicatas_turma = df_filt.groupby(['Nome_Estudante', 'Escola']).agg({
            'Turma': 'nunique' if 'Turma' in df_filt.columns else 'count'
        }).reset_index()
        
        estudantes_multiplas_turmas = duplicatas_turma[duplicatas_turma['Turma'] > 1]
        
        # Métricas Principais
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Em Múltiplas Escolas", len(estudantes_multiplas_escolas))
        
        with col2:
            st.metric("Em Múltiplas Turmas", len(estudantes_multiplas_turmas))
        
        with col3:
            total_duplicatas = len(estudantes_multiplas_escolas) + len(estudantes_multiplas_turmas)
            st.metric("Total Duplicatas", total_duplicatas)
        
        with col4:
            percentual = (total_duplicatas / len(df_filt['Nome_Estudante'].unique())) * 100 if len(df_filt['Nome_Estudante'].unique()) > 0 else 0
            st.metric("Percentual", f"{percentual:.1f}%")
        
        # Tabelas Detalhadas
        if len(estudantes_multiplas_escolas) > 0 or len(estudantes_multiplas_turmas) > 0:
            
            # 1. Estudantes em Múltiplas Escolas (Detalhado)
            if len(estudantes_multiplas_escolas) > 0:
                st.markdown("#### 🏫 Estudantes em Múltiplas Escolas")
                
                # Criar tabela detalhada mostrando escola + turma para cada estudante
                duplicatas_escola_detalhadas = []
                for _, row in estudantes_multiplas_escolas.iterrows():
                    nome = row['Nome_Estudante']
                    dados_estudante = df_filt[df_filt['Nome_Estudante'] == nome]
                    
                    # Para cada escola do estudante, mostrar a turma correspondente
                    for _, linha in dados_estudante.iterrows():
                        cpf_limpo = re.sub(r'[^0-9]', '', str(linha['CPF'])) if 'CPF' in linha and pd.notna(linha['CPF']) else 'N/A'
                        duplicatas_escola_detalhadas.append({
                            'Nome': nome,
                            'Escola': linha['Escola'],
                            'Turma': linha['Turma'] if 'Turma' in linha else 'N/A',
                            'CPF': cpf_limpo,
                            'Situacao': linha['Situacao'] if 'Situacao' in linha else 'N/A'
                        })
                
                df_duplicatas_escola = pd.DataFrame(duplicatas_escola_detalhadas)
                st.dataframe(df_duplicatas_escola, use_container_width=True)
            
            # 2. Estudantes em Múltiplas Turmas (mesma escola) - Detalhado
            if len(estudantes_multiplas_turmas) > 0:
                st.markdown("#### 🎓 Estudantes em Múltiplas Turmas (Mesma Escola)")
                
                # Criar tabela detalhada mostrando cada linha de turma
                duplicatas_turma_detalhadas = []
                for _, row in estudantes_multiplas_turmas.iterrows():
                    nome = row['Nome_Estudante']
                    escola = row['Escola']
                    dados_estudante = df_filt[(df_filt['Nome_Estudante'] == nome) & 
                                            (df_filt['Escola'] == escola)]
                    
                    # Para cada turma do estudante na mesma escola
                    for _, linha in dados_estudante.iterrows():
                        cpf_limpo = re.sub(r'[^0-9]', '', str(linha['CPF'])) if 'CPF' in linha and pd.notna(linha['CPF']) else 'N/A'
                        duplicatas_turma_detalhadas.append({
                            'Nome': nome,
                            'Escola': escola,
                            'Turma': linha['Turma'] if 'Turma' in linha else 'N/A',
                            'CPF': cpf_limpo,
                            'Situacao': linha['Situacao'] if 'Situacao' in linha else 'N/A'
                        })
                
                df_duplicatas_turma = pd.DataFrame(duplicatas_turma_detalhadas)
                st.dataframe(df_duplicatas_turma, use_container_width=True)
            else:
                st.info("ℹ️ Nenhum estudante encontrado em múltiplas turmas da mesma escola.")
            
            # Botão de Download com Abas Separadas
            st.markdown("#### 💾 Download dos Dados")
            
            # Preparar dados para download em abas separadas
            if len(estudantes_multiplas_escolas) > 0 or len(estudantes_multiplas_turmas) > 0:
                
                # Converter para Excel com abas separadas
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    
                    # Aba 1: Duplicatas por Escola (Detalhado)
                    if len(estudantes_multiplas_escolas) > 0:
                        duplicatas_escola_download = []
                        for _, row in estudantes_multiplas_escolas.iterrows():
                            nome = row['Nome_Estudante']
                            dados_estudante = df_filt[df_filt['Nome_Estudante'] == nome]
                            
                            # Para cada escola do estudante, mostrar a turma correspondente
                            for _, linha in dados_estudante.iterrows():
                                cpf_limpo = re.sub(r'[^0-9]', '', str(linha['CPF'])) if 'CPF' in linha and pd.notna(linha['CPF']) else 'N/A'
                                duplicatas_escola_download.append({
                                    'Nome': nome,
                                    'Escola': linha['Escola'],
                                    'Turma': linha['Turma'] if 'Turma' in linha else 'N/A',
                                    'CPF': cpf_limpo,
                                    'Situacao': linha['Situacao'] if 'Situacao' in linha else 'N/A'
                                })
                        
                        df_escola_download = pd.DataFrame(duplicatas_escola_download)
                        df_escola_download.to_excel(writer, sheet_name='Múltiplas_Escolas', index=False)
                    
                    # Aba 2: Duplicatas por Turma (Detalhado)
                    if len(estudantes_multiplas_turmas) > 0:
                        duplicatas_turma_download = []
                        for _, row in estudantes_multiplas_turmas.iterrows():
                            nome = row['Nome_Estudante']
                            escola = row['Escola']
                            dados_estudante = df_filt[(df_filt['Nome_Estudante'] == nome) & 
                                                    (df_filt['Escola'] == escola)]
                            
                            # Para cada turma do estudante na mesma escola
                            for _, linha in dados_estudante.iterrows():
                                cpf_limpo = re.sub(r'[^0-9]', '', str(linha['CPF'])) if 'CPF' in linha and pd.notna(linha['CPF']) else 'N/A'
                                duplicatas_turma_download.append({
                                    'Nome': nome,
                                    'Escola': escola,
                                    'Turma': linha['Turma'] if 'Turma' in linha else 'N/A',
                                    'CPF': cpf_limpo,
                                    'Situacao': linha['Situacao'] if 'Situacao' in linha else 'N/A'
                                })
                        
                        df_turma_download = pd.DataFrame(duplicatas_turma_download)
                        df_turma_download.to_excel(writer, sheet_name='Múltiplas_Turmas', index=False)
                    
                    # Aba 3: Resumo Geral
                    resumo_geral = pd.DataFrame({
                        'Tipo_Duplicata': ['Múltiplas Escolas', 'Múltiplas Turmas', 'Total'],
                        'Quantidade': [
                            len(estudantes_multiplas_escolas),
                            len(estudantes_multiplas_turmas),
                            len(estudantes_multiplas_escolas) + len(estudantes_multiplas_turmas)
                        ],
                        'Percentual': [
                            f"{(len(estudantes_multiplas_escolas) / len(df_filt['Nome_Estudante'].unique())) * 100:.1f}%" if len(df_filt['Nome_Estudante'].unique()) > 0 else "0%",
                            f"{(len(estudantes_multiplas_turmas) / len(df_filt['Nome_Estudante'].unique())) * 100:.1f}%" if len(df_filt['Nome_Estudante'].unique()) > 0 else "0%",
                            f"{((len(estudantes_multiplas_escolas) + len(estudantes_multiplas_turmas)) / len(df_filt['Nome_Estudante'].unique())) * 100:.1f}%" if len(df_filt['Nome_Estudante'].unique()) > 0 else "0%"
                        ]
                    })
                    resumo_geral.to_excel(writer, sheet_name='Resumo', index=False)
                
                st.download_button(
                    label="📥 Baixar Relatório Completo (Excel com Abas)",
                    data=output.getvalue(),
                    file_name=f"duplicatas_censo_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.success("✅ Nenhuma duplicata encontrada nos dados filtrados!")
    
    
    # Dados Brutos (Opcional)
    with st.expander("📄 Ver todos os dados", expanded=False):
        st.dataframe(df_filt, use_container_width=True)

def criar_interface_conteudo_aplicado(df):
    """Cria interface específica para análise de conteúdo aplicado"""
    
    # Header específico para conteúdo aplicado
    st.markdown("""
    <div style="text-align: center; padding: 40px 20px; background: linear-gradient(135deg, #059669, #10b981); border-radius: 15px; margin-bottom: 30px; box-shadow: 0 8px 25px rgba(5, 150, 105, 0.3);">
        <h1 style="color: white; margin: 0; font-size: 2.2em; font-weight: 700; text-shadow: 0 2px 4px rgba(0,0,0,0.3);">Superintendência Regional de Educação de Gurupi TO</h1>
        <h2 style="color: white; margin: 15px 0 0 0; font-weight: 600; font-size: 1.8em; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Painel SGE - Conteúdo Aplicado</h2>
        <h3 style="color: rgba(255,255,255,0.95); margin: 10px 0 0 0; font-weight: 500; font-size: 1.4em;">Análise de Atividades e Conteúdos Registrados</h3>
        <p style="color: rgba(255,255,255,0.8); margin: 10px 0 0 0; font-size: 1.1em; font-weight: 400;">Registros de Conteúdo Aplicado</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Métricas gerais
    st.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
        <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Visão Geral dos Registros</h3>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric(
            label="Total de Registros", 
            value=f"{len(df):,}".replace(",", "."),
            help="Total de atividades/conteúdos registrados"
        )
    
    with col2:
        disciplinas_unicas = df["Disciplina"].nunique() if "Disciplina" in df.columns else 0
        st.metric(
            label="Disciplinas", 
            value=disciplinas_unicas,
            help="Número de disciplinas diferentes"
        )
    
    with col3:
        status_unicos = df["Status"].nunique() if "Status" in df.columns else 0
        st.metric(
            label="Status Diferentes", 
            value=status_unicos,
            help="Número de status diferentes"
        )
    
    with col4:
        if "Data" in df.columns:
            periodo_cobertura = f"{df['Data'].min().strftime('%d/%m/%Y')} a {df['Data'].max().strftime('%d/%m/%Y')}"
            st.metric(
                label="Período", 
                value=periodo_cobertura,
                help="Período coberto pelos registros"
            )
        else:
            st.metric("Período", "N/A")
    
    with col5:
        # Mostrar disciplina com mais registros
        if "Disciplina" in df.columns:
            disciplina_top = df["Disciplina"].value_counts().index[0] if len(df) > 0 else "N/A"
            qtd_top = df["Disciplina"].value_counts().iloc[0] if len(df) > 0 else 0
            st.metric(
                label="Disciplina Top", 
                value=f"{disciplina_top}",
                delta=f"{qtd_top} registros",
                help="Disciplina com maior número de registros"
            )
        else:
            st.metric("Disciplina Top", "N/A")
    
    # Função para classificar por bimestre baseado nas datas
    def classificar_bimestre(data):
        """Classifica a data em bimestre baseado nos períodos definidos"""
        if pd.isna(data):
            return "Sem Data"
        
        # Converter para datetime se necessário
        if not isinstance(data, pd.Timestamp):
            data = pd.to_datetime(data, errors='coerce')
        
        if pd.isna(data):
            return "Sem Data"
        
        # Definir períodos dos bimestres (ano 2025)
        bimestre1_inicio = pd.to_datetime("2025-02-03")
        bimestre1_fim = pd.to_datetime("2025-04-03")
        
        bimestre2_inicio = pd.to_datetime("2025-04-04")
        bimestre2_fim = pd.to_datetime("2025-06-27")
        
        bimestre3_inicio = pd.to_datetime("2025-08-04")
        bimestre3_fim = pd.to_datetime("2025-10-11")
        
        bimestre4_inicio = pd.to_datetime("2025-10-12")
        bimestre4_fim = pd.to_datetime("2025-12-19")
        
        # Classificar por bimestre
        if bimestre1_inicio <= data <= bimestre1_fim:
            return "1º Bimestre"
        elif bimestre2_inicio <= data <= bimestre2_fim:
            return "2º Bimestre"
        elif bimestre3_inicio <= data <= bimestre3_fim:
            return "3º Bimestre"
        elif bimestre4_inicio <= data <= bimestre4_fim:
            return "4º Bimestre"
        else:
            return "Fora do Período Letivo"
    
    # Adicionar coluna de bimestre se houver dados de data
    if "Data" in df.columns:
        df["Bimestre"] = df["Data"].apply(classificar_bimestre)
        
        
        
        # Análise por Bimestres
        st.markdown("""
        <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
            <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análise por Bimestres</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Contagem por bimestre
        contagem_bimestres = df["Bimestre"].value_counts().reset_index()
        contagem_bimestres.columns = ["Bimestre", "Quantidade"]
        
        # Ordenar por bimestre (1º, 2º, 3º, 4º)
        ordem_bimestres = ["1º Bimestre", "2º Bimestre", "3º Bimestre", "4º Bimestre", "Fora do Período Letivo", "Sem Data"]
        contagem_bimestres["Ordem"] = contagem_bimestres["Bimestre"].map({b: i for i, b in enumerate(ordem_bimestres)})
        contagem_bimestres = contagem_bimestres.sort_values("Ordem").reset_index(drop=True)
        
        # Criar colunas para mostrar bimestres
        num_bimestres = len(contagem_bimestres)
        num_colunas_bim = min(num_bimestres, 6)
        cols_bimestres = st.columns(num_colunas_bim)
        
        # Mostrar bimestres em cards
        for i, (_, row) in enumerate(contagem_bimestres.iterrows()):
            col_index = i % num_colunas_bim
            with cols_bimestres[col_index]:
                # Definir cor baseada no bimestre
                if "1º" in row['Bimestre']:
                    cor_borda = "#3b82f6"  # Azul
                elif "2º" in row['Bimestre']:
                    cor_borda = "#10b981"  # Verde
                elif "3º" in row['Bimestre']:
                    cor_borda = "#f59e0b"  # Amarelo
                elif "4º" in row['Bimestre']:
                    cor_borda = "#ef4444"  # Vermelho
                else:
                    cor_borda = "#6b7280"  # Cinza
                
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f0f9ff, #e0f2fe); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid {cor_borda};">
                    <div style="font-size: 0.9em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">{row['Bimestre']}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #1e40af; margin: 8px 0;">{row['Quantidade']}</div>
                    <div style="font-size: 1.1em; color: #64748b; font-weight: 600;">registros</div>
                </div>
                """, unsafe_allow_html=True)
        
        # Gráfico de barras por bimestre
        fig_bimestres = px.bar(contagem_bimestres, x="Bimestre", y="Quantidade", 
                              title="Registros por Bimestre",
                              color="Bimestre",
                              color_discrete_map={
                                  "1º Bimestre": "#3b82f6",
                                  "2º Bimestre": "#10b981", 
                                  "3º Bimestre": "#f59e0b",
                                  "4º Bimestre": "#ef4444",
                                  "Fora do Período Letivo": "#6b7280",
                                  "Sem Data": "#9ca3af"
                              })
        fig_bimestres.update_layout(xaxis_tickangle=45, showlegend=False)
        st.plotly_chart(fig_bimestres, use_container_width=True)
        
        # Análise detalhada por bimestre - disciplinas em cada bimestre
        st.markdown("""
        <div style="background: linear-gradient(135deg, #7c3aed, #a855f7); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(124, 58, 237, 0.2);">
            <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Disciplinas por Bimestre</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Criar análise por bimestre e disciplina
        bimestre_disciplina = df.groupby(['Bimestre', 'Disciplina']).size().reset_index(name='Quantidade')
        
        # Ordenar por bimestre e quantidade
        ordem_bimestres = ["1º Bimestre", "2º Bimestre", "3º Bimestre", "4º Bimestre", "Fora do Período Letivo", "Sem Data"]
        bimestre_disciplina['Ordem_Bimestre'] = bimestre_disciplina['Bimestre'].map({b: i for i, b in enumerate(ordem_bimestres)})
        bimestre_disciplina = bimestre_disciplina.sort_values(['Ordem_Bimestre', 'Quantidade'], ascending=[True, False])
        
        # Mostrar cada bimestre com suas disciplinas
        for bimestre in ordem_bimestres:
            if bimestre in bimestre_disciplina['Bimestre'].values:
                disciplinas_bimestre = bimestre_disciplina[bimestre_disciplina['Bimestre'] == bimestre]
                
                # Definir cor do bimestre
                if "1º" in bimestre:
                    cor_bimestre = "#3b82f6"  # Azul
                elif "2º" in bimestre:
                    cor_bimestre = "#10b981"  # Verde
                elif "3º" in bimestre:
                    cor_bimestre = "#f59e0b"  # Amarelo
                elif "4º" in bimestre:
                    cor_bimestre = "#ef4444"  # Vermelho
                else:
                    cor_bimestre = "#6b7280"  # Cinza
                
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f8fafc, #f1f5f9); border-radius: 8px; padding: 20px; margin: 15px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-left: 4px solid {cor_bimestre};">
                    <h4 style="color: {cor_bimestre}; margin: 0 0 15px 0; font-size: 1.3em; font-weight: 700;">{bimestre}</h4>
                </div>
                """, unsafe_allow_html=True)
                
                # Criar colunas para as disciplinas deste bimestre
                num_disciplinas = len(disciplinas_bimestre)
                num_colunas_disc = min(num_disciplinas, 4)  # Máximo 4 colunas
                cols_disciplinas = st.columns(num_colunas_disc)
                
                # Mostrar disciplinas em cards
                for i, (_, row) in enumerate(disciplinas_bimestre.iterrows()):
                    col_index = i % num_colunas_disc
                    with cols_disciplinas[col_index]:
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #ffffff, #f8fafc); border-radius: 6px; padding: 12px; margin: 5px 0; box-shadow: 0 1px 4px rgba(0,0,0,0.1); border-left: 3px solid {cor_bimestre};">
                            <div style="font-size: 0.9em; font-weight: 600; color: #374151; margin-bottom: 6px;">{row['Disciplina']}</div>
                            <div style="font-size: 1.5em; font-weight: 700; color: {cor_bimestre}; margin: 6px 0;">{row['Quantidade']}</div>
                            <div style="font-size: 0.9em; color: #6b7280; font-weight: 500;">registros</div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Gráfico de barras para este bimestre
                fig_bimestre_disc = px.bar(disciplinas_bimestre, x="Disciplina", y="Quantidade", 
                                          title=f"Disciplinas - {bimestre}",
                                          color="Disciplina",
                                          color_discrete_sequence=px.colors.qualitative.Set3)
                fig_bimestre_disc.update_layout(xaxis_tickangle=45, showlegend=False, height=300)
                st.plotly_chart(fig_bimestre_disc, use_container_width=True)
    
    # Adicionar seção com disciplinas (todas ou filtradas) - será movida para depois dos filtros
    
    # Filtros específicos para conteúdo aplicado
    st.sidebar.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
        <h2 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Filtros - Conteúdo</h2>
        <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1em; font-weight: 500;">Filtre os registros de conteúdo</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Filtros
    disciplinas_opcoes = sorted(df["Disciplina"].dropna().unique().tolist()) if "Disciplina" in df.columns else []
    status_opcoes = sorted(df["Status"].dropna().unique().tolist()) if "Status" in df.columns else []
    bimestres_opcoes = sorted(df["Bimestre"].dropna().unique().tolist()) if "Bimestre" in df.columns else []
    
    # Filtro de Data
    if "Data" in df.columns:
        st.sidebar.markdown("""
        <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
            <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">📅 Período</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Obter datas mínima e máxima
        data_min = df["Data"].min()
        data_max = df["Data"].max()
        
        # Filtro de data com slider
        data_range = st.sidebar.date_input(
            "Selecione o período:",
            value=(data_min.date(), data_max.date()),
            min_value=data_min.date(),
            max_value=data_max.date(),
            help="Selecione o período para filtrar os registros"
        )
        
        # Converter para datetime se necessário
        if len(data_range) == 2:
            data_inicio = pd.to_datetime(data_range[0])
            data_fim = pd.to_datetime(data_range[1])
        else:
            data_inicio = data_min
            data_fim = data_max
    
    # Filtro de Disciplina
    st.sidebar.markdown("""
    <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
        <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">📚 Disciplina</h3>
    </div>
    """, unsafe_allow_html=True)
    
    disciplina_sel = st.sidebar.multiselect(
        "Selecione as disciplinas:", 
        disciplinas_opcoes, 
        help="Filtre por disciplinas específicas"
    )
    
    # Filtro de Status
    st.sidebar.markdown("""
    <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
        <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">✅ Status</h3>
    </div>
    """, unsafe_allow_html=True)
    
    status_sel = st.sidebar.multiselect(
        "Selecione os status:", 
        status_opcoes, 
        help="Filtre por status específicos"
    )
    
    # Filtro de Bimestre
    if "Bimestre" in df.columns and len(bimestres_opcoes) > 0:
        st.sidebar.markdown("""
        <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
            <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">📅 Bimestre</h3>
        </div>
        """, unsafe_allow_html=True)
        
        bimestre_sel = st.sidebar.multiselect(
            "Selecione os bimestres:", 
            bimestres_opcoes, 
            help="Filtre por bimestres específicos"
        )
    else:
        bimestre_sel = []
    
    # Aplicar filtros
    df_filtrado = df.copy()
    
    # Filtro por data
    if "Data" in df.columns and 'data_inicio' in locals() and 'data_fim' in locals():
        df_filtrado = df_filtrado[
            (df_filtrado["Data"] >= data_inicio) & 
            (df_filtrado["Data"] <= data_fim)
        ]
    
    # Filtro por disciplina
    if disciplina_sel:
        df_filtrado = df_filtrado[df_filtrado["Disciplina"].isin(disciplina_sel)]
    
    # Filtro por status
    if status_sel:
        df_filtrado = df_filtrado[df_filtrado["Status"].isin(status_sel)]
    
    # Filtro por bimestre
    if bimestre_sel:
        df_filtrado = df_filtrado[df_filtrado["Bimestre"].isin(bimestre_sel)]
    
    # Verificar se há filtros aplicados (agora que as variáveis estão definidas)
    tem_filtros = (
        ('data_inicio' in locals() and 'data_fim' in locals() and 
         (data_inicio != df["Data"].min() or data_fim != df["Data"].max())) or
        disciplina_sel or 
        status_sel or
        bimestre_sel
    )
    
    # Determinar título e dados baseado nos filtros
    if tem_filtros:
        titulo_secao = "Disciplinas Filtradas"
        dados_disciplinas = df_filtrado["Disciplina"].value_counts().reset_index() if len(df_filtrado) > 0 else pd.DataFrame()
    else:
        titulo_secao = "Todas as Disciplinas"
        dados_disciplinas = df["Disciplina"].value_counts().reset_index()
    
    dados_disciplinas.columns = ["Disciplina", "Quantidade"]
    
    # Adicionar seção com disciplinas (todas ou filtradas)
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
        <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">{titulo_secao}</h3>
    </div>
    """, unsafe_allow_html=True)
    
    if len(dados_disciplinas) > 0:
        # Calcular número de colunas necessárias (máximo 6 para não ficar muito pequeno)
        num_disciplinas = len(dados_disciplinas)
        num_colunas = min(num_disciplinas, 6)
        
        # Criar colunas dinamicamente
        cols_disciplinas = st.columns(num_colunas)
        
        # Mostrar disciplinas em cards
        for i, (_, row) in enumerate(dados_disciplinas.iterrows()):
            col_index = i % num_colunas
            with cols_disciplinas[col_index]:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(5, 150, 105, 0.15); border-left: 4px solid #059669;">
                    <div style="font-size: 0.9em; font-weight: 600; color: #047857; margin-bottom: 8px;">{row['Disciplina']}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #047857; margin: 8px 0;">{row['Quantidade']}</div>
                    <div style="font-size: 1.1em; color: #64748b; font-weight: 600;">registros</div>
                </div>
                """, unsafe_allow_html=True)
        
        # Se há mais de 6 disciplinas, mostrar aviso
        if num_disciplinas > 6:
            st.info(f"Mostrando as primeiras 6 disciplinas de {num_disciplinas} total. Use os filtros para focar em disciplinas específicas.")
    
    # Mostrar informações dos filtros aplicados
    st.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
        <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Dados Filtrados</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Métricas dos dados filtrados
    col_filt1, col_filt2, col_filt3 = st.columns(3)
    
    with col_filt1:
        st.metric(
            label="Registros Filtrados", 
            value=f"{len(df_filtrado):,}".replace(",", "."),
            delta=f"{len(df_filtrado) - len(df)}" if len(df_filtrado) != len(df) else "0",
            help="Total de registros após aplicar os filtros"
        )
    
    with col_filt2:
        if len(df_filtrado) > 0 and "Disciplina" in df_filtrado.columns:
            disciplinas_filtradas = df_filtrado["Disciplina"].nunique()
            st.metric(
                label="Disciplinas no Filtro", 
                value=disciplinas_filtradas,
                help="Número de disciplinas nos dados filtrados"
            )
        else:
            st.metric("Disciplinas no Filtro", "0")
    
    with col_filt3:
        if len(df_filtrado) > 0 and "Data" in df_filtrado.columns:
            periodo_filtrado = f"{df_filtrado['Data'].min().strftime('%d/%m/%Y')} a {df_filtrado['Data'].max().strftime('%d/%m/%Y')}"
            st.metric(
                label="Período Filtrado", 
                value=periodo_filtrado,
                help="Período dos dados filtrados"
            )
        else:
            st.metric("Período Filtrado", "N/A")
    
    # Análise por Disciplina
    st.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
        <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análise por Disciplina</h3>
    </div>
    """, unsafe_allow_html=True)
    
    if len(df_filtrado) > 0:
        # Contagem por disciplina
        contagem_disciplina = df_filtrado["Disciplina"].value_counts().reset_index()
        contagem_disciplina.columns = ["Disciplina", "Quantidade"]
        
        # Gráfico de barras
        fig = px.bar(contagem_disciplina, x="Disciplina", y="Quantidade", 
                    title="Registros por Disciplina",
                    color="Quantidade",
                    color_continuous_scale="Viridis")
        fig.update_layout(xaxis_tickangle=45)
        st.plotly_chart(fig, use_container_width=True)
        
        # Tabela detalhada
        st.markdown("### Registros Detalhados")
        st.dataframe(df_filtrado, use_container_width=True)
        
        # Botão de exportação
        col_export1, col_export2 = st.columns([1, 4])
        with col_export1:
            if st.button("📊 Exportar Dados", key="export_conteudo", help="Baixar planilha com análise de conteúdo aplicado"):
                excel_data = criar_excel_formatado(df_filtrado, "Conteudo_Aplicado")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="conteudo_aplicado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Nenhum registro encontrado com os filtros aplicados.")

def mapear_bimestre(periodo: str) -> int | None:
    """Mapeia 'Primeiro Bimestre' -> 1, 'Segundo Bimestre' -> 2, etc."""
    if not isinstance(periodo, str):
        return None
    p = periodo.lower()
    if "primeiro" in p or "1º" in p or "1o" in p:
        return 1
    if "segundo" in p or "2º" in p or "2o" in p:
        return 2
    if "terceiro" in p or "3º" in p or "3o" in p:
        return 3
    if "quarto" in p or "4º" in p or "4o" in p:
        return 4
    return None

def classificar_status_b1_b2(n1, n2, media12):
    """
    Regras:
      - 'Vermelho Duplo': n1<6 e n2<6
      - 'Queda p/ Vermelho': n1>=6 e n2<6
      - 'Recuperou': n1<6 e n2>=6
      - 'Verde': n1>=6 e n2>=6
      - Se faltar n1 ou n2, retorna 'Incompleto'
    """
    if pd.isna(n1) or pd.isna(n2):
        return "Incompleto"
    if n1 < MEDIA_APROVACAO and n2 < MEDIA_APROVACAO:
        return "Vermelho Duplo"
    if n1 >= MEDIA_APROVACAO and n2 < MEDIA_APROVACAO:
        return "Queda p/ Vermelho"
    if n1 < MEDIA_APROVACAO and n2 >= MEDIA_APROVACAO:
        return "Recuperou"
    return "Verde"

def classificar_status_b1_b2_b3(n1, n2, n3, media123):
    """
    Classificação considerando 3 bimestres:
      - 'Vermelho Triplo': n1<6, n2<6 e n3<6
      - 'Vermelho Duplo': duas notas abaixo de 6
      - 'Queda Recente': n1>=6 e/ou n2>=6, mas n3<6
      - 'Recuperação': estava abaixo e melhorou no 3º bimestre
      - 'Verde': todas as notas >= 6
      - 'Incompleto': falta alguma nota
    """
    # Verificar se falta alguma nota
    if pd.isna(n1) or pd.isna(n2) or pd.isna(n3):
        return "Incompleto"
    
    # Contar quantas notas estão abaixo da média
    notas_abaixo = sum([n1 < MEDIA_APROVACAO, n2 < MEDIA_APROVACAO, n3 < MEDIA_APROVACAO])
    
    if notas_abaixo == 0:
        return "Verde"  # Todas acima de 6
    elif notas_abaixo == 3:
        return "Vermelho Triplo"  # Todas abaixo de 6
    elif notas_abaixo == 2:
        return "Vermelho Duplo"  # Duas abaixo de 6
    else:  # notas_abaixo == 1
        # Verificar se é queda recente ou recuperação
        if n3 < MEDIA_APROVACAO:
            return "Queda Recente"  # Estava bem mas caiu no 3º
        else:
            return "Recuperação"  # Estava mal mas melhorou

def classificar_status_b1_b2_b3_b4(n1, n2, n3, n4, media_final):
    """
    Classificação considerando 4 bimestres:
      - 'Vermelho Quádruplo': todas as 4 notas < 6
      - 'Vermelho Triplo': três notas abaixo de 6
      - 'Vermelho Duplo': duas notas abaixo de 6
      - 'Queda Final': estava bem mas caiu no 4º bimestre
      - 'Recuperação Final': estava mal mas melhorou no 4º bimestre
      - 'Verde': todas as notas >= 6
      - 'Incompleto': falta alguma nota
    """
    # Verificar se falta alguma nota
    if pd.isna(n1) or pd.isna(n2) or pd.isna(n3) or pd.isna(n4):
        return "Incompleto"
    
    # Contar quantas notas estão abaixo da média
    notas_abaixo = sum([n1 < MEDIA_APROVACAO, n2 < MEDIA_APROVACAO, n3 < MEDIA_APROVACAO, n4 < MEDIA_APROVACAO])
    
    if notas_abaixo == 0:
        return "Verde"  # Todas acima de 6
    elif notas_abaixo == 4:
        return "Vermelho Quádruplo"  # Todas abaixo de 6
    elif notas_abaixo == 3:
        return "Vermelho Triplo"  # Três abaixo de 6
    elif notas_abaixo == 2:
        return "Vermelho Duplo"  # Duas abaixo de 6
    else:  # notas_abaixo == 1
        # Verificar se é queda final ou recuperação final
        if n4 < MEDIA_APROVACAO:
            return "Queda Final"  # Estava bem mas caiu no 4º
        else:
            return "Recuperação Final"  # Estava mal mas melhorou no 4º

def classificar_aprovacao_final(media_final):
    """
    Classifica se o aluno está aprovado ou reprovado baseado na média final.
    Aprovado: média final >= 6.0
    Reprovado: média final < 6.0
    Incompleto: se não houver média final
    """
    if pd.isna(media_final):
        return "Incompleto"
    elif media_final >= MEDIA_APROVACAO:
        return "Aprovado"
    else:
        return "Reprovado"

def criar_excel_formatado(df, nome_planilha="Dados"):
    """
    Cria um arquivo Excel formatado usando pandas (método mais simples e confiável)
    """
    # Usar pandas para criar o Excel diretamente
    output = BytesIO()
    
    # Criar o arquivo Excel usando pandas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        # Acessar a planilha para formatação
        workbook = writer.book
        worksheet = writer.sheets[nome_planilha]
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()

def calcula_indicadores(df):
    """
    Cria um dataframe por Aluno-Disciplina com:
      N1, N2, N3, N4, Media123, Soma123, ReqMediaProx1 (quanto precisa no próximo bimestre para fechar 6 no ano), Classificacao
    """
    # Criar coluna Bimestre
    df = df.copy()
    df["Bimestre"] = df["Periodo"].apply(mapear_bimestre)

    # Pivot por (Aluno, Turma, Disciplina)
    # Detectar coluna de aluno/estudante
    coluna_aluno = None
    for col in ["Aluno", "Nome_Estudante", "Estudante"]:
        if col in df.columns:
            coluna_aluno = col
            break
    
    pivot = df.pivot_table(
        index=["Escola", "Turma", coluna_aluno, "Disciplina"],
        columns="Bimestre",
        values="Nota",
        aggfunc="mean"
    ).reset_index()

    # Renomear colunas 1..4 para N1..N4 (se existirem)
    rename_cols = {}
    for b in [1, 2, 3, 4]:
        if b in pivot.columns:
            rename_cols[b] = f"N{b}"
    pivot = pivot.rename(columns=rename_cols)

    # Obter as notas dos 4 bimestres
    n1 = pivot.get("N1", pd.Series([np.nan] * len(pivot)))
    n2 = pivot.get("N2", pd.Series([np.nan] * len(pivot)))
    n3 = pivot.get("N3", pd.Series([np.nan] * len(pivot)))
    n4 = pivot.get("N4", pd.Series([np.nan] * len(pivot)))
    
    # Se não existir a coluna, criar uma série de NaN
    if isinstance(n1, float):
        n1 = pd.Series([np.nan] * len(pivot))
    if isinstance(n2, float):
        n2 = pd.Series([np.nan] * len(pivot))
    if isinstance(n3, float):
        n3 = pd.Series([np.nan] * len(pivot))
    if isinstance(n4, float):
        n4 = pd.Series([np.nan] * len(pivot))
    
    # Calcular métricas dos 3 primeiros bimestres (para compatibilidade)
    pivot["Soma123"] = n1.fillna(0) + n2.fillna(0) + n3.fillna(0)
    # Média dos 3 bimestres (se algum for NaN, considerar apenas os disponíveis)
    pivot["Media123"] = (n1 + n2 + n3) / 3
    
    # Manter também as métricas antigas para compatibilidade
    pivot["Soma12"] = n1.fillna(0) + n2.fillna(0)
    pivot["Media12"] = (n1 + n2) / 2

    # Calcular métricas finais dos 4 bimestres
    pivot["SomaFinal"] = n1.fillna(0) + n2.fillna(0) + n3.fillna(0) + n4.fillna(0)
    # Média final dos 4 bimestres
    pivot["MediaFinal"] = (n1 + n2 + n3 + n4) / 4

    # Quanto precisa no próximo bimestre (N4) para fechar soma >= 24 (quando ainda não tem N4)
    pivot["PrecisaSomarProx1"] = SOMA_FINAL_ALVO - pivot["Soma123"]
    pivot["ReqMediaProx1"] = pivot["PrecisaSomarProx1"]
    
    # Manter também as métricas antigas para compatibilidade
    pivot["PrecisaSomarProx2"] = SOMA_FINAL_ALVO - pivot["Soma12"]
    pivot["ReqMediaProx2"] = pivot["PrecisaSomarProx2"] / 2

    # Classificação com 3 bimestres (para quando ainda não tem N4)
    # Se N3 não existe (None/NaN), marca como Incompleto já que esperamos 3 bimestres
    pivot["Classificacao"] = [
        classificar_status_b1_b2_b3(_n1, _n2, _n3, _m123) if pd.notna(_n3) and pd.isna(_n4)
        else "Incompleto"  # Se falta N3, é incompleto
        for _n1, _n2, _n3, _n4, _m123, _m12 in zip(
            pivot.get("N1", np.nan), 
            pivot.get("N2", np.nan), 
            pivot.get("N3", np.nan),
            pivot.get("N4", np.nan),
            pivot["Media123"],
            pivot["Media12"]
        )
    ]

    # Classificação com 4 bimestres (quando N4 está disponível)
    # Se N4 existe, usar classificação com 4 bimestres
    pivot["ClassificacaoFinal"] = [
        classificar_status_b1_b2_b3_b4(_n1, _n2, _n3, _n4, _m_final) if pd.notna(_n4)
        else _class_antiga  # Se falta N4, usar classificação anterior
        for _n1, _n2, _n3, _n4, _m_final, _class_antiga in zip(
            pivot.get("N1", np.nan), 
            pivot.get("N2", np.nan), 
            pivot.get("N3", np.nan),
            pivot.get("N4", np.nan),
            pivot["MediaFinal"],
            pivot["Classificacao"]
        )
    ]

    # Classificação de Aprovação/Reprovação Final
    pivot["StatusFinal"] = pivot["MediaFinal"].apply(classificar_aprovacao_final)

    # Flags de alerta
    # "Corda Bamba": precisa de nota >= 7 no próximo bimestre (ou média >= 7 nos próximos 2)
    pivot["CordaBamba"] = (pivot["ReqMediaProx1"] >= 7) | (pivot["ReqMediaProx2"] >= 7)

    # "Alerta": qualquer situação crítica ou Corda Bamba
    pivot["Alerta"] = pivot["Classificacao"].isin([
        "Vermelho Triplo", "Vermelho Duplo", "Queda p/ Vermelho", "Queda Recente"
    ]) | pivot["CordaBamba"]

    return pivot

# -----------------------------
# Controle de Acesso
# -----------------------------
# Inicializar variáveis de sessão
if 'logado' not in st.session_state:
    st.session_state.logado = True
if 'usuario' not in st.session_state or not st.session_state.usuario:
    st.session_state.usuario = {'nome': 'Visitante'}
if 'mostrar_alterar_senha' not in st.session_state:
    st.session_state.mostrar_alterar_senha = False
if 'mostrar_instrucoes' not in st.session_state:
    st.session_state.mostrar_instrucoes = False
if 'mostrar_admin' not in st.session_state:
    st.session_state.mostrar_admin = False
if 'admin_logado' not in st.session_state:
    st.session_state.admin_logado = False
if 'mostrar_relatorio' not in st.session_state:
    st.session_state.mostrar_relatorio = False
if 'mostrar_stats_usuario' not in st.session_state:
    st.session_state.mostrar_stats_usuario = False
if 'mostrar_sobre' not in st.session_state:
    st.session_state.mostrar_sobre = False
if 'email_aguardando_codigo' not in st.session_state:
    st.session_state.email_aguardando_codigo = ""
if 'login_time' not in st.session_state:
    st.session_state.login_time = datetime.now()
if 'sessao_expirada' not in st.session_state:
    st.session_state.sessao_expirada = False

# Verificar se deve mostrar tela de instruções
if st.session_state.mostrar_instrucoes:
    tela_instrucoes()
    st.stop()


# Verificar se deve mostrar modal sobre
if st.session_state.mostrar_sobre:
    tela_sobre()
    st.stop()

# Verificar se deve mostrar área administrativa
if st.session_state.mostrar_admin:
    if MONITORING_AVAILABLE:
        if not st.session_state.admin_logado:
            tela_admin()
            st.stop()
        else:
            # Verificar qual tela administrativa mostrar
            if st.session_state.mostrar_relatorio:
                relatorio_completo()
                st.stop()
            elif st.session_state.mostrar_stats_usuario:
                estatisticas_usuario()
                st.stop()
            else:
                dashboard_admin()
                st.stop()
    else:
        st.error("Sistema de monitoramento não disponível. Verifique as dependências do Firebase.")
        if st.button("⬅️ Voltar", key="btn_voltar_admin"):
            st.session_state.mostrar_admin = False
            st.rerun()
        st.stop()

# -----------------------------
# UI – Entrada de dados
# -----------------------------
# Header com boas-vindas personalizadas
st.markdown(f"""
<div style="position: relative; overflow: hidden; text-align: center; padding: 48px 24px; background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 55%, #60a5fa 100%); border-radius: 28px; margin-bottom: 28px; box-shadow: 0 22px 50px rgba(30, 64, 175, 0.24); border: 1px solid rgba(255,255,255,0.14);">
    <div style="position:absolute; inset:auto -80px -80px auto; width:220px; height:220px; background:rgba(255,255,255,0.08); border-radius:50%;"></div>
    <div style="position:absolute; inset:-60px auto auto -50px; width:160px; height:160px; background:rgba(255,255,255,0.08); border-radius:50%;"></div>
    <div style="position: relative; z-index: 2;">
        <span style="display:inline-block; padding:8px 16px; border-radius:999px; background:rgba(255,255,255,0.14); color:#dbeafe; font-size:0.95rem; font-weight:700; letter-spacing:0.04em;">PAINEL INTERATIVO SGE</span>
        <h1 style="color: white; margin: 18px 0 8px 0; font-size: 2.4em; font-weight: 800; text-shadow: 0 2px 4px rgba(0,0,0,0.28);">Superintendência Regional de Educação de Porto Nacional TO</h1>
        <h2 style="color: #eff6ff; margin: 0; font-weight: 600; font-size: 1.45em;">Notas, Frequência, Riscos e Alertas</h2>
        <p style="color: rgba(255,255,255,0.85); margin: 12px 0 0 0; font-size: 1.05em; font-weight: 500;">Análise integrada dos 1º, 2º, 3º e 4º bimestres</p>
        <div style="margin: 22px auto 0 auto; max-width: 520px; padding: 16px 18px; background: rgba(255,255,255,0.12); border-radius: 18px; border: 1px solid rgba(255,255,255,0.18); backdrop-filter: blur(6px);">
            <p style="color: white; margin: 0; font-size: 1.12em; font-weight: 700;">👋 Olá, {st.session_state.usuario['nome']}!</p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# Barra de navegação simplificada
col_nav1, col_nav2, col_nav3, col_nav4 = st.columns([2.2, 1, 1, 1])

with col_nav2:
    if st.button("📘 Instruções", use_container_width=True, key="btn_instrucoes_header"):
        st.session_state.mostrar_instrucoes = True
        st.rerun()

with col_nav3:
    if st.button("ℹ️ Sobre", use_container_width=True, key="btn_sobre"):
        st.session_state.mostrar_sobre = True

with col_nav4:
    st.markdown("<div style='text-align:center; padding-top:0.4rem; color:#475569; font-weight:700;'>Acesso livre</div>", unsafe_allow_html=True)

st.markdown("---")

col_upl, col_info = st.columns([1, 2])
with col_upl:
    st.markdown("### 📂 Carregar dados")
    arquivo = st.file_uploader("Planilha (.xlsx) do SGE", type=["xlsx"], help="Faça upload da planilha ou salve como 'dados.xlsx' na pasta")
with col_info:
    st.markdown("### 🚀 Como usar")
    st.markdown("""
    **1.** Carregue sua planilha no uploader ou salve como `dados.xlsx`  
    **2.** Use os filtros na barra lateral para focar em turmas/disciplinas específicas  
    **3.** Analise os alertas, frequência e riscos dos alunos  
    **4.** Identifique quem precisa de atenção imediata
    """)

# Carregar
try:
    df = carregar_dados(arquivo)
    
    # Verificar tipo de planilha e rotear para interface apropriada
    tipo_planilha = df.attrs.get('tipo_planilha', 'notas_frequencia')

    if tipo_planilha == 'notas_frequencia' and "Bimestre" not in df.columns and "Periodo" in df.columns:
        df["Bimestre"] = df["Periodo"].apply(mapear_bimestre)
    
    if tipo_planilha == 'conteudo_aplicado':
        # Mostrar interface específica para conteúdo aplicado
        criar_interface_conteudo_aplicado(df)
        
        # Assinatura discreta do criador
        st.markdown("---")
        st.markdown(
            """
            <div style="text-align: center; margin-top: 40px; padding: 20px;">
                <p style="margin: 0;">
                    Desenvolvido por <strong style="color: #059669;">Alexandre Tolentino</strong> • 
                    <em>Painel SGE - Conteúdo Aplicado</em>
                </p>
            </div>
            """, 
            unsafe_allow_html=True
        )
    elif tipo_planilha == 'censo_escolar':
        # Mostrar interface específica para censo escolar
        criar_interface_censo_escolar(df)
        
        # Assinatura discreta do criador
        st.markdown("---")
        st.markdown(
            """
            <div style="text-align: center; margin-top: 40px; padding: 20px;">
                <p style="margin: 0;">
                    Desenvolvido por <strong style="color: #059669;">Alexandre Tolentino</strong> • 
                    <em>Painel SGE - Censo Escolar</em>
                </p>
            </div>
            """, 
            unsafe_allow_html=True
        )
        st.stop()
    else:
        # Continuar com interface padrão de notas/frequência
        pass
        
except FileNotFoundError:
    st.error("Não encontrei `dados.xlsx` na pasta e nenhum arquivo foi enviado no uploader.")
    
    # Assinatura discreta do criador (quando não há dados)
    st.markdown("---")
    st.markdown(
        """
        <div style="text-align: center; margin-top: 40px; padding: 20px;">
            <p style="margin: 0;">
                Desenvolvido por <strong style="color: #1e40af;">Alexandre Tolentino</strong> • 
                <em>Painel SGE</em>
            </p>
        </div>
        """, 
        unsafe_allow_html=True
    )
    st.stop()

# Conferência mínima - Dados Gerais
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Visão Geral dos Dados</h3>
</div>
""", unsafe_allow_html=True)

colA, colB, colC, colD, colE = st.columns(5)

with colA:
    st.metric(
        label="Registros", 
        value=f"{len(df):,}".replace(",", "."),
        help="Total de linhas de dados na planilha"
    )
with colB:
    st.metric(
        label="Escolas", 
        value=df["Escola"].nunique() if "Escola" in df.columns else 0,
        help="Número de escolas diferentes"
    )
with colC:
    st.metric(
        label="Turmas", 
        value=df["Turma"].nunique() if "Turma" in df.columns else 0,
        help="Número de turmas diferentes"
    )
with colD:
    st.metric(
        label="Disciplinas", 
        value=df["Disciplina"].nunique() if "Disciplina" in df.columns else 0,
        help="Número de disciplinas diferentes"
    )
with colE:
    st.metric(
        label="Status", 
        value=df["Status"].nunique() if "Status" in df.columns else 0,
        help="Número de status diferentes"
    )

# Adicionar métrica de total de estudantes únicos
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">👥 Total de Estudantes</h3>
</div>
""", unsafe_allow_html=True)

col_total = st.columns(1)[0]
with col_total:
    # Detectar coluna de aluno/estudante
    coluna_aluno = None
    for col in ["Aluno", "Nome_Estudante", "Estudante"]:
        if col in df.columns:
            coluna_aluno = col
            break
    
    total_estudantes = df[coluna_aluno].nunique() if coluna_aluno else 0
    st.metric(
        label="Estudantes Únicos", 
        value=f"{total_estudantes:,}".replace(",", "."),
        help="Total de estudantes únicos na escola (sem repetição por disciplina)"
    )


# -----------------------------
# Filtros laterais
# -----------------------------
st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Filtros</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1em; font-weight: 500;">Filtre os dados para análise específica</p>
</div>
""", unsafe_allow_html=True)

escolas = sorted(df["Escola"].dropna().unique().tolist()) if "Escola" in df.columns else []
status_opcoes = sorted(df["Status"].dropna().unique().tolist()) if "Status" in df.columns else []
bimestres_opcoes = [1, 2, 3, 4] if "Bimestre" in df.columns else []
bimestres_opcoes = [1, 2, 3, 4] if "Bimestre" in df.columns else []

st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
    <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">Escola</h3>
</div>
""", unsafe_allow_html=True)
escola_sel = st.sidebar.selectbox("Selecione a escola:", ["Todas"] + escolas, help="Filtre por escola específica")

st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
    <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">Status</h3>
</div>
""", unsafe_allow_html=True)
# Botões de ação rápida para status
col_s1, col_s2 = st.sidebar.columns(2)
with col_s1:
    if st.button("Todas", key="btn_todas_status", help="Selecionar todos os status"):
        st.session_state.status_selecionados = status_opcoes
with col_s2:
    if st.button("Limpar", key="btn_limpar_status", help="Limpar seleção"):
        st.session_state.status_selecionados = []

# Inicializar estado se não existir
if 'status_selecionados' not in st.session_state:
    st.session_state.status_selecionados = []

status_sel = st.sidebar.multiselect(
    "Selecione os status:", 
    status_opcoes, 
    default=st.session_state.status_selecionados,
    help="Use os botões acima para seleção rápida"
)

# Filtrar dados baseado na escola e status selecionados para mostrar opções relevantes
df_temp = df.copy()
if escola_sel != "Todas":
    df_temp = df_temp[df_temp["Escola"] == escola_sel]
if status_sel:  # Se algum status foi selecionado
    df_temp = df_temp[df_temp["Status"].isin(status_sel)]
else:  # Se nenhum status selecionado, mostra todos
    pass  # Mantém todos os status

st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(59, 130, 246, 0.14); border-left: 3px solid #2563eb;">
    <h3 style="color: #1d4ed8; margin: 0; font-size: 1em; font-weight: 600;">Bimestre</h3>
</div>
""", unsafe_allow_html=True)
col_b1, col_b2 = st.sidebar.columns(2)
with col_b1:
    if st.button("Todos", key="btn_todos_bimestres", help="Selecionar todos os bimestres"):
        st.session_state.bimestres_selecionados = bimestres_opcoes
with col_b2:
    if st.button("Limpar", key="btn_limpar_bimestres", help="Limpar seleção"):
        st.session_state.bimestres_selecionados = []
if 'bimestres_selecionados' not in st.session_state:
    st.session_state.bimestres_selecionados = bimestres_opcoes.copy() if bimestres_opcoes else []
bimestre_sel = st.sidebar.multiselect(
    "Selecione os bimestres:",
    bimestres_opcoes,
    default=st.session_state.bimestres_selecionados,
    format_func=lambda x: f"{x}º Bimestre",
    help="Filtre os dados por um ou mais bimestres"
)
if bimestre_sel:
    df_temp = df_temp[df_temp["Bimestre"].isin(bimestre_sel)]

turmas = sorted(df_temp["Turma"].dropna().unique().tolist()) if "Turma" in df_temp.columns else []
disciplinas = sorted(df_temp["Disciplina"].dropna().unique().tolist()) if "Disciplina" in df_temp.columns else []
# Detectar coluna de aluno/estudante
coluna_aluno_temp = None
for col in ["Aluno", "Nome_Estudante", "Estudante"]:
    if col in df_temp.columns:
        coluna_aluno_temp = col
        break

alunos = sorted(df_temp[coluna_aluno_temp].dropna().unique().tolist()) if coluna_aluno_temp else []

# Filtros com interface melhorada
st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
    <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">Turmas</h3>
</div>
""", unsafe_allow_html=True)
# Botões de ação rápida para turmas
col_t1, col_t2 = st.sidebar.columns(2)
with col_t1:
    if st.button("Todas", key="btn_todas_turmas", help="Selecionar todas as turmas"):
        st.session_state.turmas_selecionadas = turmas
with col_t2:
    if st.button("Limpar", key="btn_limpar_turmas", help="Limpar seleção"):
        st.session_state.turmas_selecionadas = []

# Inicializar estado se não existir
if 'turmas_selecionadas' not in st.session_state:
    st.session_state.turmas_selecionadas = []

turma_sel = st.sidebar.multiselect(
    "Selecione as turmas:", 
    turmas, 
    default=st.session_state.turmas_selecionadas,
    help="Use os botões acima para seleção rápida"
)

st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
    <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">Disciplinas</h3>
</div>
""", unsafe_allow_html=True)
# Botões de ação rápida para disciplinas
col_d1, col_d2 = st.sidebar.columns(2)
with col_d1:
    if st.button("Todas", key="btn_todas_disc", help="Selecionar todas as disciplinas"):
        st.session_state.disciplinas_selecionadas = disciplinas
with col_d2:
    if st.button("Limpar", key="btn_limpar_disc", help="Limpar seleção"):
        st.session_state.disciplinas_selecionadas = []

# Inicializar estado se não existir
if 'disciplinas_selecionadas' not in st.session_state:
    st.session_state.disciplinas_selecionadas = []

disc_sel = st.sidebar.multiselect(
    "Selecione as disciplinas:", 
    disciplinas, 
    default=st.session_state.disciplinas_selecionadas,
    help="Use os botões acima para seleção rápida"
)

st.sidebar.markdown("""
<div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 6px; padding: 8px 12px; margin: 6px 0; box-shadow: 0 1px 4px rgba(5, 150, 105, 0.1); border-left: 3px solid #059669;">
    <h3 style="color: #047857; margin: 0; font-size: 1em; font-weight: 600;">👤 Aluno</h3>
</div>
""", unsafe_allow_html=True)
aluno_sel = st.sidebar.selectbox("Selecione o aluno:", ["Todos"] + alunos, help="Filtre por aluno específico")

df_filt = df.copy()
if escola_sel != "Todas":
    df_filt = df_filt[df_filt["Escola"] == escola_sel]
if status_sel:  # Se algum status foi selecionado
    df_filt = df_filt[df_filt["Status"].isin(status_sel)]
else:  # Se nenhum status selecionado, mostra todos
    pass  # Mantém todos os status
if bimestre_sel:
    df_filt = df_filt[df_filt["Bimestre"].isin(bimestre_sel)]
if turma_sel:  # Se alguma turma foi selecionada
    df_filt = df_filt[df_filt["Turma"].isin(turma_sel)]
else:  # Se nenhuma turma selecionada, mostra todas
    pass  # Mantém todas as turmas

if disc_sel:  # Se alguma disciplina foi selecionada
    df_filt = df_filt[df_filt["Disciplina"].isin(disc_sel)]
else:  # Se nenhuma disciplina selecionada, mostra todas
    pass  # Mantém todas as disciplinas
if aluno_sel != "Todos":
    df_filt = df_filt[df_filt[coluna_aluno] == aluno_sel]

# Total de Estudantes Únicos (após filtros)
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Total de Estudantes (Filtrado)</h3>
</div>
""", unsafe_allow_html=True)

col_total_filt = st.columns(1)[0]
with col_total_filt:
    total_estudantes_filt = df_filt[coluna_aluno].nunique()
    st.metric(
        label="Estudantes Únicos", 
        value=f"{total_estudantes_filt:,}".replace(",", "."),
        help="Total de estudantes únicos considerando os filtros aplicados"
    )

# Métricas de Frequência na Visão Geral (após filtros)
if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
    st.markdown("""
    <div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
        <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Resumo de Frequência</h3>
    </div>
    """, unsafe_allow_html=True)
    
    colF1, colF2, colF3, colF4, colF5 = st.columns(5)
    
    # Função para classificar frequência (reutilizando a existente)
    def classificar_frequencia_geral(freq):
        if pd.isna(freq):
            return "Sem dados"
        elif freq < 75:
            return "Reprovado"
        elif freq < 80:
            return "Alto Risco"
        elif freq < 90:
            return "Risco Moderado"
        elif freq < 95:
            return "Ponto de Atenção"
        else:
            return "Meta Favorável"
    
    # Calcular frequências para visão geral (usando dados filtrados)
    # Agrupar apenas por Aluno para evitar duplicação quando aluno está em múltiplas turmas
    # Verificar qual coluna de aluno existe
    # Verificar qual coluna de aluno existe
    if "Aluno" in df_filt.columns:
        coluna_aluno = "Aluno"
    elif "Nome_Estudante" in df_filt.columns:
        coluna_aluno = "Nome_Estudante"
    elif "Estudante" in df_filt.columns:
        coluna_aluno = "Estudante"
    else:
        # Tentar encontrar uma coluna que contenha "aluno" ou "estudante"
        colunas_possiveis = [col for col in df_filt.columns if "aluno" in col.lower() or "estudante" in col.lower()]
        if colunas_possiveis:
            coluna_aluno = colunas_possiveis[0]
        else:
            st.error("Não foi possível encontrar uma coluna de aluno/estudante. Colunas disponíveis: " + ", ".join(df_filt.columns))
            st.stop()
    
    if "Frequencia Anual" in df_filt.columns:
        # Agrupar por aluno E turma (igual à tabela detalhada) para garantir consistência
        freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
        freq_geral = freq_geral.rename(columns={"Frequencia Anual": "Frequencia"})
    else:
        # Agrupar por aluno E turma (igual à tabela detalhada) para garantir consistência
        freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
    
    freq_geral["Classificacao_Freq"] = freq_geral["Frequencia"].apply(classificar_frequencia_geral)
    
    # Contar alunos únicos por classificação com priorização
    # Um aluno só é contado na pior categoria que ele possui (para evitar duplicação)
    alunos_por_classificacao = {}
    alunos_ja_contados = set()
    
    # Ordem de prioridade (da pior para a melhor)
    ordem_prioridade = ["Reprovado", "Alto Risco", "Risco Moderado", "Ponto de Atenção", "Meta Favorável"]
    
    for classificacao in ordem_prioridade:
        alunos_na_categoria = set(freq_geral[freq_geral["Classificacao_Freq"] == classificacao][coluna_aluno].unique())
        # Contar apenas alunos que ainda não foram contados em categorias piores
        alunos_novos = alunos_na_categoria - alunos_ja_contados
        alunos_por_classificacao[classificacao] = len(alunos_novos)
        alunos_ja_contados.update(alunos_novos)
    
    contagem_freq_geral = pd.Series(alunos_por_classificacao)
    
    # Calcular total de alunos para porcentagem
    total_alunos_freq = contagem_freq_geral.sum()
    
    with colF1:
        valor_reprovado = contagem_freq_geral.get("Reprovado", 0)
        percent_reprovado = (valor_reprovado / total_alunos_freq * 100) if total_alunos_freq > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
            <div style="font-size: 0.9em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">< 75% (Reprovado)</div>
            <div style="font-size: 1.8em; font-weight: 700; color: #1e40af; margin: 8px 0;">{valor_reprovado}</div>
            <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_reprovado:.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)
    with colF2:
        valor_alto_risco = contagem_freq_geral.get("Alto Risco", 0)
        percent_alto_risco = (valor_alto_risco / total_alunos_freq * 100) if total_alunos_freq > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 10px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
            <div style="font-size: 0.9em; font-weight: 600; color: #0c4a6e; margin-bottom: 8px;">< 80% (Alto Risco)</div>
            <div style="font-size: 1.8em; font-weight: 700; color: #0c4a6e; margin: 8px 0;">{valor_alto_risco}</div>
            <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_alto_risco:.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)
    with colF3:
        valor_risco_moderado = contagem_freq_geral.get("Risco Moderado", 0)
        percent_risco_moderado = (valor_risco_moderado / total_alunos_freq * 100) if total_alunos_freq > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f0f9ff, #dbeafe); border-radius: 10px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(30, 64, 175, 0.15); border-left: 4px solid #1e40af;">
            <div style="font-size: 0.9em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">< 90% (Risco Moderado)</div>
            <div style="font-size: 1.8em; font-weight: 700; color: #1e40af; margin: 8px 0;">{valor_risco_moderado}</div>
            <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_risco_moderado:.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)
    with colF4:
        valor_ponto_atencao = contagem_freq_geral.get("Ponto de Atenção", 0)
        percent_ponto_atencao = (valor_ponto_atencao / total_alunos_freq * 100) if total_alunos_freq > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #eff6ff, #dbeafe); border-radius: 10px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
            <div style="font-size: 0.9em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">< 95% (Ponto Atenção)</div>
            <div style="font-size: 1.8em; font-weight: 700; color: #1e40af; margin: 8px 0;">{valor_ponto_atencao}</div>
            <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_ponto_atencao:.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)
    with colF5:
        valor_meta_favoravel = contagem_freq_geral.get("Meta Favorável", 0)
        percent_meta_favoravel = (valor_meta_favoravel / total_alunos_freq * 100) if total_alunos_freq > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
            <div style="font-size: 0.9em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">≥ 95% (Meta Favorável)</div>
            <div style="font-size: 1.8em; font-weight: 700; color: #1e40af; margin: 8px 0;">{valor_meta_favoravel}</div>
            <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_meta_favoravel:.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)

# -----------------------------
# Indicadores e tabelas de risco
# -----------------------------
indic = calcula_indicadores(df_filt)

# KPIs - Análise de Notas Baixas
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h3 style="color: white; text-align: center; margin: 0; font-size: 1.5em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análise de Notas Abaixo da Média</h3>
    <p style="color: rgba(255,255,255,0.8); margin: 10px 0 0 0; font-size: 1.1em; font-weight: 400;">Análise dos 1º, 2º, 3º e 4º Bimestres</p>
</div>
""", unsafe_allow_html=True)

# Primeira linha: Notas baixas por bimestre
st.markdown("#### 📉 Total de Notas Abaixo de 6 por Bimestre")
col1, col2, col3, col4 = st.columns(4)

notas_baixas_b1 = df_filt[df_filt["Periodo"].str.contains("Primeiro", case=False, na=False) & (df_filt["Nota"] < MEDIA_APROVACAO)]
notas_baixas_b2 = df_filt[df_filt["Periodo"].str.contains("Segundo", case=False, na=False) & (df_filt["Nota"] < MEDIA_APROVACAO)]
notas_baixas_b3 = df_filt[df_filt["Periodo"].str.contains("Terceiro", case=False, na=False) & (df_filt["Nota"] < MEDIA_APROVACAO)]
notas_baixas_b4 = df_filt[df_filt["Periodo"].str.contains("Quarto", case=False, na=False) & (df_filt["Nota"] < MEDIA_APROVACAO)]

# Número de alunos únicos com notas baixas (não disciplinas)
alunos_notas_baixas_b1 = notas_baixas_b1[coluna_aluno].nunique() if coluna_aluno in notas_baixas_b1.columns else 0
alunos_notas_baixas_b2 = notas_baixas_b2[coluna_aluno].nunique() if coluna_aluno in notas_baixas_b2.columns else 0
alunos_notas_baixas_b3 = notas_baixas_b3[coluna_aluno].nunique() if coluna_aluno in notas_baixas_b3.columns else 0
alunos_notas_baixas_b4 = notas_baixas_b4[coluna_aluno].nunique() if coluna_aluno in notas_baixas_b4.columns else 0

# Calcular porcentagens baseadas no total de estudantes filtrados
total_estudantes_para_percent = total_estudantes_filt

with col1:
    percent_notas_b1 = (len(notas_baixas_b1) / len(df_filt) * 100) if len(df_filt) > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #1e40af;">Notas < 6 – 1º Bim</div>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #1e40af; margin: 8px 0;">{len(notas_baixas_b1)}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_notas_b1:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Total de notas abaixo de 6 no 1º bimestre. Inclui todas as disciplinas e alunos.")

with col2:
    percent_notas_b2 = (len(notas_baixas_b2) / len(df_filt) * 100) if len(df_filt) > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #0c4a6e;">Notas < 6 – 2º Bim</div>
            <div style="background: rgba(12, 74, 110, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #0c4a6e;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #0c4a6e; margin: 8px 0;">{len(notas_baixas_b2)}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_notas_b2:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Total de notas abaixo de 6 no 2º bimestre. Inclui todas as disciplinas e alunos.")

with col3:
    percent_notas_b3 = (len(notas_baixas_b3) / len(df_filt) * 100) if len(df_filt) > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #1e40af;">Notas < 6 – 3º Bim</div>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #1e40af; margin: 8px 0;">{len(notas_baixas_b3)}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_notas_b3:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Total de notas abaixo de 6 no 3º bimestre. Inclui todas as disciplinas e alunos.")

with col4:
    percent_notas_b4 = (len(notas_baixas_b4) / len(df_filt) * 100) if len(df_filt) > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #991b1b;">Notas < 6 – 4º Bim</div>
            <div style="background: rgba(153, 27, 27, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #991b1b;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #991b1b; margin: 8px 0;">{len(notas_baixas_b4)}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_notas_b4:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Total de notas abaixo de 6 no 4º bimestre. Inclui todas as disciplinas e alunos.")

# Segunda linha: Alunos únicos com notas baixas
st.markdown("#### 👥 Alunos Únicos com Notas Abaixo de 6")
col5, col6, col7, col8 = st.columns(4)

with col5:
    percent_alunos_b1 = (alunos_notas_baixas_b1 / total_estudantes_para_percent * 100) if total_estudantes_para_percent > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #f0f9ff, #dbeafe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(30, 64, 175, 0.15); border-left: 4px solid #1e40af;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #1e40af;">Alunos < 6 – 1º Bim</div>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #1e40af; margin: 8px 0;">{alunos_notas_baixas_b1}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_alunos_b1:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Número de alunos únicos que tiveram pelo menos uma nota abaixo de 6 no 1º bimestre.")

with col6:
    percent_alunos_b2 = (alunos_notas_baixas_b2 / total_estudantes_para_percent * 100) if total_estudantes_para_percent > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #eff6ff, #dbeafe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #1e40af;">Alunos < 6 – 2º Bim</div>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #1e40af; margin: 8px 0;">{alunos_notas_baixas_b2}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_alunos_b2:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Número de alunos únicos que tiveram pelo menos uma nota abaixo de 6 no 2º bimestre.")

with col7:
    percent_alunos_b3 = (alunos_notas_baixas_b3 / total_estudantes_para_percent * 100) if total_estudantes_para_percent > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #1e40af;">Alunos < 6 – 3º Bim</div>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #1e40af; margin: 8px 0;">{alunos_notas_baixas_b3}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_alunos_b3:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Número de alunos únicos que tiveram pelo menos uma nota abaixo de 6 no 3º bimestre.")

with col8:
    percent_alunos_b4 = (alunos_notas_baixas_b4 / total_estudantes_para_percent * 100) if total_estudantes_para_percent > 0 else 0
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="font-size: 0.95em; font-weight: 600; color: #991b1b;">Alunos < 6 – 4º Bim</div>
            <div style="background: rgba(153, 27, 27, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #991b1b;">?</div>
        </div>
        <div style="font-size: 2em; font-weight: 700; color: #991b1b; margin: 8px 0;">{alunos_notas_baixas_b4}</div>
        <div style="font-size: 1.3em; color: #64748b; font-weight: 600;">({percent_alunos_b4:.1f}%)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip
    st.metric("", "", help="Número de alunos únicos que tiveram pelo menos uma nota abaixo de 6 no 4º bimestre.")

# KPIs - Alertas Críticos (com destaque visual)
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Alertas Críticos</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Situações que precisam de atenção imediata</p>
</div>
""", unsafe_allow_html=True)

col5, col6 = st.columns(2)

# Métricas de alerta com destaque visual (excluindo incompletos)
alerta_count = int(indic[indic["Alerta"] & (indic["Classificacao"] != "Incompleto")].sum()["Alerta"])
corda_bamba_count = int(indic["CordaBamba"].sum())

# Calcular alunos únicos em alerta e corda bamba (excluindo incompletos)
alunos_unicos_alerta = indic[indic["Alerta"] & (indic["Classificacao"] != "Incompleto")][coluna_aluno].nunique()
alunos_unicos_corda_bamba = indic[indic["CordaBamba"]][coluna_aluno].nunique()

with col5:
    st.markdown("""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <h3 style="color: #1e40af; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos-Disciplinas em ALERTA</h3>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div style="font-size: 2.5em; font-weight: 700; color: #1e40af;">{}</div>
            <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">{} alunos</div>
        </div>
    </div>
    """.format(alerta_count, alunos_unicos_alerta), unsafe_allow_html=True)
    
    # Adicionar tooltip funcional
    st.metric("", "", help="Alunos-disciplinas em situação de risco (Vermelho Duplo, Queda p/ Vermelho ou Corda Bamba). O número entre parênteses mostra quantos alunos únicos estão em alerta.")

with col6:
    st.markdown("""
    <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
        <h3 style="color: #0c4a6e; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Corda Bamba</h3>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div style="font-size: 2.5em; font-weight: 700; color: #0c4a6e;">{}</div>
            <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">{} alunos</div>
        </div>
    </div>
    """.format(corda_bamba_count, alunos_unicos_corda_bamba), unsafe_allow_html=True)
    
    # Adicionar tooltip funcional
    st.metric("", "", help="Corda Bamba são alunos que precisam tirar 7 ou mais no próximo bimestre para recuperar e sair do limite da média mínima. O número maior mostra em quantas disciplinas eles aparecem; o número entre parênteses mostra quantos alunos diferentes estão nessa condição.")

# Resumo Executivo - Dashboard Principal
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Resumo Executivo</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Visão consolidada dos principais indicadores</p>
</div>
""", unsafe_allow_html=True)

# Métricas consolidadas em cards
col_res1, col_res2, col_res3, col_res4 = st.columns(4)

with col_res1:
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 8px; padding: 15px; margin: 10px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <h3 style="color: #1e40af; margin: 0 0 5px 0; font-size: 1em; font-weight: 600;">Alertas Críticos</h3>
        <p style="color: #64748b; margin: 0 0 8px 0; font-size: 0.85em;">Situações que precisam de atenção imediata</p>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div style="font-size: 1.5em; font-weight: 700; color: #1e40af;">{alerta_count}</div>
            <div style="font-size: 1.5em; font-weight: 700; color: #64748b;">{alunos_unicos_alerta} alunos</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_res2:
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 8px; padding: 15px; margin: 10px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
        <h3 style="color: #0c4a6e; margin: 0 0 5px 0; font-size: 1em; font-weight: 600;">Corda Bamba</h3>
        <p style="color: #64748b; margin: 0 0 8px 0; font-size: 0.85em;">Precisam de nota ≥ 7 no próximo bimestre</p>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div style="font-size: 1.5em; font-weight: 700; color: #0c4a6e;">{corda_bamba_count}</div>
            <div style="font-size: 1.5em; font-weight: 700; color: #64748b;">{alunos_unicos_corda_bamba} alunos</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_res3:
    # Calcular total de alunos com notas baixas
    total_alunos_notas_baixas = max(alunos_notas_baixas_b1, alunos_notas_baixas_b2, alunos_notas_baixas_b3)
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #f0f9ff, #dbeafe); border-radius: 8px; padding: 15px; margin: 10px 0; box-shadow: 0 2px 8px rgba(30, 64, 175, 0.15); border-left: 4px solid #1e40af;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <h3 style="color: #1e40af; margin: 0; font-size: 1em; font-weight: 600;">Notas Baixas</h3>
            <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
        </div>
        <p style="color: #64748b; margin: 0 0 8px 0; font-size: 0.85em;">Alunos com notas abaixo de 6</p>
        <div style="font-size: 1.5em; font-weight: 700; color: #1e40af;">{total_alunos_notas_baixas}</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Adicionar tooltip usando st.metric
    st.metric("", "", help="Alunos únicos que tiveram pelo menos uma nota abaixo de 6 em qualquer bimestre. Considera o maior número entre 1º, 2º e 3º bimestres.")

with col_res4:
    # Calcular alunos com frequência baixa se disponível
    if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
        if "Frequencia Anual" in df_filt.columns:
            freq_baixa_count = len(df_filt[df_filt["Frequencia Anual"] < 95][coluna_aluno].unique())
        else:
            freq_baixa_count = len(df_filt[df_filt["Frequencia"] < 95][coluna_aluno].unique())
        
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #eff6ff, #dbeafe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                <h3 style="color: #1e40af; margin: 0; font-size: 1.1em; font-weight: 600;">Frequência Baixa</h3>
                <div style="background: rgba(30, 64, 175, 0.1); border-radius: 50%; width: 25px; height: 25px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: #1e40af;">?</div>
            </div>
            <p style="color: #64748b; margin: 0 0 8px 0; font-size: 0.85em;">Alunos com frequência < 95%</p>
            <div style="font-size: 2em; font-weight: 700; color: #1e40af;">{freq_baixa_count}</div>
        </div>
        """, unsafe_allow_html=True)
        
        # Adicionar tooltip usando st.metric
        st.metric("", "", help="Alunos únicos com frequência menor que 95%. Meta favorável é ≥ 95% de frequência.")
    else:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #f8fafc, #e2e8f0); border-radius: 8px; padding: 15px; margin: 10px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.1); border-left: 4px solid #64748b;">
            <h3 style="color: #374151; margin: 0 0 5px 0; font-size: 1em; font-weight: 600;">Frequência</h3>
            <p style="color: #64748b; margin: 0 0 8px 0; font-size: 0.85em;">Dados não disponíveis</p>
            <div style="font-size: 1.5em; font-weight: 700; color: #64748b;">N/A</div>
        </div>
        """, unsafe_allow_html=True)

# KPIs - Análise de Frequência
if "Frequencia Anual" in df_filt.columns:
    freq_title = "Análise de Frequência (Anual)"
    freq_subtitle = "Baseada na frequência anual dos alunos"
elif "Frequencia" in df_filt.columns:
    freq_title = "Análise de Frequência (Por Período)"
    freq_subtitle = "Baseada na frequência por período"
else:
    freq_title = "Análise de Frequência"
    freq_subtitle = "Dados de frequência não disponíveis"

st.markdown(f"""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">{freq_title}</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">{freq_subtitle}</p>
</div>
""", unsafe_allow_html=True)

col7, col8, col9, col10, col11 = st.columns(5)

# Função para classificar frequência
def classificar_frequencia(freq):
    if pd.isna(freq):
        return "Sem dados"
    elif freq < 75:
        return "Reprovado"
    elif freq < 80:
        return "Alto Risco"
    elif freq < 90:
        return "Risco Moderado"
    elif freq < 95:
        return "Ponto de Atenção"
    else:
        return "Meta Favorável"

# Calcular frequências se a coluna existir
if "Frequencia Anual" in df_filt.columns:
    # Usar frequência anual se disponível
    # Agrupar por aluno E turma (igual à tabela detalhada) para garantir consistência
    freq_atual = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
    freq_atual = freq_atual.rename(columns={"Frequencia Anual": "Frequencia"})
    freq_atual["Classificacao_Freq"] = freq_atual["Frequencia"].apply(classificar_frequencia)
    
    # Contar alunos únicos por classificação com priorização
    # Um aluno só é contado na pior categoria que ele possui (para evitar duplicação)
    alunos_por_classificacao = {}
    alunos_ja_contados = set()
    
    # Ordem de prioridade (da pior para a melhor)
    ordem_prioridade = ["Reprovado", "Alto Risco", "Risco Moderado", "Ponto de Atenção", "Meta Favorável"]
    
    for classificacao in ordem_prioridade:
        alunos_na_categoria = set(freq_atual[freq_atual["Classificacao_Freq"] == classificacao][coluna_aluno].unique())
        # Contar apenas alunos que ainda não foram contados em categorias piores
        alunos_novos = alunos_na_categoria - alunos_ja_contados
        alunos_por_classificacao[classificacao] = len(alunos_novos)
        alunos_ja_contados.update(alunos_novos)
    
    contagem_freq = pd.Series(alunos_por_classificacao)
elif "Frequencia" in df_filt.columns:
    # Usar frequência do período se anual não estiver disponível
    # Agrupar por aluno E turma (igual à tabela detalhada) para garantir consistência
    freq_atual = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
    freq_atual["Classificacao_Freq"] = freq_atual["Frequencia"].apply(classificar_frequencia)
    
    # Contar alunos únicos por classificação com priorização
    # Um aluno só é contado na pior categoria que ele possui (para evitar duplicação)
    alunos_por_classificacao = {}
    alunos_ja_contados = set()
    
    # Ordem de prioridade (da pior para a melhor)
    ordem_prioridade = ["Reprovado", "Alto Risco", "Risco Moderado", "Ponto de Atenção", "Meta Favorável"]
    
    for classificacao in ordem_prioridade:
        alunos_na_categoria = set(freq_atual[freq_atual["Classificacao_Freq"] == classificacao][coluna_aluno].unique())
        # Contar apenas alunos que ainda não foram contados em categorias piores
        alunos_novos = alunos_na_categoria - alunos_ja_contados
        alunos_por_classificacao[classificacao] = len(alunos_novos)
        alunos_ja_contados.update(alunos_novos)
    
    contagem_freq = pd.Series(alunos_por_classificacao)
    
    with col7:
        st.metric(
            label="< 75% (Reprovado)", 
            value=contagem_freq.get("Reprovado", 0),
            help="Alunos reprovados por frequência (abaixo de 75%)"
        )
    with col8:
        st.metric(
            label="< 80% (Alto Risco)", 
            value=contagem_freq.get("Alto Risco", 0),
            help="Alunos em alto risco de reprovação por frequência"
        )
    with col9:
        st.metric(
            label="< 90% (Risco Moderado)", 
            value=contagem_freq.get("Risco Moderado", 0),
            help="Alunos com risco moderado de reprovação"
        )
    with col10:
        st.metric(
            label="< 95% (Ponto Atenção)", 
            value=contagem_freq.get("Ponto de Atenção", 0),
            help="Alunos que precisam de atenção na frequência"
        )
    with col11:
        st.metric(
            label="≥ 95% (Meta Favorável)", 
            value=contagem_freq.get("Meta Favorável", 0),
            help="Alunos com frequência dentro da meta"
        )
else:
    col7.metric("< 75% (Reprovado)", "N/A")
    col8.metric("< 80% (Alto Risco)", "N/A")
    col9.metric("< 90% (Risco Moderado)", "N/A")
    col10.metric("< 95% (Ponto Atenção)", "N/A")
    col11.metric("≥ 95% (Meta Favorável)", "N/A")

# Seção expandível: Análise Detalhada de Frequência
if "Frequencia Anual" in df_filt.columns:
    expander_title = "Análise Detalhada de Frequência (Anual)"
elif "Frequencia" in df_filt.columns:
    expander_title = "Análise Detalhada de Frequência (Por Período)"
else:
    expander_title = "Análise Detalhada de Frequência"

with st.expander(expander_title):
    if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
        # Tabela de frequência por aluno e turma (agrupando por aluno e turma para mostrar turmas)
        if "Frequencia Anual" in df_filt.columns:
            freq_detalhada = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
            freq_detalhada = freq_detalhada.rename(columns={"Frequencia Anual": "Frequencia"})
        else:
            freq_detalhada = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
        freq_detalhada["Classificacao_Freq"] = freq_detalhada["Frequencia"].apply(classificar_frequencia)
        freq_detalhada = freq_detalhada.sort_values(coluna_aluno)
        
        # Função para colorir frequência
        def color_frequencia(val):
            if val == "Reprovado":
                return "background-color: #f8d7da; color: #721c24"  # Vermelho
            elif val == "Alto Risco":
                return "background-color: #f5c6cb; color: #721c24"  # Vermelho claro
            elif val == "Risco Moderado":
                return "background-color: #fff3cd; color: #856404"  # Amarelo
            elif val == "Ponto de Atenção":
                return "background-color: #ffeaa7; color: #856404"  # Amarelo claro
            elif val == "Meta Favorável":
                return "background-color: #d4edda; color: #155724"  # Verde
            else:
                return "background-color: #e2e3e5; color: #383d41"  # Cinza
        
        # Formatar frequência
        freq_detalhada["Frequencia_Formatada"] = freq_detalhada["Frequencia"].apply(
            lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
        )
        
        # Aplicar cores
        styled_freq = freq_detalhada[[coluna_aluno, "Turma", "Frequencia_Formatada", "Classificacao_Freq"]]\
            .style.applymap(color_frequencia, subset=["Classificacao_Freq"])
        
        st.dataframe(styled_freq, use_container_width=True)
        
        # Botão de exportação para frequência
        col_export5, col_export6 = st.columns([1, 4])
        with col_export5:
            if st.button("📊 Exportar Frequência", key="export_frequencia", help="Baixar planilha com análise de frequência"):
                excel_data = criar_excel_formatado(freq_detalhada[[coluna_aluno, "Turma", "Frequencia_Formatada", "Classificacao_Freq"]], "Analise_Frequencia")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="analise_frequencia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # Legenda de frequência
        st.markdown("###  Legenda de Frequência")
        col_leg1, col_leg2, col_leg3 = st.columns(3)
        with col_leg1:
            st.markdown("""
            **< 75%**: Reprovado por frequência  
            **< 80%**: Alto risco de reprovação
            """)
        with col_leg2:
            st.markdown("""
            **< 90%**: Risco moderado  
            **< 95%**: Ponto de atenção
            """)
        with col_leg3:
            st.markdown("""
            **≥ 95%**: Meta favorável  
            **Sem dados**: Frequência não informada
            """)
    else:
        st.info("Dados de frequência não disponíveis na planilha.")


st.markdown("---")

# Tabela: Alunos-Disciplinas em ALERTA (com cálculo de necessidade para 3º e 4º)
# Verificar se há dados do 4º bimestre para atualizar o título
tem_n4 = "N4" in indic.columns and indic["N4"].notna().any()
titulo_subtitulo = "Situações que precisam de atenção imediata (Baseado em N1, N2, N3 e N4)" if tem_n4 else "Situações que precisam de atenção imediata (Baseado em N1, N2 e N3)"

st.markdown(f"""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Alunos/Disciplinas em ALERTA</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">{titulo_subtitulo}</p>
</div>
""", unsafe_allow_html=True)

# Definir colunas visíveis - incluir N4 e dados finais se disponíveis
cols_visiveis = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3"]
if "N4" in indic.columns:
    cols_visiveis.append("N4")
cols_visiveis.extend(["Media123"])
if "MediaFinal" in indic.columns:
    cols_visiveis.append("MediaFinal")
cols_visiveis.append("Classificacao")
if "ClassificacaoFinal" in indic.columns:
    cols_visiveis.append("ClassificacaoFinal")
if "StatusFinal" in indic.columns:
    cols_visiveis.append("StatusFinal")
cols_visiveis.extend(["ReqMediaProx1", "CordaBamba"])

# Filtrar apenas colunas que existem no dataframe
cols_visiveis = [c for c in cols_visiveis if c in indic.columns]

# Filtrar alertas excluindo os "Incompleto" (que agora têm seção própria)
tabela_alerta = (indic[indic["Alerta"] & (indic["Classificacao"] != "Incompleto")]
                 .copy()
                 .sort_values(["Turma", coluna_aluno, "Disciplina"]))

# Formatar colunas numéricas
colunas_para_formatar = ["N1", "N2", "N3", "Media123", "ReqMediaProx1"]
if "N4" in tabela_alerta.columns:
    colunas_para_formatar.append("N4")
if "MediaFinal" in tabela_alerta.columns:
    colunas_para_formatar.append("MediaFinal")

for c in colunas_para_formatar:
    if c in tabela_alerta.columns:
        # Formatar para 1 casa decimal, removendo .0 desnecessário
        tabela_alerta[c] = tabela_alerta[c].round(1)
        tabela_alerta[c] = tabela_alerta[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)

# Função para aplicar cores na classificação (definida antes de usar)
def color_classification(val):
    if val == "Verde":
        return "background-color: #10b981; color: white; font-weight: bold;"  # Verde forte
    elif val == "Vermelho Quádruplo":
        return "background-color: #7f1d1d; color: white; font-weight: bold;"  # Vermelho muito escuro (mais escuro que triplo)
    elif val == "Vermelho Triplo":
        return "background-color: #991b1b; color: white; font-weight: bold;"  # Vermelho muito escuro
    elif val == "Vermelho Duplo":
        return "background-color: #dc2626; color: white; font-weight: bold;"  # Vermelho forte
    elif val == "Queda p/ Vermelho" or val == "Queda Recente" or val == "Queda Final":
        return "background-color: #f59e0b; color: white; font-weight: bold;"  # Laranja forte
    elif val == "Recuperou" or val == "Recuperação" or val == "Recuperação Final":
        return "background-color: #3b82f6; color: white; font-weight: bold;"  # Azul forte
    elif val == "Incompleto":
        return "background-color: #6b7280; color: white; font-weight: bold;"  # Cinza forte
    else:
        return ""

# Função para aplicar cores no status final (Aprovado/Reprovado)
def color_status_final(val):
    if val == "Aprovado":
        return "background-color: #10b981; color: white; font-weight: bold;"  # Verde forte
    elif val == "Reprovado":
        return "background-color: #dc2626; color: white; font-weight: bold;"  # Vermelho forte
    elif val == "Incompleto":
        return "background-color: #6b7280; color: white; font-weight: bold;"  # Cinza forte
    else:
        return ""

# Aplicar cores na tabela de alertas também
if len(tabela_alerta) > 0:
    # Determinar qual coluna de classificação usar
    col_classificacao = "ClassificacaoFinal" if "ClassificacaoFinal" in tabela_alerta.columns and tabela_alerta["ClassificacaoFinal"].notna().any() else "Classificacao"
    
    # Aplicar estilização
    styled_alerta = tabela_alerta[cols_visiveis].style.applymap(color_classification, subset=[col_classificacao])
    
    # Se houver StatusFinal, aplicar cores também (usar applymap novamente no styled já criado)
    if "StatusFinal" in tabela_alerta[cols_visiveis].columns:
        styled_alerta = styled_alerta.applymap(color_status_final, subset=["StatusFinal"])
    
    st.dataframe(styled_alerta, use_container_width=True)
    
    # Botão de exportação para alertas
    col_export1, col_export2 = st.columns([1, 4])
    with col_export1:
        if st.button("📊 Exportar Alertas", key="export_alertas", help="Baixar planilha com alunos em alerta"):
            excel_data = criar_excel_formatado(tabela_alerta[cols_visiveis], "Alunos_em_Alerta")
            st.download_button(
                label="Baixar Excel",
                data=excel_data,
                file_name="alunos_em_alerta.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.dataframe(pd.DataFrame(columns=cols_visiveis), use_container_width=True)

# Seção separada para alunos com status "Incompleto" - Separada por Bimestres
st.markdown("""
<div style="background: linear-gradient(135deg, #6b7280, #9ca3af); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(107, 114, 128, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Alunos/Disciplinas INCOMPLETAS</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Faltam notas para completar a avaliação - Separado por Bimestres</p>
</div>
""", unsafe_allow_html=True)

# Filtrar apenas os incompletos
incompletos = indic[indic["Classificacao"] == "Incompleto"].copy()

if len(incompletos) > 0:
    # Separar incompletos por bimestres
    # Incompletos do 1º bimestre: falta N1
    incompletos_b1 = incompletos[pd.isna(incompletos["N1"])].copy()
    
    # Incompletos do 2º bimestre: falta N2
    incompletos_b2 = incompletos[pd.isna(incompletos["N2"])].copy()
    
    # Incompletos do 3º bimestre: falta N3
    incompletos_b3 = incompletos[pd.isna(incompletos["N3"])].copy()
    
    # Incompletos do 4º bimestre: falta N4 (se a coluna existir)
    if "N4" in incompletos.columns:
        incompletos_b4 = incompletos[pd.isna(incompletos["N4"])].copy()
    else:
        incompletos_b4 = pd.DataFrame()
    
    # Criar abas para cada bimestre
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Resumo Geral", "1️⃣ 1º Bimestre", "2️⃣ 2º Bimestre", "3️⃣ 3º Bimestre", "4️⃣ 4º Bimestre"])
    
    with tab1:
        # Estatísticas gerais dos incompletos
        total_incompletos = len(incompletos)
        alunos_unicos_incompletos = incompletos[coluna_aluno].nunique()
        total_b1 = len(incompletos_b1)
        total_b2 = len(incompletos_b2)
        total_b3 = len(incompletos_b3)
        total_b4 = len(incompletos_b4) if len(incompletos_b4) > 0 else 0
        alunos_b1 = incompletos_b1[coluna_aluno].nunique() if len(incompletos_b1) > 0 else 0
        alunos_b2 = incompletos_b2[coluna_aluno].nunique() if len(incompletos_b2) > 0 else 0
        alunos_b3 = incompletos_b3[coluna_aluno].nunique() if len(incompletos_b3) > 0 else 0
        alunos_b4 = incompletos_b4[coluna_aluno].nunique() if len(incompletos_b4) > 0 else 0
        
        # Primeira linha: Resumo geral
        col_gen1, col_gen2 = st.columns(2)
        
        with col_gen1:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Total Incompletas</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #374151;">{total_incompletos}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #6b7280;">disciplinas</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_gen2:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos Afetados</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #374151;">{alunos_unicos_incompletos}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #6b7280;">alunos</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Segunda linha: Detalhamento por bimestre
        st.markdown("#### 📊 Distribuição por Bimestre")
        col_gen3, col_gen4, col_gen5, col_gen6 = st.columns(4)
        
        with col_gen3:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
                <h3 style="color: #1e40af; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Falta 1º Bimestre</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #1e40af;">{total_b1}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #64748b;">disciplinas</div>
                </div>
                <div style="font-size: 0.9em; color: #1e40af; margin-top: 5px;">({alunos_b1} alunos)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_gen4:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
                <h3 style="color: #0c4a6e; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Falta 2º Bimestre</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #0c4a6e;">{total_b2}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #64748b;">disciplinas</div>
                </div>
                <div style="font-size: 0.9em; color: #0c4a6e; margin-top: 5px;">({alunos_b2} alunos)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_gen5:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
                <h3 style="color: #1e40af; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Falta 3º Bimestre</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #1e40af;">{total_b3}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #64748b;">disciplinas</div>
                </div>
                <div style="font-size: 0.9em; color: #1e40af; margin-top: 5px;">({alunos_b3} alunos)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_gen6:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
                <h3 style="color: #991b1b; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Falta 4º Bimestre</h3>
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 2.2em; font-weight: 700; color: #991b1b;">{total_b4}</div>
                    <div style="font-size: 1.8em; font-weight: 700; color: #64748b;">disciplinas</div>
                </div>
                <div style="font-size: 0.9em; color: #991b1b; margin-top: 5px;">({alunos_b4} alunos)</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Tabela geral de incompletos
        st.markdown("### 📋 Todos os Incompletos")
        incompletos_ordenados = incompletos.sort_values(["Turma", coluna_aluno, "Disciplina"])
        
        # Formatar colunas numéricas
        colunas_para_formatar = ["N1", "N2", "N3", "Media123", "ReqMediaProx1"]
        if "N4" in incompletos_ordenados.columns:
            colunas_para_formatar.append("N4")
        if "MediaFinal" in incompletos_ordenados.columns:
            colunas_para_formatar.append("MediaFinal")
        
        for c in colunas_para_formatar:
            if c in incompletos_ordenados.columns:
                if pd.api.types.is_numeric_dtype(incompletos_ordenados[c]):
                    incompletos_ordenados[c] = incompletos_ordenados[c].round(1)
                    incompletos_ordenados[c] = incompletos_ordenados[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)
        
        # Adicionar coluna indicando qual bimestre falta
        def identificar_bimestre_faltante(row):
            if pd.isna(row.get("N1", None)):
                return "1º Bimestre"
            elif pd.isna(row.get("N2", None)):
                return "2º Bimestre"
            elif pd.isna(row.get("N3", None)):
                return "3º Bimestre"
            elif "N4" in row.index and pd.isna(row.get("N4", None)):
                return "4º Bimestre"
            else:
                return "N/A"
        
        incompletos_ordenados["Falta"] = incompletos_ordenados.apply(identificar_bimestre_faltante, axis=1)
        
        # Definir colunas da tabela geral - incluir N4 se disponível
        cols_incompletos_geral = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3"]
        if "N4" in incompletos_ordenados.columns:
            cols_incompletos_geral.append("N4")
        cols_incompletos_geral.extend(["Falta", "Classificacao"])
        if "StatusFinal" in incompletos_ordenados.columns:
            cols_incompletos_geral.append("StatusFinal")
        styled_incompletos_geral = incompletos_ordenados[cols_incompletos_geral].style.applymap(color_classification, subset=["Classificacao"])
        st.dataframe(styled_incompletos_geral, use_container_width=True)
        
        # Botão de exportação geral
        col_export_gen1, col_export_gen2 = st.columns([1, 4])
        with col_export_gen1:
            if st.button("📋 Exportar Todos", key="export_incompletos_geral", help="Baixar planilha com todos os incompletos"):
                excel_data = criar_excel_formatado(incompletos_ordenados[cols_incompletos_geral], "Todos_Incompletos")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="todos_incompletos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    with tab2:
        # Aba do 1º Bimestre
        st.markdown("### 1️⃣ Incompletos do 1º Bimestre (Falta N1)")
        
        if len(incompletos_b1) > 0:
            # Estatísticas específicas do 1º bimestre
            col_b1_1, col_b1_2 = st.columns(2)
            
            with col_b1_1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                    <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Disciplinas Incompletas</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #374151;">{total_b1}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #6b7280;">disciplinas</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_b1_2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                    <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos Afetados</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #374151;">{alunos_b1}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #6b7280;">alunos</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            # Ordenar e formatar dados do 1º bimestre
            incompletos_b1_ordenados = incompletos_b1.sort_values(["Turma", coluna_aluno, "Disciplina"])
            
            # Formatar colunas numéricas
            for c in ["N1", "N2", "Media12", "ReqMediaProx2"]:
                if c in incompletos_b1_ordenados.columns:
                    incompletos_b1_ordenados[c] = incompletos_b1_ordenados[c].round(1)
                    incompletos_b1_ordenados[c] = incompletos_b1_ordenados[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)
            
            # Mostrar tabela do 1º bimestre
            cols_incompletos_b1 = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "Media12", "Classificacao"]
            styled_incompletos_b1 = incompletos_b1_ordenados[cols_incompletos_b1].style.applymap(color_classification, subset=["Classificacao"])
            st.dataframe(styled_incompletos_b1, use_container_width=True)
            
            # Botão de exportação do 1º bimestre
            col_export_b1_1, col_export_b1_2 = st.columns([1, 4])
            with col_export_b1_1:
                if st.button("📋 Exportar 1º Bimestre", key="export_incompletos_b1", help="Baixar planilha com incompletos do 1º bimestre"):
                    excel_data = criar_excel_formatado(incompletos_b1_ordenados[cols_incompletos_b1], "Incompletos_1_Bimestre")
                    st.download_button(
                        label="Baixar Excel",
                        data=excel_data,
                        file_name="incompletos_1_bimestre.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.success("✅ Nenhum aluno com notas incompletas do 1º bimestre.")
    
    with tab3:
        # Aba do 2º Bimestre
        st.markdown("### 2️⃣ Incompletos do 2º Bimestre (Falta N2)")
        
        if len(incompletos_b2) > 0:
            # Estatísticas específicas do 2º bimestre
            col_b2_1, col_b2_2 = st.columns(2)
            
            with col_b2_1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                    <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Disciplinas Incompletas</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #374151;">{total_b2}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #6b7280;">disciplinas</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_b2_2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
                    <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos Afetados</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #374151;">{alunos_b2}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #6b7280;">alunos</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            # Ordenar e formatar dados do 2º bimestre
            incompletos_b2_ordenados = incompletos_b2.sort_values(["Turma", coluna_aluno, "Disciplina"])
            
            # Formatar colunas numéricas
            for c in ["N1", "N2", "Media12", "ReqMediaProx2"]:
                if c in incompletos_b2_ordenados.columns:
                    incompletos_b2_ordenados[c] = incompletos_b2_ordenados[c].round(1)
                    incompletos_b2_ordenados[c] = incompletos_b2_ordenados[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)
            
            # Mostrar tabela do 2º bimestre
            cols_incompletos_b2 = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "Media12", "Classificacao"]
            styled_incompletos_b2 = incompletos_b2_ordenados[cols_incompletos_b2].style.applymap(color_classification, subset=["Classificacao"])
            st.dataframe(styled_incompletos_b2, use_container_width=True)
            
            # Botão de exportação do 2º bimestre
            col_export_b2_1, col_export_b2_2 = st.columns([1, 4])
            with col_export_b2_1:
                if st.button("📋 Exportar 2º Bimestre", key="export_incompletos_b2", help="Baixar planilha com incompletos do 2º bimestre"):
                    excel_data = criar_excel_formatado(incompletos_b2_ordenados[cols_incompletos_b2], "Incompletos_2_Bimestre")
                    st.download_button(
                        label="Baixar Excel",
                        data=excel_data,
                        file_name="incompletos_2_bimestre.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.success("✅ Nenhum aluno com notas incompletas do 2º bimestre.")
    
    with tab4:
        # Aba do 3º Bimestre
        st.markdown("### 3️⃣ Incompletos do 3º Bimestre (Falta N3)")
        
        if len(incompletos_b3) > 0:
            # Estatísticas específicas do 3º bimestre
            col_b3_1, col_b3_2 = st.columns(2)
            
            with col_b3_1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fef3c7, #fde68a); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(245, 158, 11, 0.15); border-left: 4px solid #f59e0b;">
                    <h3 style="color: #92400e; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Disciplinas Incompletas</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #92400e;">{total_b3}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">disciplinas</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_b3_2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fef3c7, #fde68a); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(245, 158, 11, 0.15); border-left: 4px solid #f59e0b;">
                    <h3 style="color: #92400e; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos Afetados</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #92400e;">{alunos_b3}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">alunos</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            # Ordenar e formatar dados do 3º bimestre
            incompletos_b3_ordenados = incompletos_b3.sort_values(["Turma", coluna_aluno, "Disciplina"])
            
            # Formatar colunas numéricas
            for c in ["N1", "N2", "N3", "Media123", "ReqMediaProx1"]:
                if c in incompletos_b3_ordenados.columns:
                    incompletos_b3_ordenados[c] = incompletos_b3_ordenados[c].round(1)
                    incompletos_b3_ordenados[c] = incompletos_b3_ordenados[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)
            
            # Mostrar tabela do 3º bimestre
            cols_incompletos_b3 = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3", "Media123", "Classificacao"]
            styled_incompletos_b3 = incompletos_b3_ordenados[cols_incompletos_b3].style.applymap(color_classification, subset=["Classificacao"])
            st.dataframe(styled_incompletos_b3, use_container_width=True)
            
            # Botão de exportação do 3º bimestre
            col_export_b3_1, col_export_b3_2 = st.columns([1, 4])
            with col_export_b3_1:
                if st.button("📋 Exportar 3º Bimestre", key="export_incompletos_b3", help="Baixar planilha com incompletos do 3º bimestre"):
                    excel_data = criar_excel_formatado(incompletos_b3_ordenados[cols_incompletos_b3], "Incompletos_3_Bimestre")
                    st.download_button(
                        label="Baixar Excel",
                        data=excel_data,
                        file_name="incompletos_3_bimestre.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.success("✅ Nenhum aluno com notas incompletas do 3º bimestre.")
    
    with tab5:
        # Aba do 4º Bimestre
        st.markdown("### 4️⃣ Incompletos do 4º Bimestre (Falta N4)")
        
        if len(incompletos_b4) > 0:
            # Estatísticas específicas do 4º bimestre
            col_b4_1, col_b4_2 = st.columns(2)
            
            with col_b4_1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
                    <h3 style="color: #374151; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Disciplinas Incompletas</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #991b1b;">{len(incompletos_b4)}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">disciplinas</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            with col_b4_2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 18px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
                    <h3 style="color: #991b1b; margin: 0 0 15px 0; font-size: 1.1em; font-weight: 600;">Alunos Afetados</h3>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div style="font-size: 2.5em; font-weight: 700; color: #991b1b;">{alunos_b4}</div>
                        <div style="font-size: 2.5em; font-weight: 700; color: #64748b;">alunos</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            # Ordenar e formatar dados do 4º bimestre
            incompletos_b4_ordenados = incompletos_b4.sort_values(["Turma", coluna_aluno, "Disciplina"])
            
            # Formatar colunas numéricas
            colunas_para_formatar_b4 = ["N1", "N2", "N3", "Media123", "ReqMediaProx1"]
            if "N4" in incompletos_b4_ordenados.columns:
                colunas_para_formatar_b4.append("N4")
            if "MediaFinal" in incompletos_b4_ordenados.columns:
                colunas_para_formatar_b4.append("MediaFinal")
            
            for c in colunas_para_formatar_b4:
                if c in incompletos_b4_ordenados.columns:
                    if pd.api.types.is_numeric_dtype(incompletos_b4_ordenados[c]):
                        incompletos_b4_ordenados[c] = incompletos_b4_ordenados[c].round(1)
                        incompletos_b4_ordenados[c] = incompletos_b4_ordenados[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)
            
            # Mostrar tabela do 4º bimestre
            cols_incompletos_b4 = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3"]
            if "N4" in incompletos_b4_ordenados.columns:
                cols_incompletos_b4.append("N4")
            cols_incompletos_b4.extend(["Media123"])
            if "MediaFinal" in incompletos_b4_ordenados.columns:
                cols_incompletos_b4.append("MediaFinal")
            cols_incompletos_b4.append("Classificacao")
            if "ClassificacaoFinal" in incompletos_b4_ordenados.columns:
                cols_incompletos_b4.append("ClassificacaoFinal")
            if "StatusFinal" in incompletos_b4_ordenados.columns:
                cols_incompletos_b4.append("StatusFinal")
            
            # Filtrar apenas colunas que existem
            cols_incompletos_b4 = [c for c in cols_incompletos_b4 if c in incompletos_b4_ordenados.columns]
            
            col_classificacao_b4 = "ClassificacaoFinal" if "ClassificacaoFinal" in incompletos_b4_ordenados.columns and incompletos_b4_ordenados["ClassificacaoFinal"].notna().any() else "Classificacao"
            styled_incompletos_b4 = incompletos_b4_ordenados[cols_incompletos_b4].style.applymap(color_classification, subset=[col_classificacao_b4])
            
            if "StatusFinal" in cols_incompletos_b4:
                styled_incompletos_b4 = styled_incompletos_b4.applymap(color_status_final, subset=["StatusFinal"])
            
            st.dataframe(styled_incompletos_b4, use_container_width=True)
            
            # Botão de exportação do 4º bimestre
            col_export_b4_1, col_export_b4_2 = st.columns([1, 4])
            with col_export_b4_1:
                if st.button("📋 Exportar 4º Bimestre", key="export_incompletos_b4", help="Baixar planilha com incompletos do 4º bimestre"):
                    excel_data = criar_excel_formatado(incompletos_b4_ordenados[cols_incompletos_b4], "Incompletos_4_Bimestre")
                    st.download_button(
                        label="Baixar Excel",
                        data=excel_data,
                        file_name="incompletos_4_bimestre.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.success("✅ Nenhum aluno com notas incompletas do 4º bimestre.")

else:
    st.info("✅ Nenhum aluno com disciplinas incompletas encontrado.")

# Seção Consolidada: Resumo por Bimestres
st.markdown("""
<div style="background: linear-gradient(135deg, #7c3aed, #a855f7); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(124, 58, 237, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">📊 Resumo dos Problemas por Bimestre</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Estudantes únicos com problemas por período</p>
</div>
""", unsafe_allow_html=True)

# Calcular estudantes únicos por bimestre
alunos_notas_baixas_b1_unicos = set()
alunos_notas_baixas_b2_unicos = set()
alunos_notas_baixas_b3_unicos = set()
alunos_notas_baixas_b4_unicos = set()

if len(notas_baixas_b1) > 0:
    alunos_notas_baixas_b1_unicos = set(notas_baixas_b1[coluna_aluno].unique())
if len(notas_baixas_b2) > 0:
    alunos_notas_baixas_b2_unicos = set(notas_baixas_b2[coluna_aluno].unique())
if len(notas_baixas_b3) > 0:
    alunos_notas_baixas_b3_unicos = set(notas_baixas_b3[coluna_aluno].unique())
if len(notas_baixas_b4) > 0:
    alunos_notas_baixas_b4_unicos = set(notas_baixas_b4[coluna_aluno].unique())

alunos_incompletos_b1_unicos = set()
alunos_incompletos_b2_unicos = set()
alunos_incompletos_b3_unicos = set()
alunos_incompletos_b4_unicos = set()

if len(incompletos) > 0:
    if len(incompletos_b1) > 0:
        alunos_incompletos_b1_unicos = set(incompletos_b1[coluna_aluno].unique())
    if len(incompletos_b2) > 0:
        alunos_incompletos_b2_unicos = set(incompletos_b2[coluna_aluno].unique())
    if len(incompletos_b3) > 0:
        alunos_incompletos_b3_unicos = set(incompletos_b3[coluna_aluno].unique())
    if len(incompletos_b4) > 0:
        alunos_incompletos_b4_unicos = set(incompletos_b4[coluna_aluno].unique())

# Criar seção por bimestres
st.markdown("### 📋 Resumo por Bimestre")

col_bim1, col_bim2, col_bim3, col_bim4 = st.columns(4)

with col_bim1:
    st.markdown("#### 1º Bimestre")
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #f3f4f6, #e5e7eb); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(107, 114, 128, 0.15); border-left: 4px solid #6b7280;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #374151; font-weight: 600;">Notas Baixas:</span>
            <span style="color: #374151; font-weight: 700; font-size: 1.2em;">{len(alunos_notas_baixas_b1_unicos)} alunos</span>
        </div>
        <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 8px;">
            <span style="color: #374151; font-weight: 600;">Incompletos:</span>
            <span style="color: #374151; font-weight: 700; font-size: 1.2em;">{len(alunos_incompletos_b1_unicos)} alunos</span>
        </div>
        <div style="border-top: 1px solid #d1d5db; margin-top: 10px; padding-top: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <span style="color: #374151; font-weight: 700;">Total 1º Bimestre:</span>
                <div style="text-align: right;">
                    <span style="color: #374151; font-weight: 700; font-size: 1.3em;">{len(alunos_notas_baixas_b1_unicos | alunos_incompletos_b1_unicos)} alunos</span>
                    <div style="color: #6b7280; font-size: 0.9em; font-weight: 600;">
                        ({(len(alunos_notas_baixas_b1_unicos | alunos_incompletos_b1_unicos) / df_filt[coluna_aluno].nunique() * 100):.1f}% do total)
                    </div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_bim2:
    st.markdown("#### 2º Bimestre")
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #e0f2fe, #b3e5fc); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(14, 165, 233, 0.15); border-left: 4px solid #0ea5e9;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #0c4a6e; font-weight: 600;">Notas Baixas:</span>
            <span style="color: #0c4a6e; font-weight: 700; font-size: 1.2em;">{len(alunos_notas_baixas_b2_unicos)} alunos</span>
        </div>
        <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 8px;">
            <span style="color: #0c4a6e; font-weight: 600;">Incompletos:</span>
            <span style="color: #0c4a6e; font-weight: 700; font-size: 1.2em;">{len(alunos_incompletos_b2_unicos)} alunos</span>
        </div>
        <div style="border-top: 1px solid #7dd3fc; margin-top: 10px; padding-top: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <span style="color: #0c4a6e; font-weight: 700;">Total 2º Bimestre:</span>
                <div style="text-align: right;">
                    <span style="color: #0c4a6e; font-weight: 700; font-size: 1.3em;">{len(alunos_notas_baixas_b2_unicos | alunos_incompletos_b2_unicos)} alunos</span>
                    <div style="color: #64748b; font-size: 0.9em; font-weight: 600;">
                        ({(len(alunos_notas_baixas_b2_unicos | alunos_incompletos_b2_unicos) / df_filt[coluna_aluno].nunique() * 100):.1f}% do total)
                    </div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_bim3:
    st.markdown("#### 3º Bimestre")
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #1e40af; font-weight: 600;">Notas Baixas:</span>
            <span style="color: #1e40af; font-weight: 700; font-size: 1.2em;">{len(alunos_notas_baixas_b3_unicos)} alunos</span>
        </div>
        <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 8px;">
            <span style="color: #1e40af; font-weight: 600;">Incompletos:</span>
            <span style="color: #1e40af; font-weight: 700; font-size: 1.2em;">{len(alunos_incompletos_b3_unicos)} alunos</span>
        </div>
        <div style="border-top: 1px solid #93c5fd; margin-top: 10px; padding-top: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <span style="color: #1e40af; font-weight: 700;">Total 3º Bimestre:</span>
                <div style="text-align: right;">
                    <span style="color: #1e40af; font-weight: 700; font-size: 1.3em;">{len(alunos_notas_baixas_b3_unicos | alunos_incompletos_b3_unicos)} alunos</span>
                    <div style="color: #64748b; font-size: 0.9em; font-weight: 600;">
                        ({(len(alunos_notas_baixas_b3_unicos | alunos_incompletos_b3_unicos) / df_filt[coluna_aluno].nunique() * 100):.1f}% do total)
                    </div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_bim4:
    st.markdown("#### 4º Bimestre")
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 8px; padding: 15px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #991b1b; font-weight: 600;">Notas Baixas:</span>
            <span style="color: #991b1b; font-weight: 700; font-size: 1.2em;">{len(alunos_notas_baixas_b4_unicos)} alunos</span>
        </div>
        <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 8px;">
            <span style="color: #991b1b; font-weight: 600;">Incompletos:</span>
            <span style="color: #991b1b; font-weight: 700; font-size: 1.2em;">{len(alunos_incompletos_b4_unicos)} alunos</span>
        </div>
        <div style="border-top: 1px solid #fca5a5; margin-top: 10px; padding-top: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <span style="color: #991b1b; font-weight: 700;">Total 4º Bimestre:</span>
                <div style="text-align: right;">
                    <span style="color: #991b1b; font-weight: 700; font-size: 1.3em;">{len(alunos_notas_baixas_b4_unicos | alunos_incompletos_b4_unicos)} alunos</span>
                    <div style="color: #dc2626; font-size: 0.9em; font-weight: 600;">
                        ({(len(alunos_notas_baixas_b4_unicos | alunos_incompletos_b4_unicos) / df_filt[coluna_aluno].nunique() * 100):.1f}% do total)
                    </div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# Tabela: Panorama Geral de Notas (todos para diagnóstico rápido)
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Panorama Geral de Notas (B1→B2→B3→B4)</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Visão completa de todos os alunos e disciplinas</p>
</div>
""", unsafe_allow_html=True)
tab_diag = indic.copy()
# Colunas numéricas que devem ser formatadas
colunas_numericas = ["N1", "N2", "N3", "N4", "Media123", "MediaFinal", "ReqMediaProx1", "Soma123", "SomaFinal"]
for c in colunas_numericas:
    if c in tab_diag.columns:
        # Verificar se a coluna é numérica antes de arredondar
        if pd.api.types.is_numeric_dtype(tab_diag[c]):
            # Formatar para 1 casa decimal, removendo .0 desnecessário
            tab_diag[c] = tab_diag[c].round(1)
            tab_diag[c] = tab_diag[c].apply(lambda x: f"{x:.1f}".rstrip('0').rstrip('.') if pd.notna(x) else x)



# Aplicar estilização - incluir colunas do 4º bimestre se disponíveis
colunas_tabela = [coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3"]
if "N4" in tab_diag.columns:
    colunas_tabela.append("N4")
colunas_tabela.extend(["Media123"])
if "MediaFinal" in tab_diag.columns:
    colunas_tabela.append("MediaFinal")
colunas_tabela.append("Classificacao")
if "ClassificacaoFinal" in tab_diag.columns:
    colunas_tabela.append("ClassificacaoFinal")
if "StatusFinal" in tab_diag.columns:
    colunas_tabela.append("StatusFinal")
colunas_tabela.append("ReqMediaProx1")

# Filtrar apenas colunas que existem no dataframe
colunas_tabela = [c for c in colunas_tabela if c in tab_diag.columns]

styled_table = tab_diag[colunas_tabela]\
    .sort_values(["Turma", coluna_aluno, "Disciplina"])\
    .style.applymap(color_classification, subset=["Classificacao"])

st.dataframe(styled_table, use_container_width=True)

# Botão de exportação para panorama de notas
col_export3, col_export4 = st.columns([1, 4])
with col_export3:
        if st.button("📊 Exportar Panorama", key="export_panorama", help="Baixar planilha com panorama geral de notas"):
            excel_data = criar_excel_formatado(tab_diag[[coluna_aluno, "Turma", "Disciplina", "N1", "N2", "N3", "Media123", "Classificacao", "ReqMediaProx1"]], "Panorama_Geral_Notas")
            st.download_button(
                label="Baixar Excel",
                data=excel_data,
                file_name="panorama_notas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# Legenda de cores
st.markdown("### 🎨 Legenda de Cores")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("""
    <div style="background-color: #10b981; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🟢 Verde: Todas as notas ≥6
    </div>
    <div style="background-color: #dc2626; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🔴 Vermelho Triplo: Risco crítico (N1, N2 e N3 < 6)
    </div>
    <div style="background-color: #ef4444; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🔴 Vermelho Duplo: Risco alto (duas notas < 6)
    </div>
    """, unsafe_allow_html=True)
with col2:
    st.markdown("""
    <div style="background-color: #f59e0b; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🟠 Queda Recente: Caiu no 3º bimestre
    </div>
    <div style="background-color: #3b82f6; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🔵 Recuperação: Melhorou no 3º bimestre
    </div>
    """, unsafe_allow_html=True)
with col3:
    st.markdown("""
    <div style="background-color: #6b7280; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        ⚪ Incompleto: Falta nota de algum bimestre
    </div>
    <div style="background-color: #8b5cf6; color: white; padding: 8px; border-radius: 5px; margin: 5px 0; font-weight: bold; text-align: center;">
        🟣 Corda Bamba: Precisa ≥7 no próximo
    </div>
    """, unsafe_allow_html=True)

st.markdown(
    """
    **Interpretação rápida (3 bimestres)**  
    - *Vermelho Triplo*: situação crítica - todas as notas abaixo de 6 (N1, N2 e N3).  
    - *Vermelho Duplo*: risco alto - duas das três notas abaixo de 6.  
    - *Queda Recente*: estava indo bem mas caiu no 3º bimestre - atenção!  
    - *Recuperação*: estava com dificuldade mas melhorou no 3º bimestre.  
    - *Corda Bamba*: para fechar média 6 no ano, precisa tirar **≥ 7,0** no 4º bimestre.
    
    **Interpretação rápida (4 bimestres - Conceito Final)**  
    - *Vermelho Quádruplo*: situação crítica - todas as 4 notas abaixo de 6.  
    - *Vermelho Triplo*: três notas abaixo de 6.  
    - *Vermelho Duplo*: duas notas abaixo de 6.  
    - *Queda Final*: estava indo bem mas caiu no 4º bimestre.  
    - *Recuperação Final*: estava com dificuldade mas melhorou no 4º bimestre.  
    - *Verde*: todas as notas ≥ 6.  
    - *Aprovado*: média final ≥ 6.0 (soma dos 4 bimestres ≥ 24).  
    - *Reprovado*: média final < 6.0 (soma dos 4 bimestres < 24).
    """
)

# Gráficos: Notas e Frequência por Disciplina (movidos para o final)
st.markdown("---")
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análises Gráficas</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Visualizações complementares dos dados</p>
</div>
""", unsafe_allow_html=True)

# Gráfico de Evolução das Notas por Turma ao Longo dos Bimestres
st.markdown("### 📈 Evolução das Notas das Turmas ao Longo dos Bimestres")

# Explicação do formato dos códigos das turmas
with st.expander("ℹ️ Como ler os códigos das turmas?", expanded=False):
    st.markdown("""
    **Formato dos códigos das turmas:**
    
    Os códigos seguem o padrão: **`XX.XX/NIVEL.MODALIDADE-TURNO`**
    
    **Exemplo:** `62.01/EF.MAT-ANL`
    - **62.01**: Código identificador da turma (pode incluir código da escola)
    - **EF**: Nível de ensino - Ensino Fundamental (ou **EM** para Ensino Médio)
    - **MAT**: Pode ser modalidade, disciplina ou turno (ex: MAT = Matutino ou Matemática)
    - **ANL**: Período letivo (ANL = Anual)
    
    **Exemplo:** `82.02/EF.VESP-ANL`
    - **82.02**: Código identificador da turma
    - **EF**: Ensino Fundamental
    - **VESP**: Pode ser modalidade ou turno (VESP = Vespertino)
    - **ANL**: Anual
    
    ⚠️ **Nota:** O formato exato pode variar conforme o sistema de origem dos dados. 
    Os códigos são gerados automaticamente pelo sistema educacional e identificam 
    unicamente cada turma no banco de dados.
    
    💡 **Dica:** Cada linha colorida no gráfico representa uma turma diferente. 
    Você pode passar o mouse sobre as linhas para ver os valores exatos de cada bimestre.
    """)

# Criar coluna Bimestre no df_filt se ainda não existir
if "Bimestre" not in df_filt.columns:
    df_filt["Bimestre"] = df_filt["Periodo"].apply(mapear_bimestre)

# Filtrar apenas registros com bimestre válido e nota válida
df_evolucao = df_filt[(df_filt["Bimestre"].notna()) & (df_filt["Nota"].notna())].copy()

if len(df_evolucao) > 0 and "Turma" in df_evolucao.columns:
    # Calcular média geral por Turma e Bimestre
    evolucao_turmas = df_evolucao.groupby(["Turma", "Bimestre"])["Nota"].mean().reset_index()
    evolucao_turmas = evolucao_turmas.rename(columns={"Nota": "Média Geral"})
    
    # Ordenar por Bimestre para garantir ordem correta no gráfico
    evolucao_turmas = evolucao_turmas.sort_values(["Turma", "Bimestre"])
    
    # Verificar se há dados suficientes
    if len(evolucao_turmas) > 0:
        # Criar gráfico de linha
        fig_evolucao = px.line(
            evolucao_turmas, 
            x="Bimestre", 
            y="Média Geral", 
            color="Turma",
            markers=True,
            title="Evolução da Média Geral das Notas por Turma ao Longo dos 4 Bimestres",
            labels={
                "Bimestre": "Bimestre",
                "Média Geral": "Média Geral das Notas",
                "Turma": "Turma"
            }
        )
        
        # Personalizar layout
        fig_evolucao.update_layout(
            xaxis_title="Bimestre",
            yaxis_title="Média Geral das Notas",
            hovermode='x unified',
            legend=dict(
                title="Turma (passe o mouse para ver detalhes)",
                orientation="v",
                yanchor="top",
                y=1,
                xanchor="left",
                x=1.02,
                font=dict(size=9)
            ),
            height=500,
            xaxis=dict(
                tickmode='linear',
                tick0=1,
                dtick=1,
                range=[0.5, 4.5]
            )
        )
        
        # Melhorar tooltip para mostrar mais informações
        fig_evolucao.update_traces(
            hovertemplate='<b>%{fullData.name}</b><br>' +
                         'Bimestre: %{x}<br>' +
                         'Média: %{y:.2f}<extra></extra>'
        )
        
        # Adicionar linha de referência na média 6.0
        fig_evolucao.add_hline(
            y=6.0, 
            line_dash="dash", 
            line_color="red", 
            annotation_text="Média Mínima (6.0)",
            annotation_position="left"
        )
        
        st.plotly_chart(fig_evolucao, use_container_width=True)
        
        # Botão de exportação
        col_export_evol1, col_export_evol2 = st.columns([1, 4])
        with col_export_evol1:
            if st.button("📊 Exportar Dados do Gráfico", key="export_grafico_evolucao", help="Baixar planilha com dados da evolução"):
                # Preparar dados para exportação
                dados_export_evol = evolucao_turmas.copy()
                dados_export_evol = dados_export_evol.rename(columns={
                    'Turma': 'Turma',
                    'Bimestre': 'Bimestre',
                    'Média Geral': 'Media_Geral'
                })
                dados_export_evol = dados_export_evol.sort_values(["Turma", "Bimestre"])
                
                excel_data = criar_excel_formatado(dados_export_evol, "Evolucao_Notas_Turmas")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="evolucao_notas_turmas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # Estatísticas resumidas
        st.markdown("**Resumo da Evolução:**")
        col_res1, col_res2, col_res3 = st.columns(3)
        
        # Calcular média total de cada turma (média dos 4 bimestres)
        medias_turmas = evolucao_turmas.groupby("Turma")["Média Geral"].mean().reset_index()
        medias_turmas = medias_turmas.rename(columns={"Média Geral": "Média Total"})
        medias_turmas = medias_turmas.sort_values("Média Total", ascending=False)
        
        with col_res1:
            # Turma com melhor média total dos 4 bimestres
            if len(medias_turmas) > 0:
                melhor_turma = medias_turmas.iloc[0]
                st.metric(
                    "Melhor Média Total (4 Bimestres)", 
                    f"{melhor_turma['Turma']}: {melhor_turma['Média Total']:.2f}"
                )
        
        with col_res2:
            # Turma com pior média total dos 4 bimestres
            if len(medias_turmas) > 0:
                pior_turma = medias_turmas.iloc[-1]
                st.metric(
                    "Pior Média Total (4 Bimestres)", 
                    f"{pior_turma['Turma']}: {pior_turma['Média Total']:.2f}"
                )
        
        with col_res3:
            # Média geral de todas as turmas (média da escola)
            if len(medias_turmas) > 0:
                media_escola = medias_turmas["Média Total"].mean()
                st.metric(
                    "Média Geral da Escola", 
                    f"{media_escola:.2f}"
                )
        
        # Ranking Top 10 Melhores Alunos
        st.markdown("---")
        st.markdown("### 🏆 Top 10 Melhores Alunos da Escola (Média dos 4 Bimestres)")
        
        # Calcular média geral por aluno (média de todas as disciplinas)
        try:
            if "MediaFinal" in indic.columns and coluna_aluno in indic.columns:
                # Agrupar por aluno e calcular média geral
                ranking_alunos = indic.groupby([coluna_aluno, "Turma"])["MediaFinal"].mean().reset_index()
                ranking_alunos = ranking_alunos.rename(columns={"MediaFinal": "Média Geral"})
                
                # Se um aluno estiver em múltiplas turmas, pegar a primeira turma
                ranking_alunos = ranking_alunos.groupby(coluna_aluno).agg({
                    "Média Geral": "mean",
                    "Turma": "first"
                }).reset_index()
                
                # Ordenar por média (maior para menor) e pegar top 10
                ranking_alunos = ranking_alunos.sort_values("Média Geral", ascending=False).head(10).reset_index(drop=True)
                
                # Adicionar coluna de posição
                ranking_alunos.insert(0, "Posição", range(1, len(ranking_alunos) + 1))
                
                # Formatar média para 2 casas decimais
                ranking_alunos["Média Geral"] = ranking_alunos["Média Geral"].round(2)
                
                # Renomear colunas para exibição
                ranking_alunos_display = ranking_alunos.copy()
                ranking_alunos_display = ranking_alunos_display.rename(columns={
                    coluna_aluno: "Aluno",
                    "Turma": "Turma",
                    "Média Geral": "Média Geral"
                })
                
                # Exibir tabela estilizada
                st.dataframe(
                    ranking_alunos_display[["Posição", "Aluno", "Turma", "Média Geral"]],
                    use_container_width=True,
                    hide_index=True
                )
                
                # Botão de exportação
                col_exp_rank1, col_exp_rank2 = st.columns([1, 4])
                with col_exp_rank1:
                    if st.button("📊 Exportar Ranking", key="export_ranking_top10", help="Baixar planilha com ranking dos top 10 alunos"):
                        dados_export_rank = ranking_alunos.copy()
                        dados_export_rank = dados_export_rank.rename(columns={
                            coluna_aluno: "Aluno",
                            "Turma": "Turma",
                            "Média Geral": "Media_Geral"
                        })
                        dados_export_rank = dados_export_rank[["Posição", "Aluno", "Turma", "Media_Geral"]]
                        
                        excel_data = criar_excel_formatado(dados_export_rank, "Ranking_Top10_Alunos")
                        st.download_button(
                            label="Baixar Excel",
                            data=excel_data,
                            file_name="ranking_top10_alunos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.info("Dados de média final não disponíveis para gerar o ranking.")
        except NameError:
            st.info("Dados de indicadores não disponíveis para gerar o ranking.")
    else:
        st.info("Não há dados suficientes para gerar o gráfico de evolução.")
else:
    st.info("Não há dados de turmas ou bimestres disponíveis para gerar o gráfico de evolução.")

st.markdown("---")

# Seção de Gráficos de Notas por Disciplina
st.markdown("### 📊 Gráficos de Notas Abaixo da Média por Disciplina")

# Gráfico Geral (1º + 2º + 3º + 4º Bimestre)
with st.expander("📈 Geral - Notas Abaixo da Média por Disciplina (1º + 2º + 3º + 4º Bimestre)"):
    base_baixas = pd.concat([notas_baixas_b1, notas_baixas_b2, notas_baixas_b3, notas_baixas_b4], ignore_index=True)
    if len(base_baixas) > 0:
        # Contar notas por disciplina
        contagem = base_baixas.groupby("Disciplina")["Nota"].count().reset_index()
        contagem = contagem.rename(columns={"Nota": "Qtd Notas < 6"})
        
        # Ordenar em ordem decrescente (maior para menor)
        contagem = contagem.sort_values("Qtd Notas < 6", ascending=False).reset_index(drop=True)
        
        # Adicionar coluna de cores intercaladas baseada na posição após ordenação
        contagem['Cor'] = ['#1e40af' if i % 2 == 0 else '#059669' for i in range(len(contagem))]
        
        fig = px.bar(contagem, x="Disciplina", y="Qtd Notas < 6", 
                    title="Notas abaixo da média (1º + 2º + 3º Bimestre)",
                    color="Cor",
                    color_discrete_map={'#1e40af': '#1e40af', '#059669': '#059669'})
        
        # Forçar a ordem das disciplinas no eixo X
        fig.update_layout(
            xaxis_title=None, 
            yaxis_title="Quantidade", 
            bargap=0.25, 
            showlegend=False, 
            xaxis_tickangle=45,
            xaxis={'categoryorder': 'array', 'categoryarray': contagem['Disciplina'].tolist()}
        )
        st.plotly_chart(fig, use_container_width=True)
        
        # Botão de exportação para dados do gráfico
        col_export_graf1, col_export_graf2 = st.columns([1, 4])
        with col_export_graf1:
            if st.button("📊 Exportar Dados do Gráfico", key="export_grafico_notas_geral", help="Baixar planilha com dados do gráfico geral"):
                # Preparar dados para exportação (remover coluna de cor)
                dados_export = contagem[['Disciplina', 'Qtd Notas < 6']].copy()
                dados_export = dados_export.rename(columns={'Qtd Notas < 6': 'Quantidade_Notas_Abaixo_6'})
                
                excel_data = criar_excel_formatado(dados_export, "Notas_Por_Disciplina_Geral")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="notas_por_disciplina_geral.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Sem notas abaixo da média para os filtros atuais.")

# Gráficos separados por bimestre
col_graf1, col_graf2, col_graf3, col_graf4 = st.columns(4)

# Gráfico 1º Bimestre
with col_graf1:
    with st.expander("📊 1º Bimestre - Notas Abaixo da Média por Disciplina"):
        if len(notas_baixas_b1) > 0:
            # Contar notas por disciplina no 1º bimestre
            contagem_b1 = notas_baixas_b1.groupby("Disciplina")["Nota"].count().reset_index()
            contagem_b1 = contagem_b1.rename(columns={"Nota": "Qtd Notas < 6"})
            
            # Ordenar em ordem decrescente (maior para menor)
            contagem_b1 = contagem_b1.sort_values("Qtd Notas < 6", ascending=False).reset_index(drop=True)
            
            # Adicionar coluna de cores intercaladas baseada na posição após ordenação
            contagem_b1['Cor'] = ['#dc2626' if i % 2 == 0 else '#ea580c' for i in range(len(contagem_b1))]
            
            fig_b1 = px.bar(contagem_b1, x="Disciplina", y="Qtd Notas < 6", 
                           title="Notas abaixo da média - 1º Bimestre",
                           color="Cor",
                           color_discrete_map={'#dc2626': '#dc2626', '#ea580c': '#ea580c'})
            
            # Forçar a ordem das disciplinas no eixo X
            fig_b1.update_layout(
                xaxis_title=None, 
                yaxis_title="Quantidade", 
                bargap=0.25, 
                showlegend=False, 
                xaxis_tickangle=45,
                xaxis={'categoryorder': 'array', 'categoryarray': contagem_b1['Disciplina'].tolist()}
            )
            st.plotly_chart(fig_b1, use_container_width=True)
            
            # Botão de exportação para dados do gráfico 1º bimestre
            if st.button("📊 Exportar 1º Bimestre", key="export_grafico_notas_b1", help="Baixar planilha com dados do 1º bimestre"):
                # Preparar dados para exportação (remover coluna de cor)
                dados_export_b1 = contagem_b1[['Disciplina', 'Qtd Notas < 6']].copy()
                dados_export_b1 = dados_export_b1.rename(columns={'Qtd Notas < 6': 'Quantidade_Notas_Abaixo_6'})
                
                excel_data = criar_excel_formatado(dados_export_b1, "Notas_Por_Disciplina_B1")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="notas_por_disciplina_1bimestre.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Sem notas abaixo da média no 1º bimestre para os filtros atuais.")

# Gráfico 2º Bimestre
with col_graf2:
    with st.expander("📊 2º Bimestre - Notas Abaixo da Média por Disciplina"):
        if len(notas_baixas_b2) > 0:
            # Contar notas por disciplina no 2º bimestre
            contagem_b2 = notas_baixas_b2.groupby("Disciplina")["Nota"].count().reset_index()
            contagem_b2 = contagem_b2.rename(columns={"Nota": "Qtd Notas < 6"})
            
            # Ordenar em ordem decrescente (maior para menor)
            contagem_b2 = contagem_b2.sort_values("Qtd Notas < 6", ascending=False).reset_index(drop=True)
            
            # Adicionar coluna de cores intercaladas baseada na posição após ordenação
            contagem_b2['Cor'] = ['#7c3aed' if i % 2 == 0 else '#a855f7' for i in range(len(contagem_b2))]
            
            fig_b2 = px.bar(contagem_b2, x="Disciplina", y="Qtd Notas < 6", 
                           title="Notas abaixo da média - 2º Bimestre",
                           color="Cor",
                           color_discrete_map={'#7c3aed': '#7c3aed', '#a855f7': '#a855f7'})
            
            # Forçar a ordem das disciplinas no eixo X
            fig_b2.update_layout(
                xaxis_title=None, 
                yaxis_title="Quantidade", 
                bargap=0.25, 
                showlegend=False, 
                xaxis_tickangle=45,
                xaxis={'categoryorder': 'array', 'categoryarray': contagem_b2['Disciplina'].tolist()}
            )
            st.plotly_chart(fig_b2, use_container_width=True)
            
            # Botão de exportação para dados do gráfico 2º bimestre
            if st.button("📊 Exportar 2º Bimestre", key="export_grafico_notas_b2", help="Baixar planilha com dados do 2º bimestre"):
                # Preparar dados para exportação (remover coluna de cor)
                dados_export_b2 = contagem_b2[['Disciplina', 'Qtd Notas < 6']].copy()
                dados_export_b2 = dados_export_b2.rename(columns={'Qtd Notas < 6': 'Quantidade_Notas_Abaixo_6'})
                
                excel_data = criar_excel_formatado(dados_export_b2, "Notas_Por_Disciplina_B2")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="notas_por_disciplina_2bimestre.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Sem notas abaixo da média no 2º bimestre para os filtros atuais.")

# Gráfico 3º Bimestre
with col_graf3:
    with st.expander("📊 3º Bimestre - Notas Abaixo da Média por Disciplina"):
        if len(notas_baixas_b3) > 0:
            # Contar notas por disciplina no 3º bimestre
            contagem_b3 = notas_baixas_b3.groupby("Disciplina")["Nota"].count().reset_index()
            contagem_b3 = contagem_b3.rename(columns={"Nota": "Qtd Notas < 6"})
            
            # Ordenar em ordem decrescente (maior para menor)
            contagem_b3 = contagem_b3.sort_values("Qtd Notas < 6", ascending=False).reset_index(drop=True)
            
            # Adicionar coluna de cores intercaladas baseada na posição após ordenação
            contagem_b3['Cor'] = ['#3b82f6' if i % 2 == 0 else '#60a5fa' for i in range(len(contagem_b3))]
            
            fig_b3 = px.bar(contagem_b3, x="Disciplina", y="Qtd Notas < 6", 
                           title="Notas abaixo da média - 3º Bimestre",
                           color="Cor",
                           color_discrete_map={'#3b82f6': '#3b82f6', '#60a5fa': '#60a5fa'})
            
            # Forçar a ordem das disciplinas no eixo X
            fig_b3.update_layout(
                xaxis_title=None, 
                yaxis_title="Quantidade", 
                bargap=0.25, 
                showlegend=False, 
                xaxis_tickangle=45,
                xaxis={'categoryorder': 'array', 'categoryarray': contagem_b3['Disciplina'].tolist()}
            )
            st.plotly_chart(fig_b3, use_container_width=True)
            
            # Botão de exportação para dados do gráfico 3º bimestre
            if st.button("📊 Exportar 3º Bimestre", key="export_grafico_notas_b3", help="Baixar planilha com dados do 3º bimestre"):
                # Preparar dados para exportação (remover coluna de cor)
                dados_export_b3 = contagem_b3[['Disciplina', 'Qtd Notas < 6']].copy()
                dados_export_b3 = dados_export_b3.rename(columns={'Qtd Notas < 6': 'Quantidade_Notas_Abaixo_6'})
                
                excel_data = criar_excel_formatado(dados_export_b3, "Notas_Por_Disciplina_B3")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="notas_por_disciplina_3bimestre.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Sem notas abaixo da média no 3º bimestre para os filtros atuais.")

# Gráfico 4º Bimestre
with col_graf4:
    with st.expander("📊 4º Bimestre - Notas Abaixo da Média por Disciplina"):
        # Diagnóstico: Verificar dados do 4º bimestre
        dados_b4_total = df_filt[df_filt["Periodo"].str.contains("Quarto", case=False, na=False)]
        if len(dados_b4_total) > 0:
            st.info(f"ℹ️ **Diagnóstico:** Encontrados {len(dados_b4_total)} registros com 'Quarto' no período. Desses, {len(notas_baixas_b4)} têm nota < 6.0")
            # Mostrar exemplos de períodos encontrados
            periodos_unicos = dados_b4_total["Periodo"].unique()[:5]  # Primeiros 5 únicos
            if len(periodos_unicos) > 0:
                st.caption(f"📝 Exemplos de períodos encontrados: {', '.join(periodos_unicos[:3])}")
        
        if len(notas_baixas_b4) > 0:
            # Contar notas por disciplina no 4º bimestre
            contagem_b4 = notas_baixas_b4.groupby("Disciplina")["Nota"].count().reset_index()
            contagem_b4 = contagem_b4.rename(columns={"Nota": "Qtd Notas < 6"})
            
            # Ordenar em ordem decrescente (maior para menor)
            contagem_b4 = contagem_b4.sort_values("Qtd Notas < 6", ascending=False).reset_index(drop=True)
            
            # Adicionar coluna de cores intercaladas baseada na posição após ordenação
            contagem_b4['Cor'] = ['#ef4444' if i % 2 == 0 else '#f87171' for i in range(len(contagem_b4))]
            
            fig_b4 = px.bar(contagem_b4, x="Disciplina", y="Qtd Notas < 6", 
                           title="Notas abaixo da média - 4º Bimestre",
                           color="Cor",
                           color_discrete_map={'#ef4444': '#ef4444', '#f87171': '#f87171'})
            
            # Forçar a ordem das disciplinas no eixo X
            fig_b4.update_layout(
                xaxis_title=None, 
                yaxis_title="Quantidade", 
                bargap=0.25, 
                showlegend=False, 
                xaxis_tickangle=45,
                xaxis={'categoryorder': 'array', 'categoryarray': contagem_b4['Disciplina'].tolist()}
            )
            st.plotly_chart(fig_b4, use_container_width=True)
            
            # Botão de exportação para dados do gráfico 4º bimestre
            if st.button("📊 Exportar 4º Bimestre", key="export_grafico_notas_b4", help="Baixar planilha com dados do 4º bimestre"):
                # Preparar dados para exportação (remover coluna de cor)
                dados_export_b4 = contagem_b4[['Disciplina', 'Qtd Notas < 6']].copy()
                dados_export_b4 = dados_export_b4.rename(columns={'Qtd Notas < 6': 'Quantidade_Notas_Abaixo_6'})
                
                excel_data = criar_excel_formatado(dados_export_b4, "Notas_Por_Disciplina_B4")
                st.download_button(
                    label="Baixar Excel",
                    data=excel_data,
                    file_name="notas_por_disciplina_4bimestre.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Sem notas abaixo da média no 4º bimestre para os filtros atuais.")

# Nova seção: Gráficos de Barras - Aprovados x Reprovados
st.markdown("---")
st.markdown("### 📊 Distribuição de Alunos: Aprovados vs Reprovados por Bimestre")

# Calcular aprovados e reprovados por bimestre (alunos únicos)
col_pizza1, col_pizza2, col_pizza3, col_pizza4 = st.columns(4)

# 1º Bimestre
with col_pizza1:
    st.markdown("#### 1º Bimestre")
    total_alunos_b1 = df_filt[df_filt["Periodo"].str.contains("Primeiro", case=False, na=False)][coluna_aluno].nunique()
    aprovados_b1 = total_alunos_b1 - alunos_notas_baixas_b1
    
    if total_alunos_b1 > 0:
        dados_barra_b1 = pd.DataFrame({
            'Status': ['Aprovados (≥6)', 'Reprovados (<6)'],
            'Quantidade': [aprovados_b1, alunos_notas_baixas_b1]
        })
        
        fig_barra_b1 = px.bar(dados_barra_b1, x='Status', y='Quantidade',
                              title=f"Total: {total_alunos_b1} alunos",
                              color='Status',
                              color_discrete_map={'Aprovados (≥6)': '#10b981', 'Reprovados (<6)': '#dc2626'},
                              text='Quantidade')
        fig_barra_b1.update_traces(texttemplate='%{text} (%{y:.0%})', textposition='outside')
        fig_barra_b1.update_layout(showlegend=False, yaxis_title="Quantidade de Alunos", xaxis_title=None)
        st.plotly_chart(fig_barra_b1, use_container_width=True)
        
        # Métricas
        st.metric("Aprovados", f"{aprovados_b1} ({aprovados_b1/total_alunos_b1*100:.1f}%)")
        st.metric("Reprovados", f"{alunos_notas_baixas_b1} ({alunos_notas_baixas_b1/total_alunos_b1*100:.1f}%)")
    else:
        st.info("Sem dados do 1º bimestre")

# 2º Bimestre
with col_pizza2:
    st.markdown("#### 2º Bimestre")
    total_alunos_b2 = df_filt[df_filt["Periodo"].str.contains("Segundo", case=False, na=False)][coluna_aluno].nunique()
    aprovados_b2 = total_alunos_b2 - alunos_notas_baixas_b2
    
    if total_alunos_b2 > 0:
        dados_barra_b2 = pd.DataFrame({
            'Status': ['Aprovados (≥6)', 'Reprovados (<6)'],
            'Quantidade': [aprovados_b2, alunos_notas_baixas_b2]
        })
        
        fig_barra_b2 = px.bar(dados_barra_b2, x='Status', y='Quantidade',
                              title=f"Total: {total_alunos_b2} alunos",
                              color='Status',
                              color_discrete_map={'Aprovados (≥6)': '#10b981', 'Reprovados (<6)': '#7c3aed'},
                              text='Quantidade')
        fig_barra_b2.update_traces(texttemplate='%{text} (%{y:.0%})', textposition='outside')
        fig_barra_b2.update_layout(showlegend=False, yaxis_title="Quantidade de Alunos", xaxis_title=None)
        st.plotly_chart(fig_barra_b2, use_container_width=True)
        
        # Métricas
        st.metric("Aprovados", f"{aprovados_b2} ({aprovados_b2/total_alunos_b2*100:.1f}%)")
        st.metric("Reprovados", f"{alunos_notas_baixas_b2} ({alunos_notas_baixas_b2/total_alunos_b2*100:.1f}%)")
    else:
        st.info("Sem dados do 2º bimestre")

# 3º Bimestre
with col_pizza3:
    st.markdown("#### 3º Bimestre")
    total_alunos_b3 = df_filt[df_filt["Periodo"].str.contains("Terceiro", case=False, na=False)][coluna_aluno].nunique()
    aprovados_b3 = total_alunos_b3 - alunos_notas_baixas_b3
    
    if total_alunos_b3 > 0:
        dados_barra_b3 = pd.DataFrame({
            'Status': ['Aprovados (≥6)', 'Reprovados (<6)'],
            'Quantidade': [aprovados_b3, alunos_notas_baixas_b3]
        })
        
        fig_barra_b3 = px.bar(dados_barra_b3, x='Status', y='Quantidade',
                              title=f"Total: {total_alunos_b3} alunos",
                              color='Status',
                              color_discrete_map={'Aprovados (≥6)': '#10b981', 'Reprovados (<6)': '#3b82f6'},
                              text='Quantidade')
        fig_barra_b3.update_traces(texttemplate='%{text} (%{y:.0%})', textposition='outside')
        fig_barra_b3.update_layout(showlegend=False, yaxis_title="Quantidade de Alunos", xaxis_title=None)
        st.plotly_chart(fig_barra_b3, use_container_width=True)
        
        # Métricas
        st.metric("Aprovados", f"{aprovados_b3} ({aprovados_b3/total_alunos_b3*100:.1f}%)")
        st.metric("Reprovados", f"{alunos_notas_baixas_b3} ({alunos_notas_baixas_b3/total_alunos_b3*100:.1f}%)")
    else:
        st.info("Sem dados do 3º bimestre")

# 4º Bimestre
with col_pizza4:
    st.markdown("#### 4º Bimestre")
    total_alunos_b4 = df_filt[df_filt["Periodo"].str.contains("Quarto", case=False, na=False)][coluna_aluno].nunique()
    aprovados_b4 = total_alunos_b4 - alunos_notas_baixas_b4
    
    if total_alunos_b4 > 0:
        dados_barra_b4 = pd.DataFrame({
            'Status': ['Aprovados (≥6)', 'Reprovados (<6)'],
            'Quantidade': [aprovados_b4, alunos_notas_baixas_b4]
        })
        
        fig_barra_b4 = px.bar(dados_barra_b4, x='Status', y='Quantidade',
                              title=f"Total: {total_alunos_b4} alunos",
                              color='Status',
                              color_discrete_map={'Aprovados (≥6)': '#10b981', 'Reprovados (<6)': '#ef4444'},
                              text='Quantidade')
        fig_barra_b4.update_traces(texttemplate='%{text} (%{y:.0%})', textposition='outside')
        fig_barra_b4.update_layout(showlegend=False, yaxis_title="Quantidade de Alunos", xaxis_title=None)
        st.plotly_chart(fig_barra_b4, use_container_width=True)
        
        # Métricas
        st.metric("Aprovados", f"{aprovados_b4} ({aprovados_b4/total_alunos_b4*100:.1f}%)")
        st.metric("Reprovados", f"{alunos_notas_baixas_b4} ({alunos_notas_baixas_b4/total_alunos_b4*100:.1f}%)")
    else:
        st.info("Sem dados do 4º bimestre")

# Seção de Análise Final de Aprovação/Reprovação
st.markdown("---")
st.markdown("""
<div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análise Final de Aprovação/Reprovação</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Conceito Final baseado na média dos 4 bimestres</p>
</div>
""", unsafe_allow_html=True)

# Calcular estatísticas finais de aprovação/reprovação
if "StatusFinal" in indic.columns and "MediaFinal" in indic.columns:
    # Filtrar apenas registros com média final calculada (não incompletos)
    indic_final = indic[indic["StatusFinal"] != "Incompleto"].copy()
    
    if len(indic_final) > 0:
        # Contagem de aprovados e reprovados
        contagem_status = indic_final["StatusFinal"].value_counts()
        total_final = len(indic_final)
        aprovados_final = contagem_status.get("Aprovado", 0)
        reprovados_final = contagem_status.get("Reprovado", 0)
        
        # Calcular por aluno único (não por aluno-disciplina)
        # IMPORTANTE: Um aluno pode ter disciplinas aprovadas E reprovadas
        alunos_com_aprovados = indic_final[indic_final["StatusFinal"] == "Aprovado"][coluna_aluno].unique()
        alunos_com_reprovados = indic_final[indic_final["StatusFinal"] == "Reprovado"][coluna_aluno].unique()
        
        alunos_aprovados = len(set(alunos_com_aprovados))  # Alunos com pelo menos 1 disciplina aprovada
        alunos_reprovados = len(set(alunos_com_reprovados))  # Alunos com pelo menos 1 disciplina reprovada
        alunos_aprovados_e_reprovados = len(set(alunos_com_aprovados) & set(alunos_com_reprovados))  # Alunos em ambos grupos
        total_alunos_final = indic_final[coluna_aluno].nunique()
        
        # Métricas principais
        col_final1, col_final2, col_final3 = st.columns(3)
        
        with col_final1:
            percent_aprovados = (aprovados_final / total_final * 100) if total_final > 0 else 0
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #d1fae5, #a7f3d0); border-radius: 10px; padding: 20px; margin: 5px 0; box-shadow: 0 2px 8px rgba(16, 185, 129, 0.15); border-left: 4px solid #10b981;">
                <div style="font-size: 0.95em; font-weight: 600; color: #065f46; margin-bottom: 8px;">✅ Aprovados (Média ≥ 6.0)</div>
                <div style="font-size: 2.5em; font-weight: 700; color: #065f46; margin: 8px 0;">{aprovados_final}</div>
                <div style="font-size: 1.1em; color: #64748b; font-weight: 500; margin-bottom: 12px;">({percent_aprovados:.1f}%) registros aluno-disciplina</div>
                <div style="border-top: 2px solid #10b981; padding-top: 12px; margin-top: 12px;">
                    <div style="font-size: 1.5em; font-weight: 700; color: #065f46; margin: 4px 0;">{alunos_aprovados}</div>
                    <div style="font-size: 1.1em; color: #059669; font-weight: 600; margin-top: 4px;">👥 alunos únicos</div>
                    <div style="font-size: 0.85em; color: #6b7280; margin-top: 6px; font-style: italic;">(com pelo menos 1 disciplina aprovada)</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_final2:
            percent_reprovados = (reprovados_final / total_final * 100) if total_final > 0 else 0
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #fee2e2, #fecaca); border-radius: 10px; padding: 20px; margin: 5px 0; box-shadow: 0 2px 8px rgba(239, 68, 68, 0.15); border-left: 4px solid #ef4444;">
                <div style="font-size: 0.95em; font-weight: 600; color: #991b1b; margin-bottom: 8px;">❌ Reprovados (Média < 6.0)</div>
                <div style="font-size: 2.5em; font-weight: 700; color: #991b1b; margin: 8px 0;">{reprovados_final}</div>
                <div style="font-size: 1.1em; color: #64748b; font-weight: 500; margin-bottom: 12px;">({percent_reprovados:.1f}%) registros aluno-disciplina</div>
                <div style="border-top: 2px solid #ef4444; padding-top: 12px; margin-top: 12px;">
                    <div style="font-size: 1.5em; font-weight: 700; color: #991b1b; margin: 4px 0;">{alunos_reprovados}</div>
                    <div style="font-size: 1.1em; color: #dc2626; font-weight: 600; margin-top: 4px;">👥 alunos únicos</div>
                    <div style="font-size: 0.85em; color: #6b7280; margin-top: 6px; font-style: italic;">(com pelo menos 1 disciplina reprovada)</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_final3:
            media_geral = indic_final["MediaFinal"].mean()
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #dbeafe, #bfdbfe); border-radius: 10px; padding: 20px; margin: 5px 0; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.15); border-left: 4px solid #3b82f6;">
                <div style="font-size: 0.95em; font-weight: 600; color: #1e40af; margin-bottom: 8px;">📊 Média Geral Final</div>
                <div style="font-size: 2.5em; font-weight: 700; color: #1e40af; margin: 8px 0;">{media_geral:.2f}</div>
                <div style="font-size: 1.1em; color: #64748b; font-weight: 500; margin-bottom: 12px;">de {total_final} registros aluno-disciplina</div>
                <div style="border-top: 2px solid #3b82f6; padding-top: 12px; margin-top: 12px;">
                    <div style="font-size: 1.5em; font-weight: 700; color: #1e40af; margin: 4px 0;">{total_alunos_final}</div>
                    <div style="font-size: 1.1em; color: #3b82f6; font-weight: 600; margin-top: 4px;">👥 alunos únicos</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Explicação sobre sobreposição
        if alunos_aprovados_e_reprovados > 0:
            st.info(f"ℹ️ **Importante:** {alunos_aprovados_e_reprovados} alunos têm disciplinas aprovadas E reprovadas simultaneamente. Por isso, a soma de alunos únicos ({alunos_aprovados} + {alunos_reprovados} = {alunos_aprovados + alunos_reprovados}) é maior que o total de {total_alunos_final} alunos únicos.")
        
        # Gráfico de distribuição final
        st.markdown("#### 📊 Distribuição Final: Aprovados vs Reprovados")
        dados_final = pd.DataFrame({
            'Status': ['Aprovados (≥6)', 'Reprovados (<6)'],
            'Quantidade': [aprovados_final, reprovados_final]
        })
        
        fig_final = px.bar(dados_final, x='Status', y='Quantidade',
                          title=f"Total: {total_final} registros (Aluno-Disciplina)",
                          color='Status',
                          color_discrete_map={'Aprovados (≥6)': '#10b981', 'Reprovados (<6)': '#ef4444'},
                          text='Quantidade')
        fig_final.update_traces(texttemplate='%{text} (%{y:.0%})', textposition='outside')
        fig_final.update_layout(showlegend=False, yaxis_title="Quantidade", xaxis_title=None)
        st.plotly_chart(fig_final, use_container_width=True)
        
        # Tabela detalhada com conceito final
        st.markdown("#### 📋 Detalhamento por Aluno-Disciplina")
        indic_detalhado = indic_final[[coluna_aluno, "Disciplina", "N1", "N2", "N3", "N4", "MediaFinal", "StatusFinal", "ClassificacaoFinal"]].copy()
        indic_detalhado = indic_detalhado.sort_values(["StatusFinal", "MediaFinal"], ascending=[False, False])
        
        # Renomear colunas para exibição
        indic_detalhado_display = indic_detalhado.copy()
        indic_detalhado_display = indic_detalhado_display.rename(columns={
            coluna_aluno: "Aluno",
            "MediaFinal": "Média Final",
            "StatusFinal": "Status Final",
            "ClassificacaoFinal": "Conceito Final"
        })
        
        # Formatar valores NaN para exibição
        indic_detalhado_display = indic_detalhado_display.fillna("-")
        
        # Adicionar emoji ao status para melhor visualização
        indic_detalhado_display["Status Final"] = indic_detalhado_display["Status Final"].apply(
            lambda x: f"✅ {x}" if x == "Aprovado" else (f"❌ {x}" if x == "Reprovado" else x)
        )
        
        st.dataframe(indic_detalhado_display, use_container_width=True, hide_index=True)
    else:
        st.info("Ainda não há dados completos dos 4 bimestres para análise final.")
else:
    st.info("Aguardando dados do 4º bimestre para calcular a análise final.")

# Gráfico: Distribuição de Frequência por Faixas
col_graf1, col_graf2 = st.columns(2)

# Gráfico: Distribuição de Frequência por Faixas
with col_graf2:
    with st.expander("Distribuição de Frequência por Faixas"):
        if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
            # Usar os mesmos dados do Resumo de Frequência
            if "Frequencia Anual" in df_filt.columns:
                freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
                freq_geral = freq_geral.rename(columns={"Frequencia Anual": "Frequencia"})
            else:
                freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
            
            freq_geral["Classificacao_Freq"] = freq_geral["Frequencia"].apply(classificar_frequencia_geral)
            contagem_freq_geral = freq_geral["Classificacao_Freq"].value_counts()
            
            # Preparar dados para o gráfico
            dados_grafico = []
            cores = {
                "Reprovado": "#dc2626",
                "Alto Risco": "#ea580c", 
                "Risco Moderado": "#d97706",
                "Ponto de Atenção": "#f59e0b",
                "Meta Favorável": "#16a34a"
            }
            
            for categoria, quantidade in contagem_freq_geral.items():
                if categoria != "Sem dados":  # Excluir "Sem dados" do gráfico
                    dados_grafico.append({
                        "Categoria": categoria,
                        "Quantidade": quantidade,
                        "Cor": cores.get(categoria, "#6b7280")
                    })
            
            if dados_grafico:
                df_grafico = pd.DataFrame(dados_grafico)
                
                # Criar gráfico de barras
                fig_freq = px.bar(df_grafico, x="Categoria", y="Quantidade", 
                                 title="Distribuição de Alunos por Faixa de Frequência",
                                 color="Categoria", 
                                 color_discrete_map=cores)
                fig_freq.update_layout(xaxis_title=None, yaxis_title="Número de Alunos", 
                                     bargap=0.25, showlegend=False, xaxis_tickangle=45)
                st.plotly_chart(fig_freq, use_container_width=True)
                
                # Botão de exportação para dados do gráfico de frequência
                col_export_graf3, col_export_graf4 = st.columns([1, 4])
                with col_export_graf3:
                    if st.button("📊 Exportar Dados do Gráfico", key="export_grafico_freq", help="Baixar planilha com dados do gráfico de frequência"):
                        # Preparar dados para exportação
                        dados_export_freq = df_grafico[['Categoria', 'Quantidade']].copy()
                        dados_export_freq = dados_export_freq.rename(columns={'Quantidade': 'Numero_Alunos'})
                        
                        excel_data = criar_excel_formatado(dados_export_freq, "Frequencia_Por_Faixa")
                        st.download_button(
                            label="Baixar Excel",
                            data=excel_data,
                            file_name="frequencia_por_faixa.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # Estatísticas adicionais
                st.markdown("**Resumo das Faixas de Frequência:**")
                col_stat1, col_stat2, col_stat3 = st.columns(3)
                with col_stat1:
                    total_alunos = contagem_freq_geral.sum()
                    st.metric("Total de Alunos", total_alunos, help="Total de alunos considerados na análise de frequência")
                with col_stat2:
                    alunos_risco = contagem_freq_geral.get("Reprovado", 0) + contagem_freq_geral.get("Alto Risco", 0)
                    st.metric("Alunos em Risco", alunos_risco, help="Alunos reprovados ou em alto risco de reprovação por frequência")
                with col_stat3:
                    alunos_meta = contagem_freq_geral.get("Meta Favorável", 0)
                    percentual_meta = (alunos_meta / total_alunos * 100) if total_alunos > 0 else 0
                    st.metric("Meta Favorável", f"{percentual_meta:.1f}%", help="Percentual de alunos com frequência ≥ 95% (meta favorável)")
            else:
                st.info("Sem dados de frequência para exibir.")
        else:
            st.info("Dados de frequência não disponíveis na planilha.")

# Seção expandível: Análise Cruzada Nota x Frequência (movida para o final)
st.markdown("---")
st.markdown("""
<div style="background: linear-gradient(135deg, #1e40af, #3b82f6); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(30, 64, 175, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">Análise Cruzada</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Cruzamento entre Notas e Frequência</p>
</div>
""", unsafe_allow_html=True)

with st.expander("Análise Cruzada: Notas x Frequência"):
    if ("Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns) and len(indic) > 0:
        # Combinar dados de notas e frequência (priorizando Frequencia Anual)
        if "Frequencia Anual" in df_filt.columns:
            freq_alunos = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
            freq_alunos = freq_alunos.rename(columns={"Frequencia Anual": "Frequencia"})
        else:
            freq_alunos = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
        freq_alunos["Classificacao_Freq"] = freq_alunos["Frequencia"].apply(classificar_frequencia)
        
        # Merge com indicadores de notas
        cruzada = indic.merge(freq_alunos, on=[coluna_aluno, "Turma"], how="left")
        
        # Criar matriz de cruzamento
        matriz_cruzada = cruzada.groupby(["Classificacao", "Classificacao_Freq"]).size().unstack(fill_value=0)
        
        if not matriz_cruzada.empty:
            st.markdown("**Matriz de Cruzamento: Classificação de Notas x Frequência**")
            st.dataframe(matriz_cruzada, use_container_width=True)
            
            # Análise de alunos com frequência abaixo de 95%
            freq_baixa = cruzada[cruzada["Frequencia"] < 95]
            
            if len(freq_baixa) > 0:
                st.markdown("### Alunos com Frequência Abaixo de 95% (Cruzamento Notas x Frequência)")
                # Mostrar apenas colunas relevantes para frequência baixa
                freq_baixa_display = freq_baixa[[coluna_aluno, "Turma", "Disciplina", "Classificacao", "Classificacao_Freq", "Frequencia"]].copy()
                # Formatar frequência
                freq_baixa_display["Frequencia"] = freq_baixa_display["Frequencia"].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
                )
                
                # Função para colorir classificações de frequência
                def color_frequencia_classification(val):
                    if val == "Reprovado":
                        return "background-color: #dc2626; color: white; font-weight: bold;"  # Vermelho forte
                    elif val == "Alto Risco":
                        return "background-color: #ea580c; color: white; font-weight: bold;"  # Laranja escuro
                    elif val == "Risco Moderado":
                        return "background-color: #f59e0b; color: white; font-weight: bold;"  # Laranja forte
                    elif val == "Ponto de Atenção":
                        return "background-color: #eab308; color: white; font-weight: bold;"  # Amarelo forte
                    elif val == "Meta Favorável":
                        return "background-color: #10b981; color: white; font-weight: bold;"  # Verde forte
                    else:
                        return ""
                
                # Aplicar cores nas duas colunas de classificação
                styled_cruzada = freq_baixa_display.style.applymap(
                    color_classification, subset=["Classificacao"]
                ).applymap(
                    color_frequencia_classification, subset=["Classificacao_Freq"]
                )
                
                st.dataframe(styled_cruzada, use_container_width=True)
                
                # Legenda para classificações de frequência
                st.markdown("### 🎨 Legenda das Classificações")
                col_leg1, col_leg2 = st.columns(2)
                
                with col_leg1:
                    st.markdown("**Classificação de Notas:**")
                    st.markdown("""
                    <div style="background-color: #10b981; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟢 Verde: Aluno está bem (N1≥6 e N2≥6)
                    </div>
                    <div style="background-color: #dc2626; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🔴 Vermelho Duplo: Risco alto (N1<6 e N2<6)
                    </div>
                    <div style="background-color: #f59e0b; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟠 Queda p/ Vermelho: Piorou (N1≥6 e N2<6)
                    </div>
                    <div style="background-color: #3b82f6; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🔵 Recuperou: Melhorou (N1<6 e N2≥6)
                    </div>
                    <div style="background-color: #6b7280; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        ⚪ Incompleto: Falta nota
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_leg2:
                    st.markdown("**Classificação de Frequência:**")
                    st.markdown("""
                    <div style="background-color: #dc2626; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🔴 Reprovado: < 75%
                    </div>
                    <div style="background-color: #ea580c; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟠 Alto Risco: < 80%
                    </div>
                    <div style="background-color: #f59e0b; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟠 Risco Moderado: < 90%
                    </div>
                    <div style="background-color: #eab308; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟡 Ponto de Atenção: < 95%
                    </div>
                    <div style="background-color: #10b981; color: white; padding: 5px; border-radius: 3px; margin: 2px 0; font-weight: bold; text-align: center;">
                        🟢 Meta Favorável: ≥ 95%
                    </div>
                    """, unsafe_allow_html=True)
                
                # Botão de exportação para alunos com frequência baixa
                col_export_freq_baixa1, col_export_freq_baixa2 = st.columns([1, 4])
                with col_export_freq_baixa1:
                    if st.button("📊 Exportar Cruzamento", key="export_freq_baixa", help="Baixar planilha com cruzamento de notas e frequência (alunos com frequência < 95%)"):
                        excel_data = criar_excel_formatado(freq_baixa_display, "Cruzamento_Notas_Freq")
                        st.download_button(
                            label="Baixar Excel",
                            data=excel_data,
                            file_name="cruzamento_notas_frequencia.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.info("Todos os alunos têm frequência ≥ 95% (Meta Favorável).")
        else:
            st.info("Dados insuficientes para análise cruzada.")
    else:
        st.info("Dados de frequência ou notas não disponíveis para análise cruzada.")

# Botão para baixar todas as planilhas em uma única planilha Excel
st.markdown("---")
st.markdown("""
<div style="background: linear-gradient(135deg, #059669, #10b981); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(5, 150, 105, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">📊 Exportação Completa</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Baixar todas as análises em uma única planilha Excel</p>
</div>
""", unsafe_allow_html=True)

col_export_all1, col_export_all2 = st.columns([1, 4])
with col_export_all1:
    if st.button("📊 Baixar Tudo", key="export_tudo", help="Baixar todas as análises em uma única planilha Excel com múltiplas abas"):
        # Criar arquivo Excel com múltiplas abas
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Aba 1: Alunos em Alerta
            if len(tabela_alerta) > 0:
                tabela_alerta[cols_visiveis].to_excel(writer, sheet_name="Alunos_em_Alerta", index=False)
            
            # Aba 2: Panorama Geral de Notas
            tab_diag[[coluna_aluno, "Turma", "Disciplina", "N1", "N2", "Media12", "Classificacao", "ReqMediaProx2"]].to_excel(
                writer, sheet_name="Panorama_Geral_Notas", index=False)
            
            # Aba 3: Análise de Frequência (se disponível)
            if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
                if "Frequencia Anual" in df_filt.columns:
                    freq_detalhada = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
                    freq_detalhada = freq_detalhada.rename(columns={"Frequencia Anual": "Frequencia"})
                else:
                    freq_detalhada = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
                
                freq_detalhada["Classificacao_Freq"] = freq_detalhada["Frequencia"].apply(classificar_frequencia)
                freq_detalhada["Frequencia_Formatada"] = freq_detalhada["Frequencia"].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
                )
                freq_detalhada[[coluna_aluno, "Turma", "Frequencia_Formatada", "Classificacao_Freq"]].to_excel(
                    writer, sheet_name="Analise_Frequencia", index=False)
            
            # Aba 4: Notas por Disciplina (se houver dados)
            base_baixas = pd.concat([notas_baixas_b1, notas_baixas_b2], ignore_index=True)
            if len(base_baixas) > 0:
                contagem = base_baixas.groupby("Disciplina")["Nota"].count().reset_index()
                contagem = contagem.rename(columns={"Nota": "Quantidade_Notas_Abaixo_6"})
                contagem = contagem.sort_values("Quantidade_Notas_Abaixo_6", ascending=False).reset_index(drop=True)
                contagem.to_excel(writer, sheet_name="Notas_Por_Disciplina", index=False)
            
            # Aba 5: Frequência por Faixas (se disponível)
            if "Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns:
                if "Frequencia Anual" in df_filt.columns:
                    freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
                    freq_geral = freq_geral.rename(columns={"Frequencia Anual": "Frequencia"})
                else:
                    freq_geral = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
                
                freq_geral["Classificacao_Freq"] = freq_geral["Frequencia"].apply(classificar_frequencia_geral)
                contagem_freq_geral = freq_geral["Classificacao_Freq"].value_counts()
                
                dados_grafico = []
                for categoria, quantidade in contagem_freq_geral.items():
                    if categoria != "Sem dados":
                        dados_grafico.append({
                            "Categoria": categoria,
                            "Numero_Alunos": quantidade
                        })
                
                if dados_grafico:
                    df_grafico = pd.DataFrame(dados_grafico)
                    df_grafico.to_excel(writer, sheet_name="Frequencia_Por_Faixa", index=False)
            
            # Aba 6: Cruzamento Notas x Frequência (se disponível)
            if ("Frequencia Anual" in df_filt.columns or "Frequencia" in df_filt.columns) and len(indic) > 0:
                if "Frequencia Anual" in df_filt.columns:
                    freq_alunos = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia Anual"].last().reset_index()
                    freq_alunos = freq_alunos.rename(columns={"Frequencia Anual": "Frequencia"})
                else:
                    freq_alunos = df_filt.groupby([coluna_aluno, "Turma"])["Frequencia"].last().reset_index()
                
                freq_alunos["Classificacao_Freq"] = freq_alunos["Frequencia"].apply(classificar_frequencia)
                cruzada = indic.merge(freq_alunos, on=[coluna_aluno, "Turma"], how="left")
                freq_baixa = cruzada[cruzada["Frequencia"] < 95]
                
                if len(freq_baixa) > 0:
                    freq_baixa_display = freq_baixa[[coluna_aluno, "Turma", "Disciplina", "Classificacao", "Classificacao_Freq", "Frequencia"]].copy()
                    freq_baixa_display["Frequencia"] = freq_baixa_display["Frequencia"].apply(
                        lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A"
                    )
                    freq_baixa_display.to_excel(writer, sheet_name="Cruzamento_Notas_Freq", index=False)
            
            # Aba 7: Alunos Duplicados (se houver)
            alunos_turmas = df_filt.groupby(coluna_aluno)["Turma"].nunique().reset_index()
            alunos_turmas = alunos_turmas.rename(columns={"Turma": "Qtd_Turmas"})
            alunos_duplicados = alunos_turmas[alunos_turmas["Qtd_Turmas"] > 1].copy()
            
            if len(alunos_duplicados) > 0:
                # Criar formato com colunas separadas para cada turma
                export_data = []
                for _, row in alunos_duplicados.iterrows():
                    aluno = row[coluna_aluno]
                    qtd_turmas = row["Qtd_Turmas"]
                    turmas_aluno = df_filt[df_filt[coluna_aluno] == aluno]["Turma"].unique().tolist()
                    turmas_aluno = sorted(turmas_aluno)
                    
                    # Criar linha com colunas separadas
                    linha = {
                        coluna_aluno: aluno,
                        "Qtd_Turmas": qtd_turmas
                    }
                    
                    # Adicionar cada turma em uma coluna separada
                    for i, turma in enumerate(turmas_aluno, 1):
                        linha[f"Turma_{i}"] = turma
                    
                    # Preencher colunas vazias com None para alunos com menos turmas
                    max_turmas = alunos_duplicados["Qtd_Turmas"].max()
                    for i in range(len(turmas_aluno) + 1, max_turmas + 1):
                        linha[f"Turma_{i}"] = None
                    
                    export_data.append(linha)
                
                df_export = pd.DataFrame(export_data)
                df_export = df_export.sort_values(["Qtd_Turmas", coluna_aluno], ascending=[False, True])
                df_export.to_excel(writer, sheet_name="Alunos_Duplicados", index=False)
        
        output.seek(0)
        st.download_button(
            label="📥 Baixar Planilha Completa",
            data=output.getvalue(),
            file_name="painel_sge_completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Seção: Identificação de Alunos em Múltiplas Turmas
st.markdown("---")
st.markdown("""
<div style="background: linear-gradient(135deg, #dc2626, #ef4444); border-radius: 12px; padding: 25px; margin: 20px 0; box-shadow: 0 4px 15px rgba(220, 38, 38, 0.2);">
    <h2 style="color: white; text-align: center; margin: 0; font-size: 1.7em; font-weight: 700; text-shadow: 0 1px 3px rgba(0,0,0,0.3);">🔍 Identificação de Alunos Duplicados</h2>
    <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 8px 0 0 0; font-size: 1.1em; font-weight: 500;">Detecção de alunos que aparecem em múltiplas turmas</p>
</div>
""", unsafe_allow_html=True)

# Identificar alunos em múltiplas turmas
alunos_turmas = df_filt.groupby(coluna_aluno)["Turma"].nunique().reset_index()
alunos_turmas = alunos_turmas.rename(columns={"Turma": "Qtd_Turmas"})

# Filtrar apenas alunos com mais de uma turma
alunos_duplicados = alunos_turmas[alunos_turmas["Qtd_Turmas"] > 1].copy()

if len(alunos_duplicados) > 0:
    # Criar dataframe detalhado com todas as turmas de cada aluno duplicado
    alunos_detalhados = []
    
    for _, row in alunos_duplicados.iterrows():
        aluno = row[coluna_aluno]
        qtd_turmas = row["Qtd_Turmas"]
        
        # Obter todas as turmas deste aluno
        turmas_aluno = df_filt[df_filt[coluna_aluno] == aluno]["Turma"].unique().tolist()
        turmas_str = ", ".join(sorted(turmas_aluno))
        
        alunos_detalhados.append({
            coluna_aluno: aluno,
            "Qtd_Turmas": qtd_turmas,
            "Turmas": turmas_str
        })
    
    df_alunos_duplicados = pd.DataFrame(alunos_detalhados)
    df_alunos_duplicados = df_alunos_duplicados.sort_values(["Qtd_Turmas", coluna_aluno], ascending=[False, True])
    
    # Função para colorir quantidade de turmas
    def color_qtd_turmas(val):
        if val == 2:
            return "background-color: #fef3c7; color: #92400e"  # Amarelo para duplicidade
        elif val == 3:
            return "background-color: #fed7aa; color: #9a3412"  # Laranja para triplicidade
        elif val >= 4:
            return "background-color: #fecaca; color: #991b1b"  # Vermelho para 4+ turmas
        else:
            return ""
    
    # Aplicar cores
    styled_duplicados = df_alunos_duplicados.style.applymap(color_qtd_turmas, subset=["Qtd_Turmas"])
    
    st.dataframe(styled_duplicados, use_container_width=True)
    
    # Métricas resumidas
    col_dup1, col_dup2, col_dup3 = st.columns(3)
    
    with col_dup1:
        total_duplicados = len(df_alunos_duplicados)
        st.metric(
            label="Total de Alunos Duplicados", 
            value=total_duplicados,
            help="Alunos que aparecem em mais de uma turma"
        )
    
    with col_dup2:
        duplicidade = len(df_alunos_duplicados[df_alunos_duplicados["Qtd_Turmas"] == 2])
        st.metric(
            label="Duplicidade (2 turmas)", 
            value=duplicidade,
            help="Alunos que aparecem em exatamente 2 turmas"
        )
    
    with col_dup3:
        triplicidade_mais = len(df_alunos_duplicados[df_alunos_duplicados["Qtd_Turmas"] >= 3])
        st.metric(
            label="Triplicidade+ (3+ turmas)", 
            value=triplicidade_mais,
            help="Alunos que aparecem em 3 ou mais turmas"
        )
    
    # Botão de exportação
    col_export_dup1, col_export_dup2 = st.columns([1, 4])
    with col_export_dup1:
        if st.button("📊 Exportar Duplicados", key="export_duplicados", help="Baixar planilha com alunos em múltiplas turmas"):
            # Criar formato com colunas separadas para cada turma
            export_data = []
            for _, row in df_alunos_duplicados.iterrows():
                aluno = row[coluna_aluno]
                qtd_turmas = row["Qtd_Turmas"]
                turmas_aluno = df_filt[df_filt[coluna_aluno] == aluno]["Turma"].unique().tolist()
                turmas_aluno = sorted(turmas_aluno)
                
                # Criar linha com colunas separadas
                linha = {
                    coluna_aluno: aluno,
                    "Qtd_Turmas": qtd_turmas
                }
                
                # Adicionar cada turma em uma coluna separada
                for i, turma in enumerate(turmas_aluno, 1):
                    linha[f"Turma_{i}"] = turma
                
                # Preencher colunas vazias com None para alunos com menos turmas
                max_turmas = df_alunos_duplicados["Qtd_Turmas"].max()
                for i in range(len(turmas_aluno) + 1, max_turmas + 1):
                    linha[f"Turma_{i}"] = None
                
                export_data.append(linha)
            
            df_export = pd.DataFrame(export_data)
            excel_data = criar_excel_formatado(df_export, "Alunos_Duplicados")
            st.download_button(
                label="Baixar Excel",
                data=excel_data,
                file_name="alunos_duplicados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Legenda
    st.markdown("### Legenda de Cores")
    col_leg_dup1, col_leg_dup2, col_leg_dup3 = st.columns(3)
    with col_leg_dup1:
        st.markdown("""
        **2 turmas**: Duplicidade (amarelo)  
        **3 turmas**: Triplicidade (laranja)
        """)
    with col_leg_dup2:
        st.markdown("""
        **4+ turmas**: Múltiplas turmas (vermelho)  
        **Ação**: Verificar dados
        """)
    with col_leg_dup3:
        st.markdown("""
        **Possíveis causas**:  
        • Erro de digitação  
        • Transferência não registrada
        """)
    
    # Aviso importante
    st.warning("""
    ⚠️ **Atenção**: Alunos em múltiplas turmas podem indicar:
    - Erros de digitação nos dados
    - Transferências não registradas adequadamente
    - Inconsistências na base de dados
    
    Recomenda-se verificar e corrigir essas situações.
    """)
    
else:
    st.success("✅ **Excelente!** Não foram encontrados alunos em múltiplas turmas. Os dados estão consistentes.")
    
    # Mostrar estatística geral
    col_stats1, col_stats2 = st.columns(2)
    with col_stats1:
        total_alunos_unicos = df_filt[coluna_aluno].nunique()
        st.metric("Total de Alunos Únicos", total_alunos_unicos, help="Número total de alunos únicos nos dados filtrados")
    
    with col_stats2:
        total_turmas = df_filt["Turma"].nunique()
        st.metric("Total de Turmas", total_turmas, help="Número total de turmas nos dados filtrados")

# Assinatura discreta do criador
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center; margin-top: 40px; padding: 20px;">
        <p style="margin: 0;">
            Desenvolvido por <strong style="color: #1e40af;">Alexandre Tolentino</strong> • 
            <em>Painel SGE - Sistema de Gestão Escolar</em>
        </p>
    </div>
    """, 
    unsafe_allow_html=True
)
